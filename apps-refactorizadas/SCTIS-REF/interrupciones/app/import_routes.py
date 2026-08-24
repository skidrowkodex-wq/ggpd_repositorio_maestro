from flask import Blueprint, render_template, request, jsonify, session, g, has_request_context
from app.db import query, query_one, commit_db
from app.auth import login_required
import openpyxl
import os
import uuid
import json
import difflib
from datetime import datetime, date, time, timedelta
import re


def normalizar_nombre(nombre):
    if not nombre:
        return nombre
    nombre = re.sub(r'\s+', ' ', nombre.strip())
    return ' '.join(w.capitalize() for w in nombre.split())


def normalizar_para_matching(nombre):
    if not nombre:
        return ''
    n = nombre.strip().upper()
    n = re.sub(r'\s*\d+[\.,]?\d*(?:/[\d]+[\.,]?\d*)*\s*[Kk][Vv]\.?', '', n)
    n = re.sub(r'\s*\d+[\.,]?\d*(?:/[\d]+[\.,]?\d*)+\s*', ' ', n)
    n = re.sub(r'\s+\d+[\.,]\d+\s*$', '', n)
    n = re.sub(r'\s+(\d{2,})\s*$', '', n)
    n = re.sub(r'\s*\([^)]*\)\s*', ' ', n)
    n = re.sub(r'\s*-\s*', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def asegurar_despachador(nombre_completo, estado_codigo):
    if not nombre_completo or not nombre_completo.strip():
        return None
    nombre = normalizar_nombre(nombre_completo)
    existente = query_one(
        "SELECT despachador_id FROM sctis.despachador WHERE LOWER(nombre_completo) = LOWER(%s)",
        (nombre,)
    )
    if existente:
        return existente['despachador_id']
    result = query(
        "INSERT INTO sctis.despachador (nombre_completo, estado_codigo, activo) VALUES (%s, %s, true) RETURNING despachador_id",
        (nombre, estado_codigo)
    )
    return result[0]['despachador_id']

def cargar_aliases_assets(estado_codigo):
    if not estado_codigo:
        return {'subestaciones': {}, 'circuitos': {}}
    rows = query("""
        SELECT alias_nombre, se_referencia, asset_type, asset_id
        FROM sctis.asset_alias
        WHERE estado_codigo = %s
    """, (estado_codigo,))
    se_map = {}
    ci_map = {}
    for r in (rows or []):
        if r['asset_type'] == 'SUBSTATION':
            se_map[normalizar_para_matching(r['alias_nombre'])] = r['asset_id']
        else:
            ci_map[(normalizar_para_matching(r['alias_nombre']),
                    normalizar_para_matching(r['se_referencia']))] = r['asset_id']
    return {'subestaciones': se_map, 'circuitos': ci_map}


def guardar_aliases_assets(se_map, ci_map, estado_codigo, usuario=None):
    guardados = 0
    for se_orig, asset_id in (se_map or {}).items():
        if not asset_id or not str(se_orig).strip():
            continue
        query("""
            INSERT INTO sctis.asset_alias
                (estado_codigo, asset_type, alias_nombre, se_referencia, asset_id, usuario)
            VALUES (%s, 'SUBSTATION', %s, '', %s, %s)
            ON CONFLICT (estado_codigo, asset_type, alias_nombre, se_referencia)
            DO UPDATE SET asset_id = EXCLUDED.asset_id, usuario = EXCLUDED.usuario,
                          updated_at = now()
        """, (estado_codigo, str(se_orig).strip(), asset_id, usuario))
        guardados += 1
    for ci_key, asset_id in (ci_map or {}).items():
        if not asset_id:
            continue
        ci, se = str(ci_key).split('|', 1) if '|' in str(ci_key) else (str(ci_key), '')
        if not ci.strip():
            continue
        query("""
            INSERT INTO sctis.asset_alias
                (estado_codigo, asset_type, alias_nombre, se_referencia, asset_id, usuario)
            VALUES (%s, 'CIRCUITO', %s, %s, %s, %s)
            ON CONFLICT (estado_codigo, asset_type, alias_nombre, se_referencia)
            DO UPDATE SET asset_id = EXCLUDED.asset_id, usuario = EXCLUDED.usuario,
                          updated_at = now()
        """, (estado_codigo, str(ci).strip(), str(se).strip(), asset_id, usuario))
        guardados += 1
    return guardados


def preclasificar_activo(nombre, catalogo):
    if not nombre or not catalogo:
        return 'PROBABLE_NUEVO', None
    norm = normalizar_para_matching(nombre)
    if not norm:
        return 'PROBABLE_NUEVO', None
    mejor_ratio = 0.0
    mejor_id = None
    for a in catalogo:
        cand = normalizar_para_matching(a.get('asset_name', ''))
        if not cand:
            continue
        ratio = difflib.SequenceMatcher(None, norm, cand).ratio()
        if ratio > mejor_ratio:
            mejor_ratio = ratio
            mejor_id = a['asset_id']
    if mejor_ratio >= 0.85:
        return 'PROBABLE_TYPEO', mejor_id
    if mejor_ratio >= 0.60:
        return 'PROBABLE_ALIAS', mejor_id
    return 'PROBABLE_NUEVO', None


def es_registro_incompleto(record):
    if not record:
        return True
    f_fin = record.get('fecha_fin')
    h_fin = record.get('hora_fin')
    f_ini = record.get('fecha_inicio')
    h_ini = record.get('hora_inicio')
    hrs = record.get('horas')
    dur = record.get('duracion_str') or record.get('duracion')

    # Si no tiene fecha_fin ni hora_fin
    if not f_fin and not h_fin:
        try:
            h_val = float(hrs) if hrs is not None else 0
            if h_val <= 0:
                return True
        except (ValueError, TypeError):
            return True

    # Verificar valores string nulos
    if f_fin and str(f_fin).strip().lower() in ('', 'none', 'null', 'nan', 'nonetnone'):
        return True

    # Si fecha_inicio y fecha_fin son exactamente idénticas hasta el segundo y no hay horas registradas
    dt_i = parse_flexible_datetime(f_ini)
    dt_f = parse_flexible_datetime(f_fin)
    if dt_i and dt_f:
        if (dt_f - dt_i).total_seconds() == 0:
            try:
                h_val = float(hrs) if hrs is not None else 0
                if h_val <= 0:
                    return True
            except (ValueError, TypeError):
                return True
    elif not dt_f:
        try:
            h_val = float(hrs) if hrs is not None else 0
            if h_val <= 0:
                return True
        except (ValueError, TypeError):
            return True

    return False


def insert_audit_carga_excepcional(user_id, username, estado_codigo, filename, token, total_registros, registros_incompletos, declaracion, ip_address, user_agent):
    conn = get_db()
    db_type = getattr(g, 'db_type', 'postgres')
    sql = """
        INSERT INTO "sctis.audit_admin_carga_excepcional"
            (user_id, username, estado_codigo, filename, token, total_registros, registros_incompletos, declaracion_no_repudio, ip_address, user_agent)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    params = (user_id, username, estado_codigo, filename, token, total_registros, registros_incompletos, declaracion, ip_address, user_agent)
    if db_type == 'postgres':
        res = query(sql + " RETURNING audit_id", params, fetch=True)
        return res[0]['audit_id'] if res else None
    else:
        adapted_sql = adapt_sql_for_sqlite(sql)
        cur = conn.cursor()
        cur.execute(adapted_sql, params)
        conn.commit()
        return cur.lastrowid



bp = Blueprint('importar', __name__)
UPLOAD_DIR = '/tmp/sctis_imports'

MAX_ROWS = 2000

COL_MAP_FORMATO = {
    'A': 'estado_nombre',
    'B': 'sistema',
    'C': 'jefatura',
    'D': 'subestacion',
    'E': 'circuito',
    'F': 'fecha_inicio',
    'G': 'hora_inicio',
    'H': 'fecha_fin',
    'I': 'hora_fin',
    'J': 'duracion_str',
    'K': 'carga',
    'L': 'frec',
    'M': 'horas',
    'N': 'tti_cto',
    'O': 'senal',
    'P': 'causa',
    'Q': 'sub_causa',
    'R': 'observacion',
    'S': 'sectores',
    'T': 'ciudad',
    'U': 'kva',
}

ESTADOS_DB = None
CAUSAS_DB = None


def _load_caches():
    global ESTADOS_DB, CAUSAS_DB
    if not ESTADOS_DB:
        ESTADOS_DB = query("SELECT state_code, state_name FROM common.states ORDER BY state_name") or []
    if not CAUSAS_DB:
        CAUSAS_DB = query("SELECT causa_id, causa_codigo, causa_nombre FROM sctis.causa ORDER BY causa_nombre") or []
    return ESTADOS_DB, CAUSAS_DB


def procesar_causas_preview(causas_orig_list, causas_db):
    procesadas = []
    auto_count = 0
    pendientes_count = 0

    for item in (causas_orig_list or []):
        orig = (item.get('original') or '').strip()
        count = item.get('count', 0)
        sub_c = item.get('sub_causa') or ''
        norm_orig = normalizar_texto(orig)

        matched_id = None
        matched_nombre = None
        status = 'PENDIENTE_SCTIS'
        sugerencia_id = None
        sugerencia_nombre = None
        best_ratio = 0.0

        for c in (causas_db or []):
            norm_c = normalizar_texto(c['causa_nombre'])
            norm_code = normalizar_texto(c.get('causa_codigo', ''))

            if norm_orig == norm_c or norm_orig == norm_code:
                matched_id = c['causa_id']
                matched_nombre = c['causa_nombre']
                status = 'HOMOLOGADO'
                best_ratio = 1.0
                break

            ratio = difflib.SequenceMatcher(None, norm_orig, norm_c).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                sugerencia_id = c['causa_id']
                sugerencia_nombre = c['causa_nombre']

        if not matched_id and best_ratio >= 0.82:
            matched_id = sugerencia_id
            matched_nombre = sugerencia_nombre
            status = 'HOMOLOGADO'

        if matched_id:
            auto_count += 1
        else:
            pendientes_count += 1

        procesadas.append({
            'original': orig,
            'count': count,
            'sub_causa': sub_c,
            'status': status,
            'matched_id': matched_id,
            'matched_nombre': matched_nombre,
            'sugerencia_id': sugerencia_id,
            'sugerencia_nombre': sugerencia_nombre,
            'confianza': round(best_ratio * 100)
        })

    return procesadas, auto_count, pendientes_count


def detectar_activos_inconsistentes(records, estado_codigo):
    if not estado_codigo or not records:
        return {'subestaciones': [], 'circuitos': []}

    subestaciones_unicas = {}
    circuitos_unicos = {}
    for rec in records:
        se = (rec.get('subestacion') or '').strip()
        ci = (rec.get('circuito') or '').strip()
        if se:
            if se not in subestaciones_unicas:
                subestaciones_unicas[se] = {'count': 0, 'circuitos': set()}
            subestaciones_unicas[se]['count'] += 1
            if ci:
                subestaciones_unicas[se]['circuitos'].add(ci)
        if ci and se:
            key = f'{ci}|{se}'
            if key not in circuitos_unicos:
                circuitos_unicos[key] = {'circuito': ci, 'subestacion': se, 'count': 0}
            circuitos_unicos[key]['count'] += 1

    # Aliases aprendidos del estado (resolución automática de selecciones previas)
    aliases = cargar_aliases_assets(estado_codigo)
    aliases_se = aliases['subestaciones']
    aliases_ci = aliases['circuitos']

    # Catalogo de subestaciones del estado
    cat_se = query("""
        SELECT asset_id, 
               COALESCE(asset_name_normalizado, asset_name) AS asset_name,
               asset_name AS asset_name_original,
               asset_code
        FROM common.assets
        WHERE asset_type = 'SUBSTATION' AND state_code = %s AND is_active = true
        ORDER BY asset_name
    """, (estado_codigo,))

    inconsistentes_se = []
    for nombre_se, info in subestaciones_unicas.items():
        match = None
        norm_se = normalizar_para_matching(nombre_se)
        if norm_se in aliases_se:
            continue
        for a in cat_se:
            norm_a = normalizar_para_matching(a['asset_name'])
            norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
            norm_code = normalizar_para_matching(a.get('asset_code', ''))
            if norm_a == norm_se or norm_orig == norm_se or norm_code == norm_se:
                match = a
                break
        if not match:
            for a in cat_se:
                norm_a = normalizar_para_matching(a['asset_name'])
                if norm_se in norm_a or norm_a in norm_se:
                    match = a
                    break
        if not match:
            clasif, sug_id = preclasificar_activo(nombre_se, cat_se)
            inconsistentes_se.append({
                'original': nombre_se,
                'rows': info['count'],
                'sugerencias': [{'asset_id': a['asset_id'], 'asset_name': a['asset_name']} for a in cat_se],
                'mapped_to': None,
                'clasificacion': clasif,
                'sugerencia_alias': sug_id,
            })

    inconsistentes_ci = []
    if inconsistentes_se:
        cat_ci = query("""
            SELECT a.asset_id, 
                   COALESCE(a.asset_name_normalizado, a.asset_name) AS asset_name,
                   a.asset_name AS asset_name_original,
                   a.elemento_tipo,
                   a.elemento_codigo,
                   a.parent_asset_id
            FROM common.assets a
            JOIN common.assets p ON p.asset_id = a.parent_asset_id
            WHERE a.asset_type = 'CIRCUITO' AND p.state_code = %s AND a.is_active = true
            ORDER BY a.asset_name
        """, (estado_codigo,))
    else:
        cat_ci = []

    for key, info in circuitos_unicos.items():
        se_original = info['subestacion']
        norm_se = normalizar_para_matching(se_original)
        se_match = None
        for a in cat_se:
            norm_a = normalizar_para_matching(a['asset_name'])
            norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
            if norm_a == norm_se or norm_orig == norm_se:
                se_match = a
                break
        ci_nombre = info['circuito']
        norm_ci = normalizar_para_matching(ci_nombre)
        if (norm_ci, normalizar_para_matching(se_original)) in aliases_ci:
            continue
        match = None
        if se_match:
            for a in cat_ci:
                if a['parent_asset_id'] == se_match['asset_id']:
                    norm_a = normalizar_para_matching(a['asset_name'])
                    norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
                    if norm_a == norm_ci or norm_orig == norm_ci or norm_ci in norm_a or norm_a in norm_ci:
                        match = a
                        break
        if not match:
            parent_id = se_match['asset_id'] if se_match else None
            sugerencias_ci = [a for a in cat_ci if parent_id is None or a['parent_asset_id'] == parent_id] if cat_ci else []
            clasif, sug_id = preclasificar_activo(ci_nombre, sugerencias_ci if sugerencias_ci else cat_ci)
            inconsistentes_ci.append({
                'original': ci_nombre,
                'subestacion_original': se_original,
                'rows': info['count'],
                'sugerencias': [{'asset_id': a['asset_id'], 'asset_name': a['asset_name']} for a in sugerencias_ci],
                'mapped_to': None,
                'clasificacion': clasif,
                'sugerencia_alias': sug_id,
            })

    return {'subestaciones': inconsistentes_se, 'circuitos': inconsistentes_ci}


def parse_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return float(value) if not isinstance(value, bool) else value
    s = str(value).strip()
    return s if s else None


def parse_hora(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', s):
        return s
    return s


def completar_fecha_hora(fecha_str, hora_str):
    """Combina una fecha con un valor que puede ser solo hora."""
    if not fecha_str or not hora_str:
        return hora_str or None
    try:
        fecha = datetime.fromisoformat(fecha_str).date()
    except (ValueError, TypeError):
        try:
            fecha = datetime.strptime(str(fecha_str).strip(), '%d/%m/%Y').date()
        except (ValueError, TypeError):
            return hora_str

    hora = str(hora_str).strip()
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', hora):
        parts = hora.split(':')
        h, m = int(parts[0]), int(parts[1])
        s = int(parts[2]) if len(parts) > 2 else 0
        dt = datetime(fecha.year, fecha.month, fecha.day, h, m, s)
        return dt.isoformat()

    try:
        dt = datetime.fromisoformat(hora)
        if dt.year == 1900:
            dt = dt.replace(year=fecha.year, month=fecha.month, day=fecha.day)
        return dt.isoformat()
    except (ValueError, TypeError):
        return hora_str


MESES_ES = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']


def parse_time_only(val):
    """Extrae un objeto time a partir de cadenas como '15:40:00', '15:40', '03:40 p.m.', '07:15 a.m', '12:13 a.m'."""
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    s = str(val).strip()
    if not s or s.lower() in ('none', 'null', 'nan', 'nat', '-', 's/i', 'n/a'):
        return None
    is_pm = bool(re.search(r'\b(p\.?\s*m\.?|pm)\b', s, re.IGNORECASE) or 'p.m' in s.lower() or 'pm' in s.lower())
    is_am = bool(re.search(r'\b(a\.?\s*m\.?|am)\b', s, re.IGNORECASE) or 'a.m' in s.lower() or 'am' in s.lower())
    s_clean = re.sub(r'(p\.?\s*m\.?|a\.?\s*m\.?|pm|am)', '', s, flags=re.IGNORECASE).strip(' .')
    m = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', s_clean)
    if m:
        hr, mn, sc = m.groups()
        hr = int(hr)
        mn = int(mn)
        sc = int(sc) if sc is not None else 0
        if is_pm:
            if 1 <= hr <= 11:
                hr += 12
        elif is_am:
            if hr == 12:
                hr = 0
        if 0 <= hr <= 23 and 0 <= mn <= 59 and 0 <= sc <= 59:
            return time(hr, mn, sc)
    return None


def parse_flexible_datetime(val):
    """
    Parsea de manera ultra robusta cualquier representación de fecha y hora,
    incluyendo formatos mixtos venezolanos como:
      - '03/01/2026 15:24 p.m' (hora 24h con sufijo redundante)
      - '31/07/2026 02:13 p.m' (hora 12h con sufijo pm -> 14:13)
      - '24/07/2026 12:13 a.m' (medianoche 12h -> 00:13)
      - '25/07/2026 07:15 a.m'
      - Objetos datetime o date de openpyxl
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())

    s = str(val).strip()
    if not s or s.lower() in ('none', 'null', 'nan', 'nat', '-', 's/i', 'n/a'):
        return None

    # Detectar presencia de indicador meridiano (am / pm)
    is_pm = bool(re.search(r'\b(p\.?\s*m\.?|pm)\b', s, re.IGNORECASE) or 'p.m' in s.lower() or 'pm' in s.lower())
    is_am = bool(re.search(r'\b(a\.?\s*m\.?|am)\b', s, re.IGNORECASE) or 'a.m' in s.lower() or 'am' in s.lower())

    # Limpiar sufijos meridianos para aislar números
    s_clean = re.sub(r'(p\.?\s*m\.?|a\.?\s*m\.?|pm|am)', '', s, flags=re.IGNORECASE).strip(' .')
    # Eliminar posibles dobles espacios o caracteres extra
    s_clean = re.sub(r'\s+', ' ', s_clean)

    # Patrón común: DD/MM/YYYY o YYYY-MM-DD con hora HH:MM(:SS) opcional
    m = re.match(r'^(\d{1,4})[/\-\.](\d{1,2})[/\-\.](\d{2,4})(?:[\sT,]+(\d{1,2}):(\d{2})(?::(\d{2}))?)?$', s_clean)
    if m:
        p1, p2, p3, hr, mn, sc = m.groups()
        has_time = (hr is not None and mn is not None)
        hr = int(hr) if hr is not None else 0
        mn = int(mn) if mn is not None else 0
        sc = int(sc) if sc is not None else 0

        # Discernir si p1 es año o día
        if len(p1) == 4:
            year, month, day = int(p1), int(p2), int(p3)
        elif len(p3) == 4:
            day, month, year = int(p1), int(p2), int(p3)
        elif len(p3) == 2:
            day, month, year = int(p1), int(p2), int('20' + p3)
        else:
            day, month, year = int(p1), int(p2), int(p3)

        # Ajuste meridiano si se capturó hora
        if has_time:
            if is_pm:
                if 1 <= hr <= 11:
                    hr += 12
            elif is_am:
                if hr == 12:
                    hr = 0

        try:
            return datetime(year, month, day, hr, mn, sc)
        except ValueError:
            pass

    # Intentos con formatos estándar de datetime
    try:
        return datetime.fromisoformat(s_clean)
    except Exception:
        pass

    for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M',
                '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s_clean, fmt)
        except Exception:
            pass

    return None


def normalizar_a_date(val):
    dt = parse_flexible_datetime(val)
    return dt.date() if dt else None


def normalizar_a_datetime(val):
    return parse_flexible_datetime(val)


def parse_cell(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return float(value) if not isinstance(value, bool) else value
    s = str(value).strip()
    return s if s else None


def normalizar_hora_24h(val):
    """Normaliza cualquier formato de hora (12h, 24h, floats, datetime) a formato estricto 24h HH:MM:SS."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%H:%M:%S')
    if isinstance(val, date) and hasattr(val, 'strftime'):
        return val.strftime('%H:%M:%S')
    if isinstance(val, (int, float)):
        if 0 <= val < 1:
            total_sec = int(val * 86400)
            h = total_sec // 3600
            m = (total_sec % 3600) // 60
            s = total_sec % 60
            return f'{h:02d}:{m:02d}:{s:02d}'

    s = str(val).strip()
    if not s or s.lower() in ['none', 'null', '00:00:00', '0']:
        return None

    # Limpiar comas o caracteres tipográficos espurios (ej. '08,:54:00,')
    s = re.sub(r'[,;]', '', s).strip()

    m = re.search(r'(\d{1,2})[:\.](\d{2})(?:[:\.](\d{2}))?\s*([ap]\.?m\.?)?', s, re.IGNORECASE)
    if m:
        h = int(m.group(1))
        minute = int(m.group(2))
        sec = int(m.group(3)) if m.group(3) else 0
        ampm = (m.group(4) or '').lower().replace('.', '').strip()

        if 'p' in ampm:
            if h < 12:
                h += 12
            # Si h >= 12 (ej. '15:24 p.m.'), se mantiene en 15 (ya en 24h)
        elif 'a' in ampm:
            if h == 12:
                h = 0

        if h >= 24:
            h = h % 24

        return f'{h:02d}:{minute:02d}:{sec:02d}'
    return s


def parse_hora(value):
    return normalizar_hora_24h(value)


def separar_fecha_hora(val):
    """Extrae y separa fecha ('YYYY-MM-DD') y hora ('HH:MM:SS' en 24h) desde cadenas mixtas."""
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d'), val.strftime('%H:%M:%S')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d'), None

    s = str(val).strip()
    if not s or s.lower() in ['none', 'null']:
        return None, None

    fecha_part = None
    hora_part = None

    # 1. Fecha ISO YYYY-MM-DD
    m_iso = re.search(r'(\d{4})[\/\-](\d{1,2})[\/\-](\d{1,2})', s)
    if m_iso:
        year, month, day = int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3))
        fecha_part = f'{year:04d}-{month:02d}-{day:02d}'
    else:
        # 2. Fecha Latina DD/MM/YYYY
        m_lat = re.search(r'(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{2,4})', s)
        if m_lat:
            p1, p2, p3 = int(m_lat.group(1)), int(m_lat.group(2)), int(m_lat.group(3))
            if p3 < 100:
                p3 += 2000
            day, month, year = p1, p2, p3
            fecha_part = f'{year:04d}-{month:02d}-{day:02d}'

    # 3. Hora normalizada 24h
    m_time = re.search(r'(\d{1,2}[:\.]\d{2}(?:[:\.]\d{2})?\s*(?:[ap]\.?m\.?)?)', s, re.IGNORECASE)
    if m_time:
        hora_part = normalizar_hora_24h(m_time.group(1))

    return fecha_part, hora_part


def completar_fecha_hora(fecha_str, hora_str):
    """Combina una fecha con un valor que puede ser solo hora o fecha+hora."""
    if not fecha_str and not hora_str:
        return None
    f_p, h_p = separar_fecha_hora(fecha_str)
    _, h_from_hora = separar_fecha_hora(hora_str)
    hora_final = h_from_hora or h_p or '00:00:00'
    if f_p:
        return f'{f_p} {hora_final}'
    return str(hora_str or fecha_str).strip()


def combinar_fecha_hora(record):
    """Separa componentes y los recombina limpiamente derivando fecha_falla."""
    raw_ini = record.get('fecha_inicio')
    raw_h_ini = record.get('hora_inicio')
    f_ini, h_ini = separar_fecha_hora(raw_ini)
    if not h_ini and raw_h_ini:
        h_ini = normalizar_hora_24h(raw_h_ini)

    raw_fin = record.get('fecha_fin')
    raw_h_fin = record.get('hora_fin')
    f_fin, h_fin = separar_fecha_hora(raw_fin)
    if not h_fin and raw_h_fin:
        h_fin = normalizar_hora_24h(raw_h_fin)

    if f_ini and h_ini:
        record['fecha_inicio'] = f'{f_ini} {h_ini}'
    elif f_ini:
        record['fecha_inicio'] = f_ini
    elif raw_ini:
        record['fecha_inicio'] = str(raw_ini)

    if f_fin and h_fin:
        record['fecha_fin'] = f'{f_fin} {h_fin}'
    elif f_fin:
        record['fecha_fin'] = f_fin
    elif raw_fin:
        record['fecha_fin'] = str(raw_fin)

    if h_ini:
        record['hora_inicio'] = h_ini
    if h_fin:
        record['hora_fin'] = h_fin

    if f_ini and not record.get('fecha_falla'):
        record['fecha_falla'] = f_ini
    elif record.get('fecha_falla'):
        f_f, _ = separar_fecha_hora(record['fecha_falla'])
        if f_f:
            record['fecha_falla'] = f_f

    # Derivar duración en horas si está vacía o en 0
    h_val = record.get('horas')
    try:
        h_float = float(h_val) if h_val is not None and str(h_val).strip() != '' else None
    except (ValueError, TypeError):
        h_float = None

    dt_ini = parse_flexible_datetime(record.get('fecha_inicio'))
    dt_fin = parse_flexible_datetime(record.get('fecha_fin'))
    if (h_float is None or h_float <= 0) and dt_ini and dt_fin:
        diff_sec = (dt_fin - dt_ini).total_seconds()
        if diff_sec >= 0:
            record['horas'] = str(round(diff_sec / 3600.0, 4))
    elif h_float is not None:
        record['horas'] = str(h_float)

    # Derivar mes
    if not record.get('mes') or str(record.get('mes')).strip() == '':
        ref = dt_ini or parse_flexible_datetime(record.get('fecha_falla'))
        if ref and 1 <= ref.month <= 12:
            record['mes'] = MESES_ES[ref.month]

    return record


def normalizar_y_extraer_registro(record):
    """Wrapper para compatibilidad con código existente."""
    return combinar_fecha_hora(record)


def detectar_estado_desde_texto(texto, estados_db):
    if not texto:
        return None, None
    t = normalizar_texto(str(texto).strip())
    for st in estados_db:
        if t == normalizar_texto(st['state_name']):
            return st['state_code'], st['state_name']
    return None, None


# ─── Parseo del FORMATO ESTABLECIDO (primera hoja) ────────────

def parse_formato(ws, estados_db, causas_db):
    records = []
    causas_originales = {}
    max_row = ws.max_row or 1

    for row_idx in range(2, min(max_row + 1, MAX_ROWS + 1)):
        cells = ws[row_idx]
        a_val = cells[0].value if len(cells) > 0 else None
        if a_val is None or str(a_val).strip() == '':
            continue

        record = {}
        for col_idx in range(22):
            col_letter = chr(ord('A') + col_idx)
            field = COL_MAP_FORMATO.get(col_letter)
            if field is None:
                continue
            val = parse_cell(cells[col_idx].value) if col_idx < len(cells) else None
            record[field] = val

        causa_orig = (record.get('causa') or '').strip()
        if causa_orig:
            key = causa_orig.upper()
            if key not in causas_originales:
                causas_originales[key] = {
                    'original': causa_orig,
                    'sub_causa': record.get('sub_causa'),
                    'count': 0,
                }
            causas_originales[key]['count'] += 1

        combinar_fecha_hora(record)

        records.append(record)

    return {
        'ok': True,
        'formato': 'formato',
        'total_rows': len(records),
        'causas_originales': list(causas_originales.values()),
        'causas_homologadas': causas_db,
        'preview': records[:20],
        '_all_records': records,
    }


# ─── Validación de estructura contra plantilla oficial ─────────

HEADERS_FORMATO_ESPERADOS = [
    (0, 'ESTADO'),
    (1, 'SISTEMA'),
    (4, 'CIRCUITO'),
    (5, 'FECHA'),
    (15, 'CAUSA'),
    (16, 'SUB'),
    (17, 'OBSERVACION'),
]


def normalizar_texto(s):
    """Elimina acentos y diéresis para comparación de forma segura."""
    if s is None:
        return ''
    import unicodedata
    s_str = str(s).strip()
    if not s_str:
        return ''
    return ''.join(
        c for c in unicodedata.normalize('NFKD', s_str) if not unicodedata.combining(c)
    ).upper().strip()


def validar_estructura(ws):
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not headers or not headers[0]:
        return False, 'El archivo no tiene fila de encabezados.'

    row = headers[0]
    total_cols = ws.max_column or sum(1 for c in row if c is not None and str(c).strip())
    if total_cols < 19:
        return False, (
            'El archivo no coincide con la plantilla oficial. '
            'Debe usar el formulario de carga Excel homologado proporcionado por la Coordinación de SCTIS. '
            f'Se detectaron {total_cols} columnas (se esperan al menos 19).'
        )

    row_strs = [normalizar_texto(str(c)) if c else '' for c in row[:30]]

    if 'ESTADO' not in row_strs[0]:
        return False, (
            'Formato de archivo no reconocido. '
            'Debe usar el formulario de carga Excel homologado '
            '(primera pestaña del archivo proporcionado por la Coordinación de SCTIS). '
            f'Encabezado detectado: "{row_strs[0][:50] if row_strs[0] else "(vacío)"}".'
        )

    problemas = []
    for col_idx, kw in HEADERS_FORMATO_ESPERADOS:
        actual = row_strs[col_idx] if col_idx < len(row_strs) else ''
        if kw not in actual and actual:
            problemas.append(f'Columna {chr(65+col_idx)}: se esperaba contener "{kw}", se encontró "{actual}"')
    if problemas:
        return False, (
            'La estructura del archivo no coincide con la plantilla oficial del Formato Establecido. '
            'Debe usar el formulario de carga Excel homologado proporcionado por la Coordinación de SCTIS. '
            'Detalles: ' + '; '.join(problemas[:3])
        )

    return True, ''


def detectar_encabezados_y_filtro(ws, max_scan_rows=25):
    """Escanea las primeras filas, omite títulos y ubica la fila exacta de encabezados."""
    HEADER_KEYWORDS = {
        'ESTADO', 'SISTEMA', 'SUBESTACION', 'SUB ESTACION', 'SUB-ESTACION', 'SUB-ESTACIÓN', 'S/E', 'CIRCUITO', 'CIRCUITOS',
        'FECHA', 'FECHA INI', 'FECHA FIN', 'HORA', 'HORA INICIO', 'INICIO', 'FIN', 'DURACION', 'DURACIÓN', 'CARGA',
        'CAUSA', 'SUB CAUSA', 'SUB-CAUSA', '(SUB-CAUSA)', 'OBSERVACION', 'OBSERVACIONES', 'KVA', 'HORAS', 'DISTRITO',
        'JEFATURA', 'SECTORES', 'CIUDAD', 'SEÑAL', 'TTI'
    }
    best_row = 1
    best_score = 0
    headers_detected = []

    scan_limit = min(max_scan_rows, (ws.max_row or 1))
    for r in range(1, scan_limit + 1):
        row_cells = ws[r]
        row_vals = [normalizar_texto(c.value) for c in row_cells[:min(len(row_cells), 50)]]
        score = sum(1 for v in row_vals if any(kw in v for kw in HEADER_KEYWORDS))
        if score > best_score:
            best_score = score
            best_row = r
            headers_detected = row_vals

    has_headers = (best_score >= 2)
    start_data_row = (best_row + 1) if has_headers else 1
    return best_row, headers_detected, start_data_row, has_headers


FORMATOS_ESTANDAR = {
    'AMAZONAS': {
        'formato_codigo': 'AMAZONAS',
        'formato_nombre': 'Formato Amazonas (Distribución)',
        'estado_codigo': 'AMA',
        'header_keywords': ['FECHA FALLA', 'SISTEMA', 'S/E', 'CIRCUITO', 'FECHA INI', 'FECHA FIN', 'CAUSA', 'SUB CAUSA'],
        'mapeo_columnas': {
            'FECHA FALLA': {'target': 'fecha_falla', 'tipo': 'fecha'},
            'SISTEMA': {'target': 'sistema', 'tipo': 'texto'},
            'S/E': {'target': 'subestacion', 'tipo': 'texto'},
            'SUBESTACION': {'target': 'subestacion', 'tipo': 'texto'},
            'CIRCUITO': {'target': 'circuito', 'tipo': 'texto'},
            'DISTRITO': {'target': 'estado_nombre', 'tipo': 'texto'},
            'JEFATURA': {'target': 'jefatura', 'tipo': 'texto'},
            'FECHA INI': {'target': 'fecha_inicio', 'tipo': 'fecha_hora'},
            'FECHA FIN': {'target': 'fecha_fin', 'tipo': 'fecha_hora'},
            'CAUSA': {'target': 'causa', 'tipo': 'texto'},
            'SUB CAUSA': {'target': 'sub_causa', 'tipo': 'texto'},
            'OBSERVACION': {'target': 'observacion', 'tipo': 'texto'},
            'USUARIO': {'target': 'despachador', 'tipo': 'texto'},
            'KVA': {'target': 'kva', 'tipo': 'numerico'},
            'T.HORAS': {'target': 'horas', 'tipo': 'numerico'},
            'MES': {'target': 'mes', 'tipo': 'texto'},
            'SECTORES': {'target': 'sectores', 'tipo': 'texto'},
            'CIUDAD': {'target': 'ciudad', 'tipo': 'texto'},
            'HORA INICIO': {'target': 'hora_inicio', 'tipo': 'hora'},
        },
        'reglas': {'fecha_falla': 'derivar_de_fecha_inicio'}
    },
    'MONAGAS': {
        'formato_codigo': 'MONAGAS',
        'formato_nombre': 'Formato Monagas (Crystal Reports)',
        'estado_codigo': 'MON',
        'header_keywords': ['SUB-ESTACION', 'SUB-ESTACIÓN', 'CIRCUITO', 'INICIO', 'DURACION', 'CARGA', 'CAUSA'],
        'mapeo_columnas': {
            'FECHA': {'target': 'fecha_inicio', 'tipo': 'fecha'},
            'SISTEMA': {'target': 'sistema', 'tipo': 'texto'},
            'DISTRITO': {'target': 'jefatura', 'tipo': 'texto'},
            'SUB-ESTACION': {'target': 'subestacion', 'tipo': 'texto'},
            'SUB-ESTACIÓN': {'target': 'subestacion', 'tipo': 'texto'},
            'SUBESTACION': {'target': 'subestacion', 'tipo': 'texto'},
            'CIRCUITO': {'target': 'circuito', 'tipo': 'texto'},
            'INICIO': {'target': 'hora_inicio', 'tipo': 'hora'},
            'DURACION': {'target': 'duracion', 'tipo': 'duracion_a_horas'},
            'DURACIÓN': {'target': 'duracion', 'tipo': 'duracion_a_horas'},
            'CARGA': {'target': 'kva', 'tipo': 'numerico'},
            'CAUSA': {'target': 'causa', 'tipo': 'texto'},
            '(SUB-CAUSA)': {'target': 'sub_causa', 'tipo': 'texto'},
            'SUB-CAUSA': {'target': 'sub_causa', 'tipo': 'texto'},
            'SUB CAUSA': {'target': 'sub_causa', 'tipo': 'texto'},
            'OBSERVACIONES': {'target': 'observacion', 'tipo': 'texto'},
            'OBSERVACION': {'target': 'observacion', 'tipo': 'texto'},
        },
        'reglas': {'fecha_falla': 'derivar_de_fecha_inicio'}
    },
    'TIRAS': {
        'formato_codigo': 'TIRAS',
        'formato_nombre': 'Formato TIRAS Estándar',
        'estado_codigo': None,
        'header_keywords': ['FECHA', 'SISTEMA', 'S/E', 'CIRCUITO', 'CAUSA', 'SUB CAUSA', 'OBSERVACION', 'KVA'],
        'mapeo_columnas': {
            'S/E': {'target': 'subestacion', 'tipo': 'texto'},
            'SUBESTACION': {'target': 'subestacion', 'tipo': 'texto'},
            'CIRCUITO': {'target': 'circuito', 'tipo': 'texto'},
            'FECHA': {'target': 'fecha_inicio', 'tipo': 'fecha'},
            'FECHA INI': {'target': 'fecha_inicio', 'tipo': 'fecha_hora'},
            'FECHA FIN': {'target': 'fecha_fin', 'tipo': 'fecha_hora'},
            'HORA INICIO': {'target': 'hora_inicio', 'tipo': 'hora'},
            'HORA FIN': {'target': 'hora_fin', 'tipo': 'hora'},
            'SISTEMA': {'target': 'sistema', 'tipo': 'texto'},
            'CAUSA': {'target': 'causa', 'tipo': 'texto'},
            'SUB CAUSA': {'target': 'sub_causa', 'tipo': 'texto'},
            'OBSERVACION': {'target': 'observacion', 'tipo': 'texto'},
            'KVA': {'target': 'kva', 'tipo': 'numerico'},
            'T.HORAS': {'target': 'horas', 'tipo': 'numerico'}
        },
        'reglas': {'fecha_falla': 'derivar_de_fecha_inicio'}
    }
}


def detectar_formato_catalogo(ws, estado_codigo=None):
    """
    Detecta el formato del Excel comparando headers con formatos estándar o catálogo DB.
    Retorna (formato_dict, None) si match, o (None, error_msg) si no.
    """
    _, header_vals, _, has_headers = detectar_encabezados_y_filtro(ws)
    headers_str = ' '.join([normalizar_texto(h) for h in header_vals if h])

    # 1. Chequeo de FORMATOS_ESTANDAR en memoria
    for fmt_key, fmt_def in FORMATOS_ESTANDAR.items():
        kws = fmt_def.get('header_keywords', [])
        if estado_codigo and fmt_def.get('estado_codigo') and fmt_def.get('estado_codigo') != estado_codigo:
            continue
        match_count = sum(1 for kw in kws if normalizar_texto(kw) in headers_str)
        if len(kws) > 0 and (match_count / len(kws) >= 0.5 or match_count >= 4):
            return fmt_def, None

    # 2. Chequeo en base de datos
    where = "activo = true"
    params = []
    if estado_codigo:
        where += " AND (estado_codigo = %s OR estado_codigo IS NULL)"
        params.append(estado_codigo)
    else:
        where += " AND estado_codigo IS NULL"

    formatos = query(f"""
        SELECT formato_id, formato_codigo, formato_nombre, estado_codigo, header_keywords,
               header_row, data_start_row, mapeo_columnas, reglas, campos_faltantes
        FROM sctis.formato_catalogo
        WHERE {where}
        ORDER BY
            CASE WHEN estado_codigo IS NOT DISTINCT FROM %s THEN 0 ELSE 1 END,
            array_length(header_keywords, 1) DESC
    """, params + [estado_codigo])

    for fmt in (formatos or []):
        keywords = fmt.get('header_keywords')
        if isinstance(keywords, str):
            try:
                keywords = json.loads(keywords)
            except Exception:
                if keywords.startswith('{') and keywords.endswith('}'):
                    keywords = [k.strip().strip('"') for k in keywords[1:-1].split(',') if k.strip()]
                else:
                    keywords = [k.strip() for k in keywords.split(',') if k.strip()]
        if not isinstance(keywords, list):
            keywords = []
        fmt['header_keywords'] = keywords

        match_count = sum(1 for kw in keywords if kw and normalizar_texto(kw) in headers_str)
        match_ratio = match_count / len(keywords) if keywords else 0
        if match_ratio >= 0.5 or match_count >= 4:
            if isinstance(fmt.get('mapeo_columnas'), str):
                try:
                    fmt['mapeo_columnas'] = json.loads(fmt['mapeo_columnas'])
                except Exception:
                    fmt['mapeo_columnas'] = {}
            if isinstance(fmt.get('reglas'), str):
                try:
                    fmt['reglas'] = json.loads(fmt['reglas'])
                except Exception:
                    fmt['reglas'] = {}
            return fmt, None

    return None, (
        'Formato no reconocido. El sistema detectó que el archivo no coincide '
        'con ningún formato predefinido del catálogo ni con el formato CTIS homologado. '
        'Verifique que el archivo esté en el formato correcto de su estado.'
    )


def parse_formato_catalogo(ws, formato, estado_codigo):
    mapeo = formato.get('mapeo_columnas', {})
    if isinstance(mapeo, str):
        try:
            mapeo = json.loads(mapeo)
        except Exception:
            mapeo = {}
    reglas = formato.get('reglas', {})
    if isinstance(reglas, str):
        try:
            reglas = json.loads(reglas)
        except Exception:
            reglas = {}

    header_row_idx, header_vals, data_start, _ = detectar_encabezados_y_filtro(ws)
    max_row = ws.max_row or data_start

    col_mapping_by_idx = {}

    # Pasada 1: Coincidencias exactas
    for col_idx, h_val in enumerate(header_vals):
        norm_h = normalizar_texto(h_val)
        if not norm_h:
            continue
        for header_key, map_info in mapeo.items():
            if normalizar_texto(header_key) == norm_h:
                if map_info and map_info.get('tipo') != 'ignore' and map_info.get('target'):
                    col_mapping_by_idx[col_idx] = map_info
                break

    # Pasada 2: Coincidencias difusas para targets aún no asignados
    used_targets = {m['target'] for m in col_mapping_by_idx.values()}
    for col_idx, h_val in enumerate(header_vals):
        if col_idx in col_mapping_by_idx:
            continue
        norm_h = normalizar_texto(h_val)
        if not norm_h or len(norm_h) < 3:
            continue
        for header_key, map_info in mapeo.items():
            norm_k = normalizar_texto(header_key)
            if len(norm_k) >= 3 and (norm_k in norm_h or norm_h in norm_k):
                if map_info and map_info.get('tipo') != 'ignore' and map_info.get('target'):
                    if map_info['target'] not in used_targets:
                        col_mapping_by_idx[col_idx] = map_info
                        used_targets.add(map_info['target'])
                break

    records = []
    causas_originales = {}

    for row_idx in range(data_start, min(max_row + 1, 2500 + data_start)):
        cells = ws[row_idx]
        if not cells or len(cells) == 0:
            continue
        if all(c.value is None or str(c.value).strip() == '' for c in cells[:min(len(cells), 8)]):
            continue

        record = {}
        for col_idx, mapping in col_mapping_by_idx.items():
            if col_idx >= len(cells):
                continue
            cell_val = cells[col_idx].value
            if cell_val is None:
                continue

            target = mapping['target']
            tipo = mapping.get('tipo', 'texto')

            if tipo == 'texto':
                record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'numerico':
                try:
                    record[target] = float(cell_val) if cell_val is not None else None
                except (ValueError, TypeError):
                    record[target] = None
            elif tipo == 'fecha':
                f_p, _ = separar_fecha_hora(cell_val)
                record[target] = f_p
            elif tipo == 'hora':
                record[target] = normalizar_hora_24h(cell_val)
            elif tipo == 'fecha_hora':
                f_p, h_p = separar_fecha_hora(cell_val)
                if f_p and h_p:
                    record[target] = f'{f_p} {h_p}'
                elif f_p:
                    record[target] = f_p
                else:
                    record[target] = str(cell_val).strip()
            elif tipo == 'duracion_a_horas':
                record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'minutos_a_horas':
                try:
                    mins = float(cell_val)
                    record[target] = str(mins / 60) if mins else None
                except (ValueError, TypeError):
                    record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'mw_a_kva':
                try:
                    mw = float(cell_val)
                    record[target] = str(mw * 1000) if mw else None
                except (ValueError, TypeError):
                    record[target] = None
            elif tipo == 'mvamin_a_kva':
                record[target] = str(cell_val).strip() if cell_val else None
            else:
                record[target] = str(cell_val).strip() if cell_val else None

        if not record.get('subestacion') and not record.get('circuito'):
            continue

        for key, val in record.items():
            if val == '' or val == 'None':
                record[key] = None

        combinar_fecha_hora(record)

        causa_orig = (record.get('causa') or '').strip()
        if causa_orig:
            key = causa_orig.upper()
            if key not in causas_originales:
                causas_originales[key] = {
                    'original': causa_orig,
                    'sub_causa': record.get('sub_causa'),
                    'count': 0,
                }
            causas_originales[key]['count'] += 1

        records.append(record)

    return {
        'ok': True,
        'formato': formato.get('formato_codigo', 'CATALOGO'),
        'formato_nombre': formato.get('formato_nombre', 'Formato Adaptativo'),
        'total_rows': len(records),
        'causas_originales': list(causas_originales.values()),
        'campos_faltantes': formato.get('campos_faltantes', []),
        'preview': records[:20],
        '_all_records': records,
    }


def generar_excel_correccion(records_errores, token, estado_codigo):
    """
    Genera un Excel con los registros rechazados para que el usuario los corrija.
    Retorna la ruta del archivo generado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros para Corrección'

    headers = ['Fila Original', 'Subestación', 'Circuito', 'Fecha Inicio',
               'Causa', 'Error', 'Observación']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for rec in records_errores:
        ws.append([
            rec.get('row_numero', ''),
            rec.get('subestacion', ''),
            rec.get('circuito', ''),
            rec.get('fecha_inicio', ''),
            rec.get('causa', ''),
            rec.get('error', ''),
            rec.get('observacion', ''),
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    filename = f'correccion_{estado_codigo}_{token[:8]}.xlsx'
    filepath = os.path.join(UPLOAD_DIR, filename)
    wb.save(filepath)
    return filepath


def sheet_oculta(filepath, sheet_name):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        oculta = wb[sheet_name].sheet_state != 'visible'
        wb.close()
        return oculta
    except Exception:
        return False


def listar_hojas_visibles(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        visibles = [s for s in wb.sheetnames if wb[s].sheet_state == 'visible']
        wb.close()
        return visibles
    except Exception:
        return []


class _Celda:
    __slots__ = ('value',)
    def __init__(self, value):
        self.value = value


class _Fila:
    __slots__ = ('_celdas',)
    def __init__(self, celdas):
        self._celdas = celdas

    def __getitem__(self, idx):
        if isinstance(idx, slice):
            return self._celdas[idx]
        if isinstance(idx, int) and 0 <= idx < len(self._celdas):
            return self._celdas[idx]
        return _Celda(None)

    def __len__(self):
        return len(self._celdas)

    def __iter__(self):
        return iter(self._celdas)


class HojaMemoria:
    """Wrapper de memoria acotado a 60 columnas y 2500 filas max con descarte de filas vacías."""
    def __init__(self, ws):
        self.title = getattr(ws, 'title', 'Sheet')
        self._filas = []
        self.max_row = 0
        self.max_column = 0
        consecutive_empty = 0

        rows_gen = ws.iter_rows(values_only=True) if hasattr(ws, 'iter_rows') else []
        for row in rows_gen:
            if len(self._filas) >= 2500:
                break

            # Limitar a 60 columnas útiles
            sliced = row[:60] if row else []
            is_empty_row = not sliced or all(v is None or str(v).strip() == '' for v in sliced)
            if is_empty_row:
                consecutive_empty += 1
                if len(self._filas) > 0 and consecutive_empty >= 10:
                    break
                elif len(self._filas) == 0 and consecutive_empty >= 50:
                    break
                continue
            else:
                consecutive_empty = 0

            celdas = [_Celda(v) for v in sliced]
            self._filas.append(_Fila(celdas))
            self.max_column = max(self.max_column, len(celdas))

        self.max_row = len(self._filas)

    def __getitem__(self, idx):
        if isinstance(idx, int) and 1 <= idx <= len(self._filas):
            return self._filas[idx - 1]
        elif isinstance(idx, int) and 0 <= idx < len(self._filas):
            return self._filas[idx]
        return _Fila([])

    def iter_rows(self, min_row=1, max_row=None, values_only=False):
        max_row = max_row or self.max_row
        for i in range(min_row, max_row + 1):
            fila = self[i]
            if values_only:
                yield tuple(c.value for c in fila)
            else:
                yield fila


def cargar_hoja_memoria(filepath, sheet_name=None):
    # 0. Si es archivo .xls binario, leer directamente con xlrd en memoria en <0.1s
    if filepath.lower().endswith(('.xls', '.xlt')):
        try:
            import xlrd
            xwb = xlrd.open_workbook(filepath)
            s_name = sheet_name if (sheet_name and sheet_name in xwb.sheet_names()) else xwb.sheet_names()[0]
            xws = xwb.sheet_by_name(s_name)
            
            class _XlrdWS:
                def __init__(self, title, xws, xwb):
                    self.title = str(title)
                    self.xws = xws
                    self.xwb = xwb
                def iter_rows(self, values_only=True):
                    for r in range(self.xws.nrows):
                        row_vals = []
                        for c in range(min(self.xws.ncols, 60)):
                            cell = self.xws.cell(r, c)
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    val = xlrd.xldate_as_datetime(cell.value, self.xwb.datemode)
                                except Exception:
                                    val = cell.value
                            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                                val = bool(cell.value)
                            elif cell.ctype == xlrd.XL_CELL_ERROR:
                                val = None
                            else:
                                val = cell.value
                            row_vals.append(val)
                        yield tuple(row_vals)
            return HojaMemoria(_XlrdWS(s_name, xws, xwb))
        except Exception:
            pass

    # 1. Intentar con openpyxl en modo read_only=True para mínimo consumo de memoria y máxima velocidad
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            hm = HojaMemoria(ws)
            if hm.max_row > 0:
                return hm
        finally:
            wb.close()
    except Exception:
        pass

    # 2. Intentar con pandas (soporta xlsx, xls, html tables y es ultra tolerante)
    try:
        import pandas as pd
        df = pd.read_excel(filepath, sheet_name=sheet_name or 0, header=None)
        if not df.empty:
            class _DummyWS:
                def __init__(self, title, rows):
                    self.title = str(title)
                    self._rows = rows
                def iter_rows(self, values_only=True):
                    for r in self._rows:
                        yield r
            
            rows_data = []
            for _, r_series in df.iloc[:2500, :60].iterrows():
                row_vals = []
                for val in r_series:
                    if pd.isna(val):
                        row_vals.append(None)
                    elif hasattr(val, 'strftime'):
                        row_vals.append(str(val))
                    else:
                        row_vals.append(val)
                rows_data.append(tuple(row_vals))
            return HojaMemoria(_DummyWS(sheet_name or 'Sheet1', rows_data))
    except Exception:
        pass

    return None


def ext_from_token(token):
    for f in os.listdir(UPLOAD_DIR):
        if f.startswith(token) and not f.endswith('_preview.json'):
            return os.path.splitext(f)[1]
    return '.xlsx'


HEADER_SYNONYMS_ORDERED = [
    ('fecha_inicio', ['FECHA INI', 'FECHA INICIO', 'FECHA DE APERTURA', 'FECHA APERTURA', 'FECHA APERT', 'F. INI', 'F. INICIO', 'FECHA DE INICIO', 'FECHA_INICIO', 'FECHA INICIAL', 'FECHA Y HORA DE INICIO', 'F_INI']),
    ('fecha_fin', ['FECHA FIN', 'FECHA FINAL', 'FECHA DE CIERRE', 'FECHA CIERRE', 'F. FIN', 'F. FINAL', 'FECHA DE FIN', 'FECHA_FIN', 'FECHA Y HORA DE FIN', 'F_FIN']),
    ('hora_inicio', ['HORA INICIO', 'HORA INI', 'HORA DE APERTURA', 'HORA APERTURA', 'H. INI', 'H. INICIO', 'HORA DE INICIO', 'HORA_INICIO', 'INICIO', 'H_INI']),
    ('hora_fin', ['HORA FIN', 'HORA FINAL', 'HORA DE CIERRE', 'HORA CIERRE', 'H. FIN', 'H. FINAL', 'HORA DE FIN', 'HORA_FIN', 'FIN', 'H_FIN']),
    ('fecha_falla', ['FECHA FALLA', 'FECHA DE FALLA', 'FECHA DE LA FALLA', 'FECHA', 'F. FALLA', 'FECHA_FALLA']),
    ('sub_causa', ['SUB CAUSA', 'SUB-CAUSA', 'SUBCAUSA', 'SUB_CAUSA', '(SUB-CAUSA)', 'SUB-CAUSAS', 'SUB CAUSAS']),
    ('causa', ['CAUSA DE LA FALLA', 'TIPO DE FALLA', 'DESCRIPCION CAUSA', 'DESCRIPCIÓN CAUSA', 'CAUSA', 'MOTIVO DEL EVENTO', 'MOTIVO', 'DESCRIPCION DE LA FALLA', 'TIPO DE EVENTO', 'CAUSA PRINCIPAL']),
    ('subestacion', ['SUB-ESTACION', 'SUB ESTACION', 'SUBESTACION', 'SUBESTACIÓN', 'S/E', 'S / E', 'SUB_ESTACION', 'SUBEST', 'S /E', 'S/ E']),
    ('circuito', ['NOMBRE DEL CIRCUITO', 'CIRCUITOS', 'CIRCUITO', 'GENERADOR_SUBESTACION CIRCUITO', 'CTO', 'ALIMENTADOR', 'CIRCUITO / RAMAL', 'BARRA / CIRCUITO', 'CIRCUITO/BARRA', 'CTO.']),
    ('horas', ['T.HORAS', 'HORAS DE DURACION', 'HORAS DE DURACIÓN', 'DURACIÓN T-HORAS', 'DURACION T-HORAS', 'TIEMPO INTERRUPCION', 'TIEMPO INTERRUPCIÓN', 'DURACION', 'DURACIÓN', 'HORAS', 'T.REP.', 'T.TOTAL', 'DURACION EN HORAS']),
    ('duracion_str', ['T. hh:mm', 'T.HH:MM', 'TIEMPO', 'DURACION EN MIN', 'DURACION_STR', 'T. hh:mm: ss', 'T. HH:MM:SS', 'T. HH:MM']),
    ('kva', ['CARGA (KVA)', 'KVA INTERRUMP', 'KVA INTERR', 'CARGA KVA', 'KVA INTERRUPIDOS', 'KVA', 'KVA PROM', 'CARGA']),
    ('mva', ['CARGA (MW)', 'CARGA (MVA)', 'CARGA MW', 'MVA', 'MW', 'MVAMIN', 'CARGA_MW']),
    ('observacion', ['OBSERVACIONES', 'OBSERVACIÓN', 'OBSERVACION', 'DESCRIPCION MATERIAL', 'DESCRIPCION', 'DESCRIPCIÓN', 'NOTA DE CIERRE', 'DETALLE', 'DESCRIPCIÓN DEL EVENTO', 'OBS']),
    ('jefatura', ['CENTRO DE SERVICIO', 'CS DE SERVICIO', 'CS', 'DISTRITO', 'JEFATURA', 'C. SERVICIO', 'GERENCIA', 'ZONA']),
    ('sistema', ['TIPO DE RED', 'RESPONSABLE', 'SISTEMA', 'NIVEL DE TENSION', 'NIVEL']),
    ('sectores', ['SECTOR AFECTADO', 'SECTORES', 'SECTOR', 'POBLACION', 'POBLACIÓN', 'MUNICIPIO', 'PARROQUIA', 'SECTORES AFECTADOS']),
    ('ciudad', ['CIUDAD', 'LOCALIDAD', 'ESTADO', 'REGION', 'REGIÓN']),
    ('despachador', ['USUARIO', 'DESPACHADOR', 'OPERADOR', 'CTO/BAR', 'PERSONAL', 'DESPACHADOR / OPERADOR']),
    ('tti_cto', ['TTI CTO.', 'TTI (F)', 'MWH (TTI)', 'TTI', 'MWH']),
    ('kva_x_h', ['KVA X H', 'KVA*H', 'KVAXH', 'KVA*SEGUNDOS']),
    ('mes', ['MES']),
]


def auto_detectar_cabeceras(ws):
    """
    Escanea las primeras 25 filas de la hoja para encontrar la fila real de encabezados
    y mapear dinámicamente cada columna reconocida.
    """
    import unicodedata, re
    best_row_idx = None
    best_mapping = {}
    best_score = 0

    max_scan = min(35, ws.max_row or 1)
    for r_idx in range(1, max_scan + 1):
        row = ws[r_idx]
        mapping = {}
        matched_targets = set()
        for col_idx in range(len(row)):
            val = str(row[col_idx].value).strip() if row[col_idx].value is not None else ''
            if not val or val.lower() == 'none':
                continue

            val_norm = normalizar_texto(val)
            val_clean = re.sub(r'[\r\n\t]+', ' ', val_norm).strip()
            val_clean_alnum = re.sub(r'[^A-Z0-9]', '', val_norm)

            for target, syns in HEADER_SYNONYMS_ORDERED:
                if target in matched_targets:
                    continue
                matched = False
                for syn in syns:
                    syn_norm = normalizar_texto(syn)
                    syn_alnum = re.sub(r'[^A-Z0-9]', '', syn_norm)
                    if val_clean == syn_norm or (syn_alnum and syn_alnum == val_clean_alnum) or (len(syn_norm) >= 4 and syn_norm in val_clean):
                        mapping[col_idx] = target
                        matched_targets.add(target)
                        matched = True
                        break
                if matched:
                    break

        score = len(matched_targets)
        if 'circuito' in matched_targets: score += 4
        if 'subestacion' in matched_targets: score += 4
        if 'fecha_inicio' in matched_targets or 'fecha_falla' in matched_targets: score += 4
        if 'causa' in matched_targets: score += 3

        if score > best_score:
            best_score = score
            best_row_idx = r_idx
            best_mapping = mapping

    return best_row_idx, best_mapping, best_score


def parse_auto_detect(ws, estados_db, causas_db, estado_codigo=None):
    """
    Parsea una hoja con encabezados detectados automáticamente en cualquier fila.
    """
    h_row, mapping, score = auto_detectar_cabeceras(ws)
    if not h_row or score < 2 or len(mapping) < 2:
        return None

    records = []
    causas_originales = {}
    max_row = ws.max_row or 1

    for r in range(h_row + 1, min(max_row + 1, MAX_ROWS + 1)):
        row = ws[r]
        rec = {}
        for c_idx, target in mapping.items():
            if c_idx < len(row):
                v = row[c_idx].value
                rec[target] = str(v).strip() if v is not None and str(v).strip() != 'None' else None

        # Descartar filas completamente vacías
        non_empty = [v for v in rec.values() if v is not None and str(v).strip() != '' and str(v).strip().lower() != 'none']
        if not non_empty:
            continue

        # Descartar filas de totales o notas al pie
        first_val = str(non_empty[0]).upper()
        if first_val.startswith(('TOTAL', 'PROMEDIO', 'RESUMEN', 'NOTA', 'ELABORADO', 'APROBADO')):
            continue

        normalizar_y_extraer_registro(rec)

        causa_orig = (rec.get('causa') or '').strip()
        if causa_orig:
            key = causa_orig.upper()
            if key not in causas_originales:
                causas_originales[key] = {
                    'original': causa_orig,
                    'sub_causa': rec.get('sub_causa'),
                    'count': 0,
                }
            causas_originales[key]['count'] += 1

        records.append(rec)

    if not records:
        return None

    return {
        'ok': True,
        'formato': 'auto_detect',
        'formato_nombre': f'Detección Automática (Fila {h_row})',
        'total_rows': len(records),
        'causas_originales': list(causas_originales.values()),
        'causas_homologadas': causas_db,
        'preview': records[:20],
        '_all_records': records,
        'es_formato_catalogo': True,
        'sheet_seleccionada': ws.title,
    }


def parse_excel(filepath, sheet_name=None, estado_codigo=None):
    estados_db, causas_db = _load_caches()
    user_estado = estado_codigo or (session.get('user', {}).get('estado_codigo') if (has_request_context() and 'user' in session) else None)

    all_sheets = []
    try:
        wb_meta = openpyxl.load_workbook(filepath, read_only=True)
        all_sheets = wb_meta.sheetnames
        wb_meta.close()
    except Exception:
        pass
    if not all_sheets:
        try:
            import pandas as pd
            xl = pd.ExcelFile(filepath)
            all_sheets = xl.sheet_names
        except Exception:
            pass

    sheets_to_try = [sheet_name] if sheet_name else (all_sheets or [None])
    best_result = None
    best_row_count = -1

    for s_name in sheets_to_try:
        try:
            ws = cargar_hoja_memoria(filepath, s_name)
        except Exception:
            continue
        if ws is None or ws.max_row == 0:
            continue

        # 1. Intentar catálogo de formatos específicos (ej. AMAZONAS, TIRAS, BD)
        fmt, _ = detectar_formato_catalogo(ws, user_estado)
        if fmt:
            result = parse_formato_catalogo(ws, fmt, user_estado or '')
            if result and result.get('total_rows', 0) > best_row_count:
                result['formato_detectado'] = fmt.get('formato_codigo', 'CATALOGO')
                result['formato_nombre'] = fmt.get('formato_nombre', 'Formato Adaptativo')
                result['es_formato_catalogo'] = True
                result['sheet_seleccionada'] = ws.title
                result['sheet_names'] = all_sheets
                result['sheet_count'] = len(all_sheets)
                best_result = result
                best_row_count = result.get('total_rows', 0)
                if sheet_name:
                    return best_result

        # 2. Intentar auto-detección inteligente de cabeceras en cualquier fila
        auto_res = parse_auto_detect(ws, estados_db, causas_db, user_estado)
        if auto_res and auto_res.get('total_rows', 0) > best_row_count:
            auto_res['sheet_seleccionada'] = ws.title
            auto_res['sheet_names'] = all_sheets
            auto_res['sheet_count'] = len(all_sheets)
            best_result = auto_res
            best_row_count = auto_res.get('total_rows', 0)
            if sheet_name:
                return best_result

        # 3. Intentar validar formato estándar SCTIS
        valido, msg = validar_estructura(ws)
        if valido:
            result = parse_formato(ws, estados_db, causas_db)
            if result and result.get('total_rows', 0) > best_row_count:
                result['es_formato_catalogo'] = False
                result['sheet_seleccionada'] = ws.title
                result['sheet_names'] = all_sheets
                result['sheet_count'] = len(all_sheets)
                best_result = result
                best_row_count = result.get('total_rows', 0)
                if sheet_name:
                    return best_result

    if best_result and best_row_count > 0:
        return best_result

    if best_result:
        return best_result

    return {
        'error': 'formato_invalido',
        'mensaje': 'No se encontraron registros de interrupciones válidos en las hojas del archivo.',
        'formato_sugerencia': 'Verifique que la hoja contenga encabezados como CIRCUITO, SUBESTACIÓN, FECHA, CAUSA.',
        'ok': False,
    }


# ─── Rutas ─────────────────────────────────────────────────────

@bp.route('/importar', methods=['GET'])
@login_required
def importar_page():
    user_estado = g.user.get('estado_codigo')
    if g.is_admin:
        estados = query("SELECT state_code, state_name FROM common.states ORDER BY state_name")
    else:
        estados = query("SELECT state_code, state_name FROM common.states WHERE state_code = %s ORDER BY state_name", (user_estado,))
    return render_template('importar.html', estados=estados, user_estado=user_estado)


def normalizar_archivo_a_xlsx(filepath, token):
    """
    Convierte cualquier archivo tabular (.xlsx, .xls binario, .xls HTML, .csv, .tsv)
    en un archivo estándar y devuelve la ruta al archivo normalizado y la lista de nombres de hojas.
    """
    ext = os.path.splitext(filepath)[1].lower()
    
    # Si ya es .xls / .xlt binario, preservarlo directamente ya que xlrd lo lee en <0.05s
    if ext in ('.xls', '.xlt'):
        out_xls = os.path.join(UPLOAD_DIR, f'{token}{ext}')
        try:
            import xlrd
            xwb = xlrd.open_workbook(filepath, on_demand=True)
            names = xwb.sheet_names()
            if filepath != out_xls:
                import shutil
                shutil.copyfile(filepath, out_xls)
            return out_xls, names
        except Exception:
            pass

    out_path = os.path.join(UPLOAD_DIR, f'{token}.xlsx')

    # 1. Intentar con openpyxl directo (formato .xlsx moderno)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        if filepath != out_path:
            import shutil
            shutil.copyfile(filepath, out_path)
        return out_path, names
    except Exception:
        pass

    # 2. Intentar con xlrd (.xls binario de Excel 97-2003)
    try:
        import xlrd
        xwb = xlrd.open_workbook(filepath)
        wb_new = openpyxl.Workbook()
        wb_new.remove(wb_new.active)
        sheet_names = []
        for sname in xwb.sheet_names():
            xws = xwb.sheet_by_name(sname)
            ws_new = wb_new.create_sheet(title=sname)
            sheet_names.append(sname)
            for r in range(xws.nrows):
                row_vals = []
                for c in range(xws.ncols):
                    cell = xws.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            val = xlrd.xldate_as_datetime(cell.value, xwb.datemode)
                        except Exception:
                            val = cell.value
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        val = bool(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        val = None
                    else:
                        val = cell.value
                    row_vals.append(val)
                ws_new.append(row_vals)
        wb_new.save(out_path)
        return out_path, sheet_names
    except Exception:
        pass

    # 3. Intentar como tabla HTML disfrazada de .xls (exportes comunes de sistemas web / SCADA)
    try:
        import pandas as pd
        tables = pd.read_html(filepath)
        if tables:
            df = tables[0]
            df.to_excel(out_path, index=False)
            return out_path, ['Sheet1']
    except Exception:
        pass

    # 4. Intentar como CSV / TSV texto plano
    try:
        import csv
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        if lines:
            delim = '\t' if '\t' in lines[0] else (';' if ';' in lines[0] else ',')
            reader = csv.reader(lines, delimiter=delim)
            wb_new = openpyxl.Workbook()
            ws_new = wb_new.active
            ws_new.title = 'Tiras'
            for r in reader:
                ws_new.append(r)
            wb_new.save(out_path)
            return out_path, ['Tiras']
    except Exception:
        pass

    raise ValueError('El archivo no pudo ser leído como Excel (.xlsx, .xls) ni como formato tabular compatible.')


def crear_excel_desde_texto_pegado(pasted_text, tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tiras'
    lines = [l.strip() for l in pasted_text.strip().splitlines() if l.strip()]
    if not lines:
        raise ValueError('No se encontraron líneas de datos')
    
    import csv
    first_line = lines[0]
    delimiter = '\t' if '\t' in first_line else (';' if ';' in first_line else ',')
    reader = csv.reader(lines, delimiter=delimiter)
    raw_rows = list(reader)
    if not raw_rows:
        raise ValueError('No se detectaron filas válidas')
    
    for r in raw_rows:
        ws.append(r)
    wb.save(tmp_path)
    return len(raw_rows)


@bp.route('/api/importar/preview', methods=['POST'])
def preview_import():
    try:
        req_json = request.get_json(silent=True) or {}
        pasted_text = request.form.get('pasted_text') or req_json.get('pasted_text')
        file = request.files.get('file')
        token = request.form.get('token') or req_json.get('token') or session.get('import_token')
        sheet_name = request.form.get('sheet_name') or req_json.get('sheet_name') or None
        user_estado = session.get('user', {}).get('estado_codigo')
        form_estado = request.form.get('estado_codigo') or req_json.get('estado_codigo')
        estado = user_estado or form_estado

        if not file and not token and not pasted_text:
            return jsonify({'error': 'No se envió ningún archivo ni datos de portapapeles', 'ok': False}), 400

        os.makedirs(UPLOAD_DIR, exist_ok=True)

        # ── Caso A: Datos pegados desde Excel (Clipboard TSV / CSV) ──
        if pasted_text:
            token = uuid.uuid4().hex
            tmp_path = os.path.join(UPLOAD_DIR, f'{token}.xlsx')
            try:
                crear_excel_desde_texto_pegado(pasted_text, tmp_path)
                session['import_token'] = token
                sheet_name = 'Tiras'
            except Exception as e:
                return jsonify({'error': f'Error procesando texto pegado de Excel: {str(e)}', 'ok': False}), 400

        # ── Caso B: Subida de archivo Excel ──
        elif file and not sheet_name:
            if file.filename == '':
                return jsonify({'error': 'Nombre de archivo vacío', 'ok': False}), 400

            ext = os.path.splitext(file.filename)[1] or '.xlsx'
            token = uuid.uuid4().hex
            raw_tmp_path = os.path.join(UPLOAD_DIR, f'{token}_raw{ext}')
            file.save(raw_tmp_path)
            session['import_token'] = token

            try:
                tmp_path, sheet_names = normalizar_archivo_a_xlsx(raw_tmp_path, token)
            except Exception as e:
                if os.path.exists(raw_tmp_path):
                    os.remove(raw_tmp_path)
                return jsonify({'error': f'El archivo no es compatible: {str(e)}', 'ok': False}), 400

            visibles = listar_hojas_visibles(tmp_path) or sheet_names
            sheet_name = None  # Permite que parse_excel escanee y auto-seleccione la hoja con registros válidos

        if not token:
            return jsonify({'error': 'Sesión expirada. Vuelva a cargar los datos.', 'ok': False}), 400

        tmp_path = os.path.join(UPLOAD_DIR, f'{token}.xlsx')
        if not os.path.exists(tmp_path):
            tmp_path = os.path.join(UPLOAD_DIR, f'{token}{ext_from_token(token)}')
        if not os.path.exists(tmp_path):
            return jsonify({'error': 'Archivo no encontrado en servidor. Vuelva a cargar los datos.', 'ok': False}), 400

        result = parse_excel(tmp_path, sheet_name, estado)
        if 'error' in result:
            status = 422 if result.get('error') == 'formato_invalido' else 400
            return jsonify(result), status

        if user_estado:
            result['estado_codigo'] = user_estado
        elif form_estado:
            result['estado_codigo'] = form_estado

        try:
            wb_sheets = openpyxl.load_workbook(tmp_path, read_only=True)
            result['sheet_names'] = wb_sheets.sheetnames
            result['sheet_count'] = len(wb_sheets.sheetnames)
            wb_sheets.close()
        except Exception:
            result['sheet_names'] = []
            result['sheet_count'] = 0

        submission_id = None
        try:
            user_username = session.get('user', {}).get('username', 'unknown')
            sub_result = query("""
                INSERT INTO audit.submissions
                    (process_code, state_code, source_filename, source_sheet,
                     sheet_names, row_count, validation_status, formato_codigo,
                     ingested_by, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING submission_id
            """, (
                'sctis_import',
                estado,
                os.path.basename(tmp_path),
                result.get('sheet_seleccionada') or sheet_name,
                str(result.get('sheet_names', [])),
                result.get('total_rows', 0),
                'PENDING',
                result.get('formato_detectado'),
                user_username,
                f"Formato: {result.get('formato_detectado', 'CTIS_HOMOLOGADO')}",
            ), fetch=True)
            if sub_result and len(sub_result) > 0:
                submission_id = sub_result[0]['submission_id']
        except Exception:
            pass

        result['submission_id'] = submission_id

        records = result.get('_all_records') or []
        incompletos_count = sum(1 for r in records if es_registro_incompleto(r))
        user_role = session.get('role_code') or session.get('user', {}).get('role_code')
        is_admin = (user_role == 'admin')

        result['registros_incompletos'] = incompletos_count
        result['alerta_incompletos'] = (incompletos_count > 0)
        result['es_admin'] = is_admin
        result['bloqueado_no_admin'] = (incompletos_count > 0 and not is_admin)

        if records and estado:
            result['activos_inconsistentes'] = detectar_activos_inconsistentes(records, estado)
        else:
            result['activos_inconsistentes'] = {'subestaciones': [], 'circuitos': []}
        result.pop('_all_records', None)

        # Process and classify causes
        _, causas_db = _load_caches()
        causas_procesadas, auto_count, pendientes_count = procesar_causas_preview(result.get('causas_originales', []), causas_db)
        result['causas_originales'] = causas_procesadas
        result['causas_homologadas'] = causas_db
        result['auto_homologadas_count'] = auto_count
        result['pendientes_count'] = pendientes_count

        preview_path = os.path.join(UPLOAD_DIR, f'{token}_preview.json')
        with open(preview_path, 'w') as f:
            json.dump(result, f, default=str)

        return jsonify(result)
    except Exception as err:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error en procesamiento: {str(err)}', 'ok': False}), 400



@bp.route('/api/importar/confirmar', methods=['POST'])
def confirmar_import():
    data = request.get_json()
    if not data or 'mapping' not in data:
        return jsonify({'error': 'Se requiere el mapeo de causas'}), 400

    user_estado = session.get('user', {}).get('estado_codigo')
    estado_codigo = data.get('estado_codigo') or user_estado
    if not estado_codigo:
        return jsonify({'error': 'Debe seleccionar un estado destino para la importación.'}), 400

    token = session.get('import_token')
    if not token:
        return jsonify({'error': 'Sesi\xf3n expirada. Vuelva a subir el archivo.'}), 400

    tmp_path = os.path.join(UPLOAD_DIR, f'{token}.xlsx')
    preview_path = os.path.join(UPLOAD_DIR, f'{token}_preview.json')

    if not os.path.exists(tmp_path):
        session.pop('import_token', None)
        return jsonify({'error': 'Archivo no encontrado. Vuelva a subir el archivo.'}), 400

    if os.path.exists(preview_path):
        with open(preview_path) as f:
            result = json.load(f)
    else:
        result = {}

    mapping = data['mapping']
    se_map = data.get('subestacion_map') or {}
    ci_map = data.get('circuito_map') or {}
    se_nuevos = data.get('se_nuevos') or []
    ci_nuevos = data.get('ci_nuevos') or []
    mes_override = data.get('mes') or result.get('mes')

    usuario_alias = session.get('user', {}).get('username')
    aliases = cargar_aliases_assets(estado_codigo)
    aliases_se = aliases['subestaciones']
    aliases_ci = aliases['circuitos']
    guardar_aliases_assets(se_map, ci_map, estado_codigo, usuario_alias)

    se_nuevos_set = set(str(n).strip() for n in se_nuevos if n and str(n).strip())
    ci_nuevos_set = set(
        f'{str(c.get("circuito")).strip()}|{str(c.get("subestacion")).strip()}'
        for c in ci_nuevos if c and c.get('circuito') and str(c.get('circuito')).strip()
    )
    filas_se = {}
    filas_ci = {}

    sheet_name = result.get('sheet_seleccionada')
    
    # Parsear el libro con parse_excel para soportar todos los formatos de catálogo y formato estándar
    parsed_excel = parse_excel(tmp_path, sheet_name, estado_codigo)
    if 'error' in parsed_excel:
        return jsonify({'error': parsed_excel.get('mensaje') or 'Error al procesar el archivo Excel.'}), 400

    all_parsed_records = parsed_excel.get('_all_records') or []
    insertados = 0
    errores = []

    # Resolver nombres catalogados
    resolver_asset = {}
    for se_orig, asset_id in se_map.items():
        if asset_id:
            a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name, asset_code FROM common.assets WHERE asset_id = %s", (asset_id,))
            if a:
                resolver_asset[se_orig] = {'asset_id': asset_id, 'asset_name': a['asset_name']}

    resolver_circuito = {}
    for ci_key, asset_id in ci_map.items():
        if asset_id:
            a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (asset_id,))
            if a:
                resolver_circuito[ci_key] = {'asset_id': asset_id, 'asset_name': a['asset_name']}

    # Pre-parse and validate records for incomplete dates/times
    records_cache = []
    incompletos_count = 0
    for row_idx, record in enumerate(all_parsed_records, start=1):
        normalizar_y_extraer_registro(record)
        is_inc = es_registro_incompleto(record)
        record['_es_incompleto'] = is_inc
        if is_inc:
            incompletos_count += 1
        records_cache.append((row_idx, record))

    user_role = session.get('role_code') or session.get('user', {}).get('role_code')
    is_admin = (user_role == 'admin')

    if incompletos_count > 0 and not is_admin:
        return jsonify({
            'error': f'Carga bloqueada por cumplimiento de norma: El archivo contiene {incompletos_count} registro(s) incompletos (sin fecha u hora de cierre). Solamente un Administrador del Sistema posee privilegios para autorizar la carga de datos incompletos.'
        }), 403

    audit_id = None
    if incompletos_count > 0 and is_admin:
        if not data.get('aceptar_no_repudio'):
            return jsonify({
                'error': f'Declaración de No Repudio requerida: Debe aceptar formalmente la Declaración de No Repudio y Responsabilidad Administrativa para autorizar la carga de {incompletos_count} registro(s) incompletos.'
            }), 400

        user_id = session.get('user_id') or session.get('user', {}).get('user_id')
        username = session.get('user', {}).get('username') or session.get('username') or 'admin'
        declaracion = data.get('declaracion_no_repudio') or f'Carga de data fuera de norma autorizada por {username} con {incompletos_count} registros incompletos.'
        ip_addr = request.remote_addr or '127.0.0.1'
        user_agent = request.headers.get('User-Agent', '')

        audit_id = insert_audit_carga_excepcional(
            user_id, username, estado_codigo,
            os.path.basename(tmp_path), token,
            len(records_cache), incompletos_count,
            declaracion, ip_addr, user_agent
        )

    # Catalogo fuzzy para fallback de matching
    cat_se_fuzzy = query("""
        SELECT asset_id, 
               COALESCE(asset_name_normalizado, asset_name) AS asset_name,
               asset_name AS asset_name_original
        FROM common.assets
        WHERE asset_type = 'SUBSTATION' AND state_code = %s AND is_active = true
    """, (estado_codigo,))
    cat_se_fuzzy_map = {}
    for a in (cat_se_fuzzy or []):
        cat_se_fuzzy_map[normalizar_para_matching(a['asset_name'])] = a
        if a.get('asset_name_original'):
            cat_se_fuzzy_map[normalizar_para_matching(a['asset_name_original'])] = a

    despachadores_cache = {}

    for row_idx, record in records_cache:
        se_orig_parsed = (record.get('subestacion') or '').strip()
        ci_orig_parsed = (record.get('circuito') or '').strip()
        if se_orig_parsed:
            filas_se[se_orig_parsed] = filas_se.get(se_orig_parsed, 0) + 1
        if se_orig_parsed and ci_orig_parsed:
            filas_ci[f'{ci_orig_parsed}|{se_orig_parsed}'] = filas_ci.get(f'{ci_orig_parsed}|{se_orig_parsed}', 0) + 1

        causa_raw = (record.get('causa') or '').strip()
        causa_orig_upper = causa_raw.upper()
        causa_id = mapping.get(causa_raw) or mapping.get(causa_orig_upper) or mapping.get(str(causa_raw))
        if isinstance(causa_id, str) and causa_id.isdigit():
            causa_id = int(causa_id)

        if not causa_id and causa_raw:
            _, causas_db = _load_caches()
            norm_raw = normalizar_texto(causa_raw)
            for c in causas_db:
                if norm_raw == normalizar_texto(c['causa_nombre']) or norm_raw == normalizar_texto(c.get('causa_codigo', '')):
                    causa_id = c['causa_id']
                    break
            if not causa_id:
                otros = next((c for c in causas_db if c['causa_codigo'] == 'OTROS' or 'OTROS' in c['causa_nombre'].upper()), None)
                if otros:
                    causa_id = otros['causa_id']

        try:
            estado = estado_codigo

            if not estado:
                errores.append(f'Fila {row_idx}: no se pudo determinar el estado')
                continue

            # Resolver subestacion
            se_orig = (record.get('subestacion') or '').strip()
            if se_orig and se_orig in se_nuevos_set:
                record['subestacion_id'] = None
            elif se_orig and se_orig in resolver_asset:
                record['subestacion'] = resolver_asset[se_orig]['asset_name']
                record['subestacion_id'] = resolver_asset[se_orig]['asset_id']
            elif se_orig:
                norm_se = normalizar_para_matching(se_orig)
                alias_id = aliases_se.get(norm_se)
                if alias_id:
                    a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name, asset_code FROM common.assets WHERE asset_id = %s", (alias_id,))
                    if a:
                        record['subestacion'] = a['asset_name']
                        record['subestacion_id'] = alias_id
                elif norm_se in cat_se_fuzzy_map:
                    a = cat_se_fuzzy_map[norm_se]
                    record['subestacion'] = a['asset_name']
                    record['subestacion_id'] = a['asset_id']
                else:
                    for norm_key, a in cat_se_fuzzy_map.items():
                        if norm_se in norm_key or norm_key in norm_se:
                            record['subestacion'] = a['asset_name']
                            record['subestacion_id'] = a['asset_id']
                            break

            # Resolver circuito
            ci_orig = (record.get('circuito') or '').strip()
            ci_key = f'{ci_orig}|{se_orig}'
            if ci_orig and ci_key in ci_nuevos_set:
                record['circuito_id'] = None
            elif ci_orig and ci_key in resolver_circuito:
                record['circuito'] = resolver_circuito[ci_key]['asset_name']
                record['circuito_id'] = resolver_circuito[ci_key]['asset_id']
            elif ci_orig and record.get('subestacion_id'):
                norm_ci = normalizar_para_matching(ci_orig)
                alias_id = aliases_ci.get((norm_ci, normalizar_para_matching(se_orig)))
                if alias_id:
                    a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (alias_id,))
                    if a:
                        record['circuito'] = a['asset_name']
                        record['circuito_id'] = alias_id
                else:
                    cat_ci_fallback = query("""
                        SELECT asset_id, COALESCE(asset_name_normalizado, asset_name) AS asset_name,
                               asset_name AS asset_name_original
                        FROM common.assets
                        WHERE asset_type = 'CIRCUITO' AND parent_asset_id = %s AND is_active = true
                    """, (record['subestacion_id'],))
                    for a in (cat_ci_fallback or []):
                        norm_a = normalizar_para_matching(a['asset_name'])
                        norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
                        if norm_a == norm_ci or norm_orig == norm_ci or norm_ci in norm_a or norm_a in norm_ci:
                            record['circuito'] = a['asset_name']
                            record['circuito_id'] = a['asset_id']
                            break

            desp = record.get('despachador')
            desp_id = None
            if desp:
                desp_clean = str(desp).strip()
                if desp_clean not in despachadores_cache:
                    despachadores_cache[desp_clean] = asegurar_despachador(desp_clean, estado)
                desp_id = despachadores_cache[desp_clean]
                if desp_id:
                    record['despachador'] = normalizar_nombre(desp_clean)

            f_falla_dt = normalizar_a_date(record.get('fecha_falla'))
            f_ini_dt = normalizar_a_datetime(record.get('fecha_inicio'))
            f_fin_dt = normalizar_a_datetime(record.get('fecha_fin'))

            if not f_ini_dt:
                if f_falla_dt:
                    f_ini_dt = datetime.combine(f_falla_dt, datetime.min.time())
                else:
                    f_ini_dt = datetime.now()

            if not f_falla_dt:
                f_falla_dt = f_ini_dt.date() if f_ini_dt else datetime.now().date()

            is_inc = record.get('_es_incompleto', False)
            if is_inc or not f_fin_dt:
                f_fin_dt = f_ini_dt
                horas = 0.0
                estado_calc = 'INCOMPLETO_EXCEPCION_ADMIN'
                es_exc = 1
                audit_id_val = audit_id
            else:
                horas = record.get('horas')
                if horas is not None:
                    try:
                        horas = float(horas)
                    except Exception:
                        horas = 0.0
                else:
                    horas = 0.0
                estado_calc = record.get('estado_calculo') or 'CALCULO VALIDO'
                es_exc = 0
                audit_id_val = None

            f_falla = f_falla_dt.isoformat() if f_falla_dt else datetime.now().date().isoformat()
            f_ini = f_ini_dt.strftime('%Y-%m-%d %H:%M:%S') if f_ini_dt else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f_fin = f_fin_dt.strftime('%Y-%m-%d %H:%M:%S') if f_fin_dt else f_ini

            kva = record.get('kva')
            if kva is not None:
                try:
                    kva = float(kva)
                except Exception:
                    kva = 0.0

            subestacion_val = record.get('subestacion')
            circuito_val = record.get('circuito')

            insert_sql = """
                INSERT INTO "sctis.tira_interrupcion"
                    (estado_codigo, fecha_falla, sistema, subestacion, subestacion_id,
                     circuito, circuito_id, jefatura,
                     fecha_inicio, fecha_fin, causa, sub_causa, observacion,
                     despachador, despachador_id, kva, horas, mes, sectores, ciudad,
                     causa_id, created_by, es_excepcion_admin, audit_id, estado_calculo)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            mes_val = mes_override or record.get('mes')
            query(insert_sql, (
                estado, f_falla, record.get('sistema'),
                subestacion_val, record.get('subestacion_id'),
                circuito_val, record.get('circuito_id'),
                record.get('jefatura'),
                f_ini, f_fin, record.get('causa'), record.get('sub_causa'),
                record.get('observacion'), record.get('despachador'), desp_id,
                kva, horas, mes_val,
                record.get('sectores'), record.get('ciudad'),
                causa_id, 'importacion_excel',
                es_exc, audit_id_val, estado_calc
            ), fetch=False, commit=False)
            insertados += 1
        except Exception as e:
            errores.append(f'Fila {row_idx}: {str(e)}')

    commit_db()

    # Encolar activos reportados como nuevos para revisión del administrador
    requests_creados = 0
    if se_nuevos_set or ci_nuevos_set:
        cat_ci_estado = query("""
            SELECT a.asset_id, COALESCE(a.asset_name_normalizado, a.asset_name) AS asset_name
            FROM common.assets a
            JOIN common.assets p ON p.asset_id = a.parent_asset_id
            WHERE a.asset_type = 'CIRCUITO' AND p.state_code = %s AND a.is_active = true
        """, (estado_codigo,)) or []
        se_request_ids = {}
        for se_n in se_nuevos_set:
            clasif, sug = preclasificar_activo(se_n, cat_se_fuzzy or [])
            try:
                res = query("""
                    INSERT INTO sctis.asset_request
                        (estado_codigo, asset_type, nombre_reportado, nombre_normalizado,
                         filas_afectadas, clasificacion, sugerencia_alias,
                         submission_id, requested_by)
                    VALUES (%s, 'SUBSTATION', %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (estado_codigo, asset_type, nombre_reportado, se_referencia)
                    DO UPDATE SET filas_afectadas = sctis.asset_request.filas_afectadas + EXCLUDED.filas_afectadas,
                                  updated_at = now()
                    RETURNING request_id
                """, (estado_codigo, se_n, normalizar_nombre(se_n),
                      filas_se.get(se_n, 0), clasif, sug, submission_id, usuario_alias))
                if res:
                    se_request_ids[se_n] = res[0]['request_id']
                    requests_creados += 1
            except Exception:
                pass
        for ci_n in ci_nuevos_set:
            ci, se = ci_n.split('|', 1)
            clasif, sug = preclasificar_activo(ci, cat_ci_estado)
            try:
                query("""
                    INSERT INTO sctis.asset_request
                        (estado_codigo, asset_type, nombre_reportado, nombre_normalizado,
                         se_referencia, se_request_id, filas_afectadas, clasificacion,
                         sugerencia_alias, submission_id, requested_by)
                    VALUES (%s, 'CIRCUITO', %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (estado_codigo, asset_type, nombre_reportado, se_referencia)
                    DO UPDATE SET filas_afectadas = sctis.asset_request.filas_afectadas + EXCLUDED.filas_afectadas,
                                  updated_at = now()
                """, (estado_codigo, ci, normalizar_nombre(ci), se,
                      se_request_ids.get(se), filas_ci.get(ci_n, 0), clasif, sug,
                      submission_id, usuario_alias), fetch=False)
                requests_creados += 1
            except Exception:
                pass

    submission_id = result.get('submission_id')
    total_intentados = result.get('total_rows', insertados + len(errores))
    rechazados = len(errores)

    if submission_id:
        try:
            query("""
                UPDATE audit.submissions SET
                    accepted_count = %s,
                    rejected_count = %s,
                    validation_status = %s
                WHERE submission_id = %s
            """, (insertados, rechazados,
                  'VALIDATED' if rechazados == 0 else 'PARTIAL',
                  submission_id), fetch=False)
        except Exception:
            pass

        if rechazados > 0:
            user_id = g.user.get('user_id') if hasattr(g, 'user') and g.user else None
            if not user_id:
                try:
                    u = query_one("SELECT user_id FROM sctis.user_profiles WHERE username = %s",
                                  (session.get('user', {}).get('username', ''),))
                    user_id = u['user_id'] if u else None
                except Exception:
                    pass

            if user_id:
                error_records = []
                for i, err in enumerate(errores[:50]):
                    parts = err.split(': ', 1)
                    row_num = parts[0].replace('Fila ', '') if parts else ''
                    error_msg = parts[1] if len(parts) > 1 else err
                    error_records.append({
                        'row_numero': row_num,
                        'error': error_msg,
                    })

                try:
                    corr_file = generar_excel_correccion(error_records, token, estado_codigo)
                    corr_filename = os.path.basename(corr_file)
                except Exception:
                    corr_filename = None

                try:
                    query("""
                        INSERT INTO sctis.tarea_pendiente
                            (usuario_id, estado_codigo, tipo_tarea, submission_id,
                             descripcion, registros_total, registros_ok, registros_rechazados,
                             archivo_correccion, estado_tarea)
                        VALUES (%s, %s, 'CORREGIR_DATOS', %s, %s, %s, %s, %s, %s, 'PENDIENTE')
                    """, (
                        user_id, estado_codigo, submission_id,
                        f'Carga con {rechazados} registros rechazados de {total_intentados} totales. '
                        f'Archivo: {result.get("source_filename", "desconocido")}',
                        total_intentados, insertados, rechazados,
                        corr_filename,
                    ), fetch=False)
                except Exception:
                    pass

                if corr_filename:
                    try:
                        query("""
                            UPDATE audit.submissions SET correction_file = %s
                            WHERE submission_id = %s
                        """, (corr_filename, submission_id), fetch=False)
                    except Exception:
                        pass

        if requests_creados > 0:
            admin = query_one("""
                SELECT up.user_id FROM sctis.user_profiles up
                JOIN sctis.user_roles r ON r.role_id = up.role_id
                WHERE r.role_code = 'admin' AND up.is_active = true
                ORDER BY up.user_id LIMIT 1
            """)
            if admin:
                try:
                    query("""
                        INSERT INTO sctis.tarea_pendiente
                            (usuario_id, estado_codigo, tipo_tarea, submission_id,
                             descripcion, registros_total, registros_ok, registros_rechazados,
                             estado_tarea)
                        VALUES (%s, %s, 'APROBAR_ACTIVO', %s, %s, %s, %s, %s, 'PENDIENTE')
                    """, (
                        admin['user_id'], estado_codigo, submission_id,
                        f'{requests_creados} activo(s) (SE/CT) reportados fuera de catálogo '
                        f'pendientes de revisión. Archivo: {result.get("source_filename", "desconocido")}',
                        requests_creados, 0, requests_creados,
                    ), fetch=False)
                except Exception:
                    pass


    # -- Enviar notificación email si está configurado --
    try:
        email_config = query_one("SELECT valor FROM sctis.configuracion WHERE clave = 'notificacion_email'")
        if email_config and email_config['valor']:
            email_to = email_config['valor']
            asunto = f"Nueva Carga de Datos ({estado_codigo})"
            cuerpo = f"Se ha realizado una nueva carga de datos para el estado {estado_codigo}.\n"
            cuerpo += f"Registros procesados exitosamente: {insertados}\n"
            if rechazados > 0:
                cuerpo += f"Se generó una tarea de corrección por {rechazados} registros rechazados.\n"
            
            # En un entorno real se usaría la cuenta de servicio o un token válido de aplicación.
            # send_notification_email(access_token_app, email_to, asunto, cuerpo)
            pass
    except Exception as e:
        print("Error al intentar notificar", e)

    for f in [tmp_path, preview_path]:
        if os.path.exists(f):
            os.remove(f)
    session.pop('import_token', None)

    return jsonify({
        'ok': True,
        'insertados': insertados,
        'errores': errores[:20],
        'total_errores': len(errores),
        'submission_id': submission_id,
        'correccion_generada': rechazados > 0,
        'requests_creados': requests_creados,
    })


@bp.route('/api/importar/limpiar', methods=['POST'])
def limpiar_import():
    token = session.pop('import_token', None)
    if token:
        for f in [os.path.join(UPLOAD_DIR, f'{token}.xlsx'), os.path.join(UPLOAD_DIR, f'{token}_preview.json')]:
            if os.path.exists(f):
                os.remove(f)
    return jsonify({'ok': True})


@bp.route('/api/importar/sugerir-causas', methods=['POST'])
@login_required
def sugerir_causas_route():
    data = request.get_json() or {}
    causas_input = data.get('causas') or []
    if isinstance(causas_input, str):
        causas_input = [causas_input]

    _, causas_db = _load_caches()
    sugerencias = {}

    for c_text in causas_input:
        if not c_text:
            continue
        norm_orig = normalizar_texto(c_text)
        best_ratio = 0.0
        best_c = None

        for c in causas_db:
            norm_c = normalizar_texto(c['causa_nombre'])
            if norm_orig == norm_c:
                best_c = c
                best_ratio = 1.0
                break
            ratio = difflib.SequenceMatcher(None, norm_orig, norm_c).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_c = c

        if best_c:
            sugerencias[c_text] = {
                'causa_id': best_c['causa_id'],
                'causa_codigo': best_c['causa_codigo'],
                'causa_nombre': best_c['causa_nombre'],
                'confianza': round(best_ratio * 100)
            }

    return jsonify({'ok': True, 'sugerencias': sugerencias})

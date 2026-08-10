from flask import Blueprint, render_template, request, jsonify, session, g
from app.db import query, query_one
from app.auth import login_required
import openpyxl
import os
import uuid
import json
import difflib
from datetime import datetime, date, timedelta
import re


def normalizar_nombre(nombre):
    if not nombre:
        return nombre
    nombre = re.sub(r'\s+', ' ', nombre.strip())
    return ' '.join(w.capitalize() for w in nombre.split())


def normalizar_para_matching(nombre):
    if not nombre:
        return ''
    n = nombre.strip().upper()
    n = re.sub(r'\s*\d+[\.,]?\d*(?:/[\d]+[\.,]?\d*)*\s*[Kk][Vv]\.?', '', n)
    n = re.sub(r'\s*\d+[\.,]?\d*(?:/[\d]+[\.,]?\d*)+\s*', ' ', n)
    n = re.sub(r'\s+\d+[\.,]\d+\s*$', '', n)
    n = re.sub(r'\s+(\d{2,})\s*$', '', n)
    n = re.sub(r'\s*\([^)]*\)\s*', ' ', n)
    n = re.sub(r'\s*-\s*', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def asegurar_despachador(nombre_completo, estado_codigo):
    if not nombre_completo or not nombre_completo.strip():
        return None
    nombre = normalizar_nombre(nombre_completo)
    existente = query_one(
        "SELECT despachador_id FROM sctis.despachador WHERE LOWER(nombre_completo) = LOWER(%s)",
        (nombre,)
    )
    if existente:
        return existente['despachador_id']
    result = query(
        "INSERT INTO sctis.despachador (nombre_completo, estado_codigo, activo) VALUES (%s, %s, true) RETURNING despachador_id",
        (nombre, estado_codigo)
    )
    return result[0]['despachador_id']

def cargar_aliases_assets(estado_codigo):
    if not estado_codigo:
        return {'subestaciones': {}, 'circuitos': {}}
    rows = query("""
        SELECT alias_nombre, se_referencia, asset_type, asset_id
        FROM sctis.asset_alias
        WHERE estado_codigo = %s
    """, (estado_codigo,))
    se_map = {}
    ci_map = {}
    for r in (rows or []):
        if r['asset_type'] == 'SUBSTATION':
            se_map[normalizar_para_matching(r['alias_nombre'])] = r['asset_id']
        else:
            ci_map[(normalizar_para_matching(r['alias_nombre']),
                    normalizar_para_matching(r['se_referencia']))] = r['asset_id']
    return {'subestaciones': se_map, 'circuitos': ci_map}


def guardar_aliases_assets(se_map, ci_map, estado_codigo, usuario=None):
    guardados = 0
    for se_orig, asset_id in (se_map or {}).items():
        if not asset_id or not str(se_orig).strip():
            continue
        query("""
            INSERT INTO sctis.asset_alias
                (estado_codigo, asset_type, alias_nombre, se_referencia, asset_id, usuario)
            VALUES (%s, 'SUBSTATION', %s, '', %s, %s)
            ON CONFLICT (estado_codigo, asset_type, alias_nombre, se_referencia)
            DO UPDATE SET asset_id = EXCLUDED.asset_id, usuario = EXCLUDED.usuario,
                          updated_at = now()
        """, (estado_codigo, str(se_orig).strip(), asset_id, usuario))
        guardados += 1
    for ci_key, asset_id in (ci_map or {}).items():
        if not asset_id:
            continue
        ci, se = str(ci_key).split('|', 1) if '|' in str(ci_key) else (str(ci_key), '')
        if not ci.strip():
            continue
        query("""
            INSERT INTO sctis.asset_alias
                (estado_codigo, asset_type, alias_nombre, se_referencia, asset_id, usuario)
            VALUES (%s, 'CIRCUITO', %s, %s, %s, %s)
            ON CONFLICT (estado_codigo, asset_type, alias_nombre, se_referencia)
            DO UPDATE SET asset_id = EXCLUDED.asset_id, usuario = EXCLUDED.usuario,
                          updated_at = now()
        """, (estado_codigo, str(ci).strip(), str(se).strip(), asset_id, usuario))
        guardados += 1
    return guardados


def preclasificar_activo(nombre, catalogo):
    if not nombre or not catalogo:
        return 'PROBABLE_NUEVO', None
    norm = normalizar_para_matching(nombre)
    if not norm:
        return 'PROBABLE_NUEVO', None
    mejor_ratio = 0.0
    mejor_id = None
    for a in catalogo:
        cand = normalizar_para_matching(a.get('asset_name', ''))
        if not cand:
            continue
        ratio = difflib.SequenceMatcher(None, norm, cand).ratio()
        if ratio > mejor_ratio:
            mejor_ratio = ratio
            mejor_id = a['asset_id']
    if mejor_ratio >= 0.85:
        return 'PROBABLE_TYPEO', mejor_id
    if mejor_ratio >= 0.60:
        return 'PROBABLE_ALIAS', mejor_id
    return 'PROBABLE_NUEVO', None


def es_registro_incompleto(record):
    if not record:
        return True
    f_fin = record.get('fecha_fin')
    h_fin = record.get('hora_fin')
    f_ini = record.get('fecha_inicio')
    h_ini = record.get('hora_inicio')
    hrs = record.get('horas')
    dur = record.get('duracion_str') or record.get('duracion')

    if not f_fin or str(f_fin).strip() in ['', 'None', 'NoneTNone']:
        return True
    if not h_fin and 'T' not in str(f_fin):
        return True
    if f_ini and f_fin and str(f_ini).strip() == str(f_fin).strip():
        if h_ini and h_fin and str(h_ini).strip() == str(h_fin).strip():
            return True
    if hrs is None or str(hrs).strip() in ['', '0', '0.0', '0,0']:
        if not dur or str(dur).strip() in ['', '00:00:00', '00:00', '0', '0,0']:
            return True
    return False


def insert_audit_carga_excepcional(user_id, username, estado_codigo, filename, token, total_registros, registros_incompletos, declaracion, ip_address, user_agent):
    conn = get_db()
    db_type = getattr(g, 'db_type', 'postgres')
    sql = """
        INSERT INTO "sctis.audit_admin_carga_excepcional"
            (user_id, username, estado_codigo, filename, token, total_registros, registros_incompletos, declaracion_no_repudio, ip_address, user_agent)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    params = (user_id, username, estado_codigo, filename, token, total_registros, registros_incompletos, declaracion, ip_address, user_agent)
    if db_type == 'postgres':
        res = query(sql + " RETURNING audit_id", params, fetch=True)
        return res[0]['audit_id'] if res else None
    else:
        adapted_sql = adapt_sql_for_sqlite(sql)
        cur = conn.cursor()
        cur.execute(adapted_sql, params)
        conn.commit()
        return cur.lastrowid



bp = Blueprint('importar', __name__)
UPLOAD_DIR = '/tmp/sctis_imports'

MAX_ROWS = 2000

COL_MAP_FORMATO = {
    'A': 'estado_nombre',
    'B': 'sistema',
    'C': 'jefatura',
    'D': 'subestacion',
    'E': 'circuito',
    'F': 'fecha_inicio',
    'G': 'hora_inicio',
    'H': 'fecha_fin',
    'I': 'hora_fin',
    'J': 'duracion_str',
    'K': 'carga',
    'L': 'frec',
    'M': 'horas',
    'N': 'tti_cto',
    'O': 'senal',
    'P': 'causa',
    'Q': 'sub_causa',
    'R': 'observacion',
    'S': 'sectores',
    'T': 'ciudad',
    'U': 'kva',
}

ESTADOS_DB = None
CAUSAS_DB = None


def _load_caches():
    global ESTADOS_DB, CAUSAS_DB
    if not ESTADOS_DB:
        ESTADOS_DB = query("SELECT state_code, state_name FROM common.states ORDER BY state_name") or []
    if not CAUSAS_DB:
        CAUSAS_DB = query("SELECT causa_id, causa_codigo, causa_nombre FROM sctis.causa ORDER BY causa_nombre") or []
    return ESTADOS_DB, CAUSAS_DB


def procesar_causas_preview(causas_orig_list, causas_db):
    procesadas = []
    auto_count = 0
    pendientes_count = 0

    for item in (causas_orig_list or []):
        orig = (item.get('original') or '').strip()
        count = item.get('count', 0)
        sub_c = item.get('sub_causa') or ''
        norm_orig = normalizar_texto(orig)

        matched_id = None
        matched_nombre = None
        status = 'PENDIENTE_SCTIS'
        sugerencia_id = None
        sugerencia_nombre = None
        best_ratio = 0.0

        for c in (causas_db or []):
            norm_c = normalizar_texto(c['causa_nombre'])
            norm_code = normalizar_texto(c.get('causa_codigo', ''))

            if norm_orig == norm_c or norm_orig == norm_code:
                matched_id = c['causa_id']
                matched_nombre = c['causa_nombre']
                status = 'HOMOLOGADO'
                best_ratio = 1.0
                break

            ratio = difflib.SequenceMatcher(None, norm_orig, norm_c).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                sugerencia_id = c['causa_id']
                sugerencia_nombre = c['causa_nombre']

        if not matched_id and best_ratio >= 0.82:
            matched_id = sugerencia_id
            matched_nombre = sugerencia_nombre
            status = 'HOMOLOGADO'

        if matched_id:
            auto_count += 1
        else:
            pendientes_count += 1

        procesadas.append({
            'original': orig,
            'count': count,
            'sub_causa': sub_c,
            'status': status,
            'matched_id': matched_id,
            'matched_nombre': matched_nombre,
            'sugerencia_id': sugerencia_id,
            'sugerencia_nombre': sugerencia_nombre,
            'confianza': round(best_ratio * 100)
        })

    return procesadas, auto_count, pendientes_count


def detectar_activos_inconsistentes(records, estado_codigo):
    if not estado_codigo or not records:
        return {'subestaciones': [], 'circuitos': []}

    subestaciones_unicas = {}
    circuitos_unicos = {}
    for rec in records:
        se = (rec.get('subestacion') or '').strip()
        ci = (rec.get('circuito') or '').strip()
        if se:
            if se not in subestaciones_unicas:
                subestaciones_unicas[se] = {'count': 0, 'circuitos': set()}
            subestaciones_unicas[se]['count'] += 1
            if ci:
                subestaciones_unicas[se]['circuitos'].add(ci)
        if ci and se:
            key = f'{ci}|{se}'
            if key not in circuitos_unicos:
                circuitos_unicos[key] = {'circuito': ci, 'subestacion': se, 'count': 0}
            circuitos_unicos[key]['count'] += 1

    # Aliases aprendidos del estado (resolución automática de selecciones previas)
    aliases = cargar_aliases_assets(estado_codigo)
    aliases_se = aliases['subestaciones']
    aliases_ci = aliases['circuitos']

    # Catalogo de subestaciones del estado
    cat_se = query("""
        SELECT asset_id, 
               COALESCE(asset_name_normalizado, asset_name) AS asset_name,
               asset_name AS asset_name_original,
               asset_code
        FROM common.assets
        WHERE asset_type = 'SUBSTATION' AND state_code = %s AND is_active = true
        ORDER BY asset_name
    """, (estado_codigo,))

    inconsistentes_se = []
    for nombre_se, info in subestaciones_unicas.items():
        match = None
        norm_se = normalizar_para_matching(nombre_se)
        if norm_se in aliases_se:
            continue
        for a in cat_se:
            norm_a = normalizar_para_matching(a['asset_name'])
            norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
            norm_code = normalizar_para_matching(a.get('asset_code', ''))
            if norm_a == norm_se or norm_orig == norm_se or norm_code == norm_se:
                match = a
                break
        if not match:
            for a in cat_se:
                norm_a = normalizar_para_matching(a['asset_name'])
                if norm_se in norm_a or norm_a in norm_se:
                    match = a
                    break
        if not match:
            clasif, sug_id = preclasificar_activo(nombre_se, cat_se)
            inconsistentes_se.append({
                'original': nombre_se,
                'rows': info['count'],
                'sugerencias': [{'asset_id': a['asset_id'], 'asset_name': a['asset_name']} for a in cat_se],
                'mapped_to': None,
                'clasificacion': clasif,
                'sugerencia_alias': sug_id,
            })

    inconsistentes_ci = []
    if inconsistentes_se:
        cat_ci = query("""
            SELECT a.asset_id, 
                   COALESCE(a.asset_name_normalizado, a.asset_name) AS asset_name,
                   a.asset_name AS asset_name_original,
                   a.elemento_tipo,
                   a.elemento_codigo,
                   a.parent_asset_id
            FROM common.assets a
            JOIN common.assets p ON p.asset_id = a.parent_asset_id
            WHERE a.asset_type = 'CIRCUITO' AND p.state_code = %s AND a.is_active = true
            ORDER BY a.asset_name
        """, (estado_codigo,))
    else:
        cat_ci = []

    for key, info in circuitos_unicos.items():
        se_original = info['subestacion']
        norm_se = normalizar_para_matching(se_original)
        se_match = None
        for a in cat_se:
            norm_a = normalizar_para_matching(a['asset_name'])
            norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
            if norm_a == norm_se or norm_orig == norm_se:
                se_match = a
                break
        ci_nombre = info['circuito']
        norm_ci = normalizar_para_matching(ci_nombre)
        if (norm_ci, normalizar_para_matching(se_original)) in aliases_ci:
            continue
        match = None
        if se_match:
            for a in cat_ci:
                if a['parent_asset_id'] == se_match['asset_id']:
                    norm_a = normalizar_para_matching(a['asset_name'])
                    norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
                    if norm_a == norm_ci or norm_orig == norm_ci or norm_ci in norm_a or norm_a in norm_ci:
                        match = a
                        break
        if not match:
            parent_id = se_match['asset_id'] if se_match else None
            sugerencias_ci = [a for a in cat_ci if parent_id is None or a['parent_asset_id'] == parent_id] if cat_ci else []
            clasif, sug_id = preclasificar_activo(ci_nombre, sugerencias_ci if sugerencias_ci else cat_ci)
            inconsistentes_ci.append({
                'original': ci_nombre,
                'subestacion_original': se_original,
                'rows': info['count'],
                'sugerencias': [{'asset_id': a['asset_id'], 'asset_name': a['asset_name']} for a in sugerencias_ci],
                'mapped_to': None,
                'clasificacion': clasif,
                'sugerencia_alias': sug_id,
            })

    return {'subestaciones': inconsistentes_se, 'circuitos': inconsistentes_ci}


def parse_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return float(value) if not isinstance(value, bool) else value
    s = str(value).strip()
    return s if s else None


def parse_hora(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', s):
        return s
    return s


def completar_fecha_hora(fecha_str, hora_str):
    """Combina una fecha con un valor que puede ser solo hora."""
    if not fecha_str or not hora_str:
        return hora_str or None
    try:
        fecha = datetime.fromisoformat(fecha_str).date()
    except (ValueError, TypeError):
        try:
            fecha = datetime.strptime(str(fecha_str).strip(), '%d/%m/%Y').date()
        except (ValueError, TypeError):
            return hora_str

    hora = str(hora_str).strip()
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', hora):
        parts = hora.split(':')
        h, m = int(parts[0]), int(parts[1])
        s = int(parts[2]) if len(parts) > 2 else 0
        dt = datetime(fecha.year, fecha.month, fecha.day, h, m, s)
        return dt.isoformat()

    try:
        dt = datetime.fromisoformat(hora)
        if dt.year == 1900:
            dt = dt.replace(year=fecha.year, month=fecha.month, day=fecha.day)
        return dt.isoformat()
    except (ValueError, TypeError):
        return hora_str


def combinar_fecha_hora(record):
    """Combina los pares fecha_inicio+hora_inicio y fecha_fin+hora_fin
    en timestamps, y deriva fecha_falla desde la fecha de inicio."""
    fecha_ini = record.get('fecha_inicio')
    hora_ini = record.get('hora_inicio')
    if fecha_ini and hora_ini:
        record['fecha_inicio'] = completar_fecha_hora(fecha_ini, hora_ini)
    else:
        record['fecha_inicio'] = hora_ini or fecha_ini

    fecha_fin = record.get('fecha_fin')
    hora_fin = record.get('hora_fin')
    if fecha_fin and hora_fin:
        record['fecha_fin'] = completar_fecha_hora(fecha_fin, hora_fin)
    else:
        record['fecha_fin'] = hora_fin or fecha_fin

    if record.get('fecha_inicio') and not record.get('fecha_falla'):
        try:
            dt = datetime.fromisoformat(str(record['fecha_inicio']))
            record['fecha_falla'] = dt.date().isoformat()
        except (ValueError, TypeError):
            try:
                record['fecha_falla'] = datetime.strptime(str(record['fecha_inicio']).strip(), '%d/%m/%Y').date().isoformat()
            except (ValueError, TypeError):
                pass
    return record


def detectar_estado_desde_texto(texto, estados_db):
    if not texto:
        return None, None
    t = normalizar_texto(str(texto).strip())
    for st in estados_db:
        if t == normalizar_texto(st['state_name']):
            return st['state_code'], st['state_name']
    return None, None


# ─── Parseo del FORMATO ESTABLECIDO (primera hoja) ────────────

def parse_formato(ws, estados_db, causas_db):
    records = []
    causas_originales = {}
    max_row = ws.max_row or 1

    for row_idx in range(2, min(max_row + 1, MAX_ROWS + 1)):
        cells = ws[row_idx]
        a_val = cells[0].value if len(cells) > 0 else None
        if a_val is None or str(a_val).strip() == '':
            continue

        record = {}
        for col_idx in range(22):
            col_letter = chr(ord('A') + col_idx)
            field = COL_MAP_FORMATO.get(col_letter)
            if field is None:
                continue
            val = parse_cell(cells[col_idx].value) if col_idx < len(cells) else None
            record[field] = val

        causa_orig = (record.get('causa') or '').strip()
        if causa_orig:
            key = causa_orig.upper()
            if key not in causas_originales:
                causas_originales[key] = {
                    'original': causa_orig,
                    'sub_causa': record.get('sub_causa'),
                    'count': 0,
                }
            causas_originales[key]['count'] += 1

        combinar_fecha_hora(record)

        records.append(record)

    return {
        'ok': True,
        'formato': 'formato',
        'total_rows': len(records),
        'causas_originales': list(causas_originales.values()),
        'causas_homologadas': causas_db,
        'preview': records[:20],
        '_all_records': records,
    }


# ─── Validación de estructura contra plantilla oficial ─────────

HEADERS_FORMATO_ESPERADOS = [
    (0, 'ESTADO'),
    (1, 'SISTEMA'),
    (4, 'CIRCUITO'),
    (5, 'FECHA'),
    (15, 'CAUSA'),
    (16, 'SUB'),
    (17, 'OBSERVACION'),
]


def normalizar_texto(s):
    """Elimina acentos y diéresis para comparación."""
    import unicodedata
    return ''.join(
        c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c)
    ).upper()


def validar_estructura(ws):
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not headers or not headers[0]:
        return False, 'El archivo no tiene fila de encabezados.'

    row = headers[0]
    total_cols = ws.max_column or sum(1 for c in row if c is not None and str(c).strip())
    if total_cols < 19:
        return False, (
            'El archivo no coincide con la plantilla oficial. '
            'Debe usar el formulario de carga Excel homologado proporcionado por la Coordinación de SCTIS. '
            f'Se detectaron {total_cols} columnas (se esperan al menos 19).'
        )

    row_strs = [normalizar_texto(str(c)) if c else '' for c in row[:30]]

    if 'ESTADO' not in row_strs[0]:
        return False, (
            'Formato de archivo no reconocido. '
            'Debe usar el formulario de carga Excel homologado '
            '(primera pestaña del archivo proporcionado por la Coordinación de SCTIS). '
            f'Encabezado detectado: "{row_strs[0][:50] if row_strs[0] else "(vacío)"}".'
        )

    problemas = []
    for col_idx, kw in HEADERS_FORMATO_ESPERADOS:
        actual = row_strs[col_idx] if col_idx < len(row_strs) else ''
        if kw not in actual and actual:
            problemas.append(f'Columna {chr(65+col_idx)}: se esperaba contener "{kw}", se encontró "{actual}"')
    if problemas:
        return False, (
            'La estructura del archivo no coincide con la plantilla oficial del Formato Establecido. '
            'Debe usar el formulario de carga Excel homologado proporcionado por la Coordinación de SCTIS. '
            'Detalles: ' + '; '.join(problemas[:3])
        )

    return True, ''


# ─── Detección de formato desde catálogo ──────────────────────

def detectar_formato_catalogo(ws, estado_codigo=None):
    """
    Detecta el formato del Excel comparando headers con el catálogo.
    Retorna (formato_dict, None) si match, o (None, error_msg) si no.
    """
    max_header_row = min(5, ws.max_row or 1)
    headers_todos = []
    for row_num in range(1, max_header_row + 1):
        row = [str(c.value).strip().upper() if c.value else '' for c in ws[row_num]]
        headers_todos.extend(row)

    headers_str = ' '.join([h for h in headers_todos if h])

    where = "activo = true"
    params = []
    if estado_codigo:
        where += " AND (estado_codigo = %s OR estado_codigo IS NULL)"
        params.append(estado_codigo)
    else:
        where += " AND estado_codigo IS NULL"

    formatos = query(f"""
        SELECT formato_id, formato_codigo, formato_nombre, header_keywords,
               header_row, data_start_row, mapeo_columnas, reglas, campos_faltantes
        FROM sctis.formato_catalogo
        WHERE {where}
        ORDER BY
            CASE WHEN estado_codigo IS NOT DISTINCT FROM %s THEN 0 ELSE 1 END,
            array_length(header_keywords, 1) DESC
    """, params + [estado_codigo])

    for fmt in (formatos or []):
        keywords = fmt['header_keywords']
        match_count = sum(1 for kw in keywords if kw.upper() in headers_str)
        match_ratio = match_count / len(keywords) if keywords else 0
        if match_ratio >= 0.6:
            return fmt, None

    return None, (
        'Formato no reconocido. El sistema detectó que el archivo no coincide '
        'con ningún formato predefinido del catálogo ni con el formato CTIS homologado. '
        'Verifique que el archivo esté en el formato correcto de su estado.'
    )


def parse_formato_catalogo(ws, formato, estado_codigo):
    """
    Parsea un Excel usando el mapeo del catálogo de formatos.
    Retorna records normalizados al formato CTIS.
    """
    mapeo = formato['mapeo_columnas']
    reglas = formato.get('reglas', {})
    data_start = formato.get('data_start_row', 2)
    max_row = ws.max_row or 1

    records = []
    causas_originales = {}

    for row_idx in range(data_start, min(max_row + 1, 2002)):
        cells = ws[row_idx]
        if not cells or cells[0].value is None:
            continue

        record = {}
        for col_idx in range(min(len(cells), 60)):
            cell_val = cells[col_idx].value
            if cell_val is None:
                continue
            col_letter = chr(65 + col_idx) if col_idx < 26 else f'{chr(64 + col_idx // 26)}{chr(65 + col_idx % 26)}'

            header_val = None
            header_row_idx = formato.get('header_row', 1) - 1
            if header_row_idx < len(list(ws.iter_rows(min_row=1, max_row=1))):
                pass
            for hrow in range(1, min(data_start, 6)):
                hcells = ws[hrow]
                if col_idx < len(hcells) and hcells[col_idx].value:
                    header_val = str(hcells[col_idx].value).strip().upper()
                    break

            if not header_val:
                continue

            mapping = None
            for header_key, map_info in mapeo.items():
                if header_key.upper() == header_val or header_key.upper() in header_val or header_val in header_key.upper():
                    mapping = map_info
                    break

            if mapping is None or mapping.get('tipo') == 'ignore' or mapping.get('target') is None:
                continue

            target = mapping['target']
            tipo = mapping.get('tipo', 'texto')

            if tipo == 'texto':
                record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'numerico':
                try:
                    record[target] = float(cell_val) if cell_val else None
                except (ValueError, TypeError):
                    record[target] = None
            elif tipo == 'fecha':
                if isinstance(cell_val, (datetime, date)):
                    record[target] = cell_val.isoformat() if isinstance(cell_val, date) else cell_val.isoformat()
                else:
                    record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'hora':
                if isinstance(cell_val, (datetime, date)):
                    record[target] = cell_val.strftime('%H:%M:%S') if hasattr(cell_val, 'strftime') else str(cell_val)
                else:
                    record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'fecha_hora':
                if isinstance(cell_val, (datetime, date)):
                    record[target] = cell_val.isoformat()
                else:
                    record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'duracion_a_horas':
                record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'minutos_a_horas':
                try:
                    mins = float(cell_val)
                    record[target] = str(mins / 60) if mins else None
                except (ValueError, TypeError):
                    record[target] = str(cell_val).strip() if cell_val else None
            elif tipo == 'mw_a_kva':
                try:
                    mw = float(cell_val)
                    record[target] = str(mw * 1000) if mw else None
                except (ValueError, TypeError):
                    record[target] = None
            elif tipo == 'mvamin_a_kva':
                record[target] = str(cell_val).strip() if cell_val else None
            else:
                record[target] = str(cell_val).strip() if cell_val else None

        if not record.get('subestacion') and not record.get('circuito'):
            continue

        for key, val in record.items():
            if val == '' or val == 'None':
                record[key] = None

        for target_field, default_rule in reglas.items():
            if default_rule.startswith('default='):
                default_val = default_rule.split('=', 1)[1]
                if record.get(target_field) is None:
                    record[target_field] = None if default_val == 'NULL' else default_val
            elif default_rule == 'derivar_de_fecha_inicio':
                if record.get('fecha_inicio') and not record.get('fecha_falla'):
                    try:
                        fi = record['fecha_inicio']
                        if isinstance(fi, str):
                            record['fecha_falla'] = fi[:10]
                        elif isinstance(fi, (date, datetime)):
                            record['fecha_falla'] = fi.isoformat()[:10]
                    except Exception:
                        pass

        causa_orig = (record.get('causa') or '').strip()
        if causa_orig:
            key = causa_orig.upper()
            if key not in causas_originales:
                causas_originales[key] = {
                    'original': causa_orig,
                    'sub_causa': record.get('sub_causa'),
                    'count': 0,
                }
            causas_originales[key]['count'] += 1

        records.append(record)

    return {
        'ok': True,
        'formato': formato['formato_codigo'],
        'formato_nombre': formato['formato_nombre'],
        'total_rows': len(records),
        'causas_originales': list(causas_originales.values()),
        'campos_faltantes': formato.get('campos_faltantes', []),
        'preview': records[:20],
        '_all_records': records,
    }


def generar_excel_correccion(records_errores, token, estado_codigo):
    """
    Genera un Excel con los registros rechazados para que el usuario los corrija.
    Retorna la ruta del archivo generado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros para Corrección'

    headers = ['Fila Original', 'Subestación', 'Circuito', 'Fecha Inicio',
               'Causa', 'Error', 'Observación']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for rec in records_errores:
        ws.append([
            rec.get('row_numero', ''),
            rec.get('subestacion', ''),
            rec.get('circuito', ''),
            rec.get('fecha_inicio', ''),
            rec.get('causa', ''),
            rec.get('error', ''),
            rec.get('observacion', ''),
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    filename = f'correccion_{estado_codigo}_{token[:8]}.xlsx'
    filepath = os.path.join(UPLOAD_DIR, filename)
    wb.save(filepath)
    return filepath


def sheet_oculta(filepath, sheet_name):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        oculta = wb[sheet_name].sheet_state != 'visible'
        wb.close()
        return oculta
    except Exception:
        return False


def listar_hojas_visibles(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        visibles = [s for s in wb.sheetnames if wb[s].sheet_state == 'visible']
        wb.close()
        return visibles
    except Exception:
        return []


class _Celda:
    __slots__ = ('value',)

    def __init__(self, value):
        self.value = value


class _Fila:
    __slots__ = ('_celdas',)

    def __init__(self, celdas):
        self._celdas = celdas

    def __getitem__(self, idx):
        if idx < len(self._celdas):
            return self._celdas[idx]
        return _Celda(None)

    def __len__(self):
        return len(self._celdas)

    def __iter__(self):
        return iter(self._celdas)


class HojaMemoria:
    """
    Wrapper sobre una hoja de openpyxl cargada en modo read_only.
    Permite acceso por fila (ws[row_idx]) y atributos max_row/max_column,
    que read_only no soporta de forma nativa.
    """

    def __init__(self, ws):
        self.title = ws.title
        self._filas = []
        self.max_row = 0
        self.max_column = 0
        for row in ws.iter_rows():
            celdas = [_Celda(c.value) for c in row]
            self._filas.append(_Fila(celdas))
            self.max_column = max(self.max_column, len(celdas))
        self.max_row = len(self._filas)

    def __getitem__(self, idx):
        if 1 <= idx <= len(self._filas):
            return self._filas[idx - 1]
        return _Fila([])

    def iter_rows(self, min_row=1, max_row=None, values_only=False):
        max_row = max_row or self.max_row
        for i in range(min_row, max_row + 1):
            fila = self[i]
            if values_only:
                yield tuple(c.value for c in fila)
            else:
                yield fila


def cargar_hoja_memoria(filepath, sheet_name=None):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        return HojaMemoria(ws)
    finally:
        wb.close()


def ext_from_token(token):
    for f in os.listdir(UPLOAD_DIR):
        if f.startswith(token) and not f.endswith('_preview.json'):
            return os.path.splitext(f)[1]
    return '.xlsx'


def parse_excel(filepath, sheet_name=None, estado_codigo=None):
    estados_db, causas_db = _load_caches()
    ws = cargar_hoja_memoria(filepath, sheet_name)
    if ws is None:
        return {'error': 'hoja_invalida', 'mensaje': 'No se pudo determinar la hoja a procesar.', 'ok': False}

    valido, msg = validar_estructura(ws)
    if not valido:
        user_estado = estado_codigo or (session.get('user', {}).get('estado_codigo') if 'user' in dir(session) else None)
        fmt, fmt_msg = detectar_formato_catalogo(ws, user_estado)
        if fmt:
            result = parse_formato_catalogo(ws, fmt, user_estado or '')
            result['formato_detectado'] = fmt['formato_codigo']
            result['formato_nombre'] = fmt['formato_nombre']
            result['es_formato_catalogo'] = True
            result['sheet_seleccionada'] = ws.title
            return result
        return {
            'error': 'formato_invalido',
            'mensaje': msg,
            'formato_sugerencia': fmt_msg,
            'ok': False,
        }

    result = parse_formato(ws, estados_db, causas_db)
    result['es_formato_catalogo'] = False
    result['sheet_seleccionada'] = ws.title

    return result


# ─── Rutas ─────────────────────────────────────────────────────

@bp.route('/importar', methods=['GET'])
@login_required
def importar_page():
    user_estado = g.user.get('estado_codigo')
    if g.is_admin:
        estados = query("SELECT state_code, state_name FROM common.states ORDER BY state_name")
    else:
        estados = query("SELECT state_code, state_name FROM common.states WHERE state_code = %s ORDER BY state_name", (user_estado,))
    return render_template('importar.html', estados=estados, user_estado=user_estado)


@bp.route('/api/importar/preview', methods=['POST'])
def preview_import():
    file = request.files.get('file')
    token = request.form.get('token') or session.get('import_token')
    sheet_name = request.form.get('sheet_name') or None

    if not file and not token:
        return jsonify({'error': 'No se envi\xf3 ning\xfan archivo'}), 400

    os.makedirs(UPLOAD_DIR, exist_ok=True)

    # ── Primer paso: subida del archivo ──
    if file and not sheet_name:
        if file.filename == '':
            return jsonify({'error': 'Nombre de archivo vac\xedo'}), 400

        ext = os.path.splitext(file.filename)[1] or '.xlsx'
        token = uuid.uuid4().hex
        tmp_path = os.path.join(UPLOAD_DIR, f'{token}{ext}')
        file.save(tmp_path)
        session['import_token'] = token

        try:
            wb_sheets = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb_sheets.sheetnames
            wb_sheets.close()
        except Exception:
            os.remove(tmp_path)
            return jsonify({'error': 'El archivo no es un Excel v\xe1lido', 'ok': False}), 400

        visibles = listar_hojas_visibles(tmp_path)
        if len(visibles) > 1:
            return jsonify({
                'ok': True,
                'requires_sheet_selection': True,
                'sheet_names': visibles,
                'sheet_count': len(visibles),
                'token': token,
            })

        sheet_name = visibles[0] if visibles else sheet_names[0]

    if not token:
        return jsonify({'error': 'Sesi\xf3n expirada. Vuelva a subir el archivo.'}), 400

    tmp_path = os.path.join(UPLOAD_DIR, f'{token}{ext_from_token(token)}')
    if not os.path.exists(tmp_path):
        return jsonify({'error': 'Archivo no encontrado. Vuelva a subir el archivo.'}), 400

    if not sheet_name:
        try:
            wb_sheets = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_name = wb_sheets.active.title
            wb_sheets.close()
        except Exception:
            sheet_name = None

    user_estado = session.get('user', {}).get('estado_codigo')
    user_username = session.get('user', {}).get('username', 'unknown')
    form_estado = request.form.get('estado_codigo')
    estado = user_estado or form_estado

    result = parse_excel(tmp_path, sheet_name, estado)
    if 'error' in result:
        status = 422 if result.get('error') == 'formato_invalido' else 400
        return jsonify(result), status

    if user_estado:
        result['estado_codigo'] = user_estado
    elif form_estado:
        result['estado_codigo'] = form_estado

    try:
        wb_sheets = openpyxl.load_workbook(tmp_path, read_only=True)
        result['sheet_names'] = wb_sheets.sheetnames
        result['sheet_count'] = len(wb_sheets.sheetnames)
        wb_sheets.close()
    except Exception:
        result['sheet_names'] = []
        result['sheet_count'] = 0

    submission_id = None
    try:
        sub_result = query("""
            INSERT INTO audit.submissions
                (process_code, state_code, source_filename, source_sheet,
                 sheet_names, row_count, validation_status, formato_codigo,
                 ingested_by, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING submission_id
        """, (
            'sctis_import',
            estado,
            os.path.basename(tmp_path),
            result.get('sheet_seleccionada') or sheet_name,
            result.get('sheet_names', []),
            result.get('total_rows', 0),
            'PENDING',
            result.get('formato_detectado'),
            user_username,
            f"Formato: {result.get('formato_detectado', 'CTIS_HOMOLOGADO')}",
        ), fetch=True)
        if sub_result and len(sub_result) > 0:
            submission_id = sub_result[0]['submission_id']
    except Exception:
        pass

    result['submission_id'] = submission_id

    records = result.get('_all_records') or []
    incompletos_count = sum(1 for r in records if es_registro_incompleto(r))
    user_role = session.get('role_code') or session.get('user', {}).get('role_code')
    is_admin = (user_role == 'admin')

    result['registros_incompletos'] = incompletos_count
    result['alerta_incompletos'] = (incompletos_count > 0)
    result['es_admin'] = is_admin
    result['bloqueado_no_admin'] = (incompletos_count > 0 and not is_admin)

    if records and estado:
        result['activos_inconsistentes'] = detectar_activos_inconsistentes(records, estado)
    else:
        result['activos_inconsistentes'] = {'subestaciones': [], 'circuitos': []}
    result.pop('_all_records', None)

    # Process and classify causes
    _, causas_db = _load_caches()
    causas_procesadas, auto_count, pendientes_count = procesar_causas_preview(result.get('causas_originales', []), causas_db)
    result['causas_originales'] = causas_procesadas
    result['causas_homologadas'] = causas_db
    result['auto_homologadas_count'] = auto_count
    result['pendientes_count'] = pendientes_count

    preview_path = os.path.join(UPLOAD_DIR, f'{token}_preview.json')
    with open(preview_path, 'w') as f:
        json.dump(result, f, default=str)

    session['import_token'] = token
    return jsonify(result)



@bp.route('/api/importar/confirmar', methods=['POST'])
def confirmar_import():
    data = request.get_json()
    if not data or 'mapping' not in data:
        return jsonify({'error': 'Se requiere el mapeo de causas'}), 400

    user_estado = session.get('user', {}).get('estado_codigo')
    estado_codigo = data.get('estado_codigo') or user_estado
    if not estado_codigo:
        return jsonify({'error': 'Debe seleccionar un estado destino para la importación.'}), 400

    token = session.get('import_token')
    if not token:
        return jsonify({'error': 'Sesi\xf3n expirada. Vuelva a subir el archivo.'}), 400

    tmp_path = os.path.join(UPLOAD_DIR, f'{token}.xlsx')
    preview_path = os.path.join(UPLOAD_DIR, f'{token}_preview.json')

    if not os.path.exists(tmp_path):
        session.pop('import_token', None)
        return jsonify({'error': 'Archivo no encontrado. Vuelva a subir el archivo.'}), 400

    if os.path.exists(preview_path):
        with open(preview_path) as f:
            result = json.load(f)
    else:
        result = {}

    mapping = data['mapping']
    se_map = data.get('subestacion_map') or {}
    ci_map = data.get('circuito_map') or {}
    se_nuevos = data.get('se_nuevos') or []
    ci_nuevos = data.get('ci_nuevos') or []
    mes_override = data.get('mes') or result.get('mes')

    usuario_alias = session.get('user', {}).get('username')
    aliases = cargar_aliases_assets(estado_codigo)
    aliases_se = aliases['subestaciones']
    aliases_ci = aliases['circuitos']
    guardar_aliases_assets(se_map, ci_map, estado_codigo, usuario_alias)

    se_nuevos_set = set(str(n).strip() for n in se_nuevos if n and str(n).strip())
    ci_nuevos_set = set(
        f'{str(c.get("circuito")).strip()}|{str(c.get("subestacion")).strip()}'
        for c in ci_nuevos if c and c.get('circuito') and str(c.get('circuito')).strip()
    )
    filas_se = {}
    filas_ci = {}

    sheet_name = result.get('sheet_seleccionada')
    ws = cargar_hoja_memoria(tmp_path, sheet_name)

    col_map = COL_MAP_FORMATO
    start_row = 2
    max_row = ws.max_row or start_row
    insertados = 0
    errores = []

    # Resolver nombres catalogados
    resolver_asset = {}
    for se_orig, asset_id in se_map.items():
        if asset_id:
            a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name, asset_code FROM common.assets WHERE asset_id = %s", (asset_id,))
            if a:
                resolver_asset[se_orig] = {'asset_id': asset_id, 'asset_name': a['asset_name']}

    resolver_circuito = {}
    for ci_key, asset_id in ci_map.items():
        if asset_id:
            a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (asset_id,))
            if a:
                resolver_circuito[ci_key] = {'asset_id': asset_id, 'asset_name': a['asset_name']}

    # Pre-parse and validate records for incomplete dates/times
    records_cache = []
    incompletos_count = 0
    for row_idx in range(start_row, max_row + 1):
        cells = ws[row_idx]
        a_val = cells[0].value if len(cells) > 0 else None
        if a_val is None or str(a_val).strip() == '':
            continue

        record = {}
        for col_letter, field in col_map.items():
            if field is None:
                continue
            col_idx = ord(col_letter) - ord('A')
            record[field] = parse_cell(cells[col_idx].value) if col_idx < len(cells) else None

        raw_f_fin = record.get('fecha_fin')
        raw_h_fin = record.get('hora_fin')
        combinar_fecha_hora(record)
        is_inc = es_registro_incompleto(record) or not raw_f_fin or not raw_h_fin
        record['_es_incompleto'] = is_inc
        if is_inc:
            incompletos_count += 1
        records_cache.append((row_idx, record))

    user_role = session.get('role_code') or session.get('user', {}).get('role_code')
    is_admin = (user_role == 'admin')

    if incompletos_count > 0 and not is_admin:
        return jsonify({
            'error': f'Carga bloqueada por cumplimiento de norma: El archivo contiene {incompletos_count} registro(s) incompletos (sin fecha u hora de cierre). Solamente un Administrador del Sistema posee privilegios para autorizar la carga de datos incompletos.'
        }), 403

    audit_id = None
    if incompletos_count > 0 and is_admin:
        if not data.get('aceptar_no_repudio'):
            return jsonify({
                'error': f'Declaración de No Repudio requerida: Debe aceptar formalmente la Declaración de No Repudio y Responsabilidad Administrativa para autorizar la carga de {incompletos_count} registro(s) incompletos.'
            }), 400

        user_id = session.get('user_id') or session.get('user', {}).get('user_id')
        username = session.get('user', {}).get('username') or session.get('username') or 'admin'
        declaracion = data.get('declaracion_no_repudio') or f'Carga de data fuera de norma autorizada por {username} con {incompletos_count} registros incompletos.'
        ip_addr = request.remote_addr or '127.0.0.1'
        user_agent = request.headers.get('User-Agent', '')

        audit_id = insert_audit_carga_excepcional(
            user_id, username, estado_codigo,
            os.path.basename(tmp_path), token,
            len(records_cache), incompletos_count,
            declaracion, ip_addr, user_agent
        )

    # Catalogo fuzzy para fallback de matching
    cat_se_fuzzy = query("""
        SELECT asset_id, 
               COALESCE(asset_name_normalizado, asset_name) AS asset_name,
               asset_name AS asset_name_original
        FROM common.assets
        WHERE asset_type = 'SUBSTATION' AND state_code = %s AND is_active = true
    """, (estado_codigo,))
    cat_se_fuzzy_map = {}
    for a in (cat_se_fuzzy or []):
        cat_se_fuzzy_map[normalizar_para_matching(a['asset_name'])] = a
        if a.get('asset_name_original'):
            cat_se_fuzzy_map[normalizar_para_matching(a['asset_name_original'])] = a

    for row_idx, record in records_cache:
        se_orig_parsed = (record.get('subestacion') or '').strip()
        ci_orig_parsed = (record.get('circuito') or '').strip()
        if se_orig_parsed:
            filas_se[se_orig_parsed] = filas_se.get(se_orig_parsed, 0) + 1
        if se_orig_parsed and ci_orig_parsed:
            filas_ci[f'{ci_orig_parsed}|{se_orig_parsed}'] = filas_ci.get(f'{ci_orig_parsed}|{se_orig_parsed}', 0) + 1

        causa_raw = (record.get('causa') or '').strip()
        causa_orig_upper = causa_raw.upper()
        causa_id = mapping.get(causa_raw) or mapping.get(causa_orig_upper) or mapping.get(str(causa_raw))
        if isinstance(causa_id, str) and causa_id.isdigit():
            causa_id = int(causa_id)

        if not causa_id and causa_raw:
            _, causas_db = _load_caches()
            norm_raw = normalizar_texto(causa_raw)
            for c in causas_db:
                if norm_raw == normalizar_texto(c['causa_nombre']) or norm_raw == normalizar_texto(c.get('causa_codigo', '')):
                    causa_id = c['causa_id']
                    break
            if not causa_id:
                otros = next((c for c in causas_db if c['causa_codigo'] == 'OTROS' or 'OTROS' in c['causa_nombre'].upper()), None)
                if otros:
                    causa_id = otros['causa_id']

        try:
            estado = estado_codigo

            if not estado:
                errores.append(f'Fila {row_idx}: no se pudo determinar el estado')
                continue

            # Resolver subestacion
            se_orig = (record.get('subestacion') or '').strip()
            if se_orig and se_orig in se_nuevos_set:
                record['subestacion_id'] = None
            elif se_orig and se_orig in resolver_asset:
                record['subestacion'] = resolver_asset[se_orig]['asset_name']
                record['subestacion_id'] = resolver_asset[se_orig]['asset_id']
            elif se_orig:
                norm_se = normalizar_para_matching(se_orig)
                alias_id = aliases_se.get(norm_se)
                if alias_id:
                    a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name, asset_code FROM common.assets WHERE asset_id = %s", (alias_id,))
                    if a:
                        record['subestacion'] = a['asset_name']
                        record['subestacion_id'] = alias_id
                elif norm_se in cat_se_fuzzy_map:
                    a = cat_se_fuzzy_map[norm_se]
                    record['subestacion'] = a['asset_name']
                    record['subestacion_id'] = a['asset_id']
                else:
                    for norm_key, a in cat_se_fuzzy_map.items():
                        if norm_se in norm_key or norm_key in norm_se:
                            record['subestacion'] = a['asset_name']
                            record['subestacion_id'] = a['asset_id']
                            break

            # Resolver circuito
            ci_orig = (record.get('circuito') or '').strip()
            ci_key = f'{ci_orig}|{se_orig}'
            if ci_orig and ci_key in ci_nuevos_set:
                record['circuito_id'] = None
            elif ci_orig and ci_key in resolver_circuito:
                record['circuito'] = resolver_circuito[ci_key]['asset_name']
                record['circuito_id'] = resolver_circuito[ci_key]['asset_id']
            elif ci_orig and record.get('subestacion_id'):
                norm_ci = normalizar_para_matching(ci_orig)
                alias_id = aliases_ci.get((norm_ci, normalizar_para_matching(se_orig)))
                if alias_id:
                    a = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (alias_id,))
                    if a:
                        record['circuito'] = a['asset_name']
                        record['circuito_id'] = alias_id
                else:
                    cat_ci_fallback = query("""
                        SELECT asset_id, COALESCE(asset_name_normalizado, asset_name) AS asset_name,
                               asset_name AS asset_name_original
                        FROM common.assets
                        WHERE asset_type = 'CIRCUITO' AND parent_asset_id = %s AND is_active = true
                    """, (record['subestacion_id'],))
                    for a in (cat_ci_fallback or []):
                        norm_a = normalizar_para_matching(a['asset_name'])
                        norm_orig = normalizar_para_matching(a.get('asset_name_original', ''))
                        if norm_a == norm_ci or norm_orig == norm_ci or norm_ci in norm_a or norm_a in norm_ci:
                            record['circuito'] = a['asset_name']
                            record['circuito_id'] = a['asset_id']
                            break

            desp = record.get('despachador')
            desp_id = None
            if desp:
                desp_id = asegurar_despachador(str(desp).strip(), estado)
                if desp_id:
                    record['despachador'] = normalizar_nombre(str(desp))

            f_falla = record.get('fecha_falla')
            if f_falla and isinstance(f_falla, str):
                try:
                    f_falla = datetime.fromisoformat(f_falla).date()
                except Exception:
                    f_falla = None

            f_ini = record.get('fecha_inicio')
            if f_ini and isinstance(f_ini, str):
                try:
                    f_ini = datetime.fromisoformat(f_ini)
                except Exception:
                    f_ini = None

            f_fin = record.get('fecha_fin')
            if f_fin and isinstance(f_fin, str):
                try:
                    f_fin = datetime.fromisoformat(f_fin)
                except Exception:
                    f_fin = None

            is_inc = record.get('_es_incompleto', False)
            if is_inc or not f_fin:
                if not f_ini:
                    f_ini = datetime.now()
                f_fin = f_ini
                horas = 0.0
                estado_calc = 'INCOMPLETO_EXCEPCION_ADMIN'
                es_exc = 1
                audit_id_val = audit_id
            else:
                horas = record.get('horas')
                if horas is not None:
                    try:
                        horas = float(horas)
                    except Exception:
                        horas = 0.0
                else:
                    horas = 0.0
                estado_calc = record.get('estado_calculo') or 'CALCULO VALIDO'
                es_exc = 0
                audit_id_val = None

            kva = record.get('kva')
            if kva is not None:
                try:
                    kva = float(kva)
                except Exception:
                    kva = 0.0

            subestacion_val = record.get('subestacion')
            circuito_val = record.get('circuito')

            insert_sql = """
                INSERT INTO "sctis.tira_interrupcion"
                    (estado_codigo, fecha_falla, sistema, subestacion, subestacion_id,
                     circuito, circuito_id, jefatura,
                     fecha_inicio, fecha_fin, causa, sub_causa, observacion,
                     despachador, despachador_id, kva, horas, mes, sectores, ciudad,
                     causa_id, created_by, es_excepcion_admin, audit_id, estado_calculo)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            mes_val = mes_override or record.get('mes')
            query(insert_sql, (
                estado, f_falla, record.get('sistema'),
                subestacion_val, record.get('subestacion_id'),
                circuito_val, record.get('circuito_id'),
                record.get('jefatura'),
                f_ini, f_fin, record.get('causa'), record.get('sub_causa'),
                record.get('observacion'), record.get('despachador'), desp_id,
                kva, horas, mes_val,
                record.get('sectores'), record.get('ciudad'),
                causa_id, 'importacion_excel',
                es_exc, audit_id_val, estado_calc
            ), fetch=False)
            insertados += 1
        except Exception as e:
            errores.append(f'Fila {row_idx}: {str(e)}')

    # Encolar activos reportados como nuevos para revisión del administrador
    requests_creados = 0
    if se_nuevos_set or ci_nuevos_set:
        cat_ci_estado = query("""
            SELECT a.asset_id, COALESCE(a.asset_name_normalizado, a.asset_name) AS asset_name
            FROM common.assets a
            JOIN common.assets p ON p.asset_id = a.parent_asset_id
            WHERE a.asset_type = 'CIRCUITO' AND p.state_code = %s AND a.is_active = true
        """, (estado_codigo,)) or []
        se_request_ids = {}
        for se_n in se_nuevos_set:
            clasif, sug = preclasificar_activo(se_n, cat_se_fuzzy or [])
            try:
                res = query("""
                    INSERT INTO sctis.asset_request
                        (estado_codigo, asset_type, nombre_reportado, nombre_normalizado,
                         filas_afectadas, clasificacion, sugerencia_alias,
                         submission_id, requested_by)
                    VALUES (%s, 'SUBSTATION', %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (estado_codigo, asset_type, nombre_reportado, se_referencia)
                    DO UPDATE SET filas_afectadas = sctis.asset_request.filas_afectadas + EXCLUDED.filas_afectadas,
                                  updated_at = now()
                    RETURNING request_id
                """, (estado_codigo, se_n, normalizar_nombre(se_n),
                      filas_se.get(se_n, 0), clasif, sug, submission_id, usuario_alias))
                if res:
                    se_request_ids[se_n] = res[0]['request_id']
                    requests_creados += 1
            except Exception:
                pass
        for ci_n in ci_nuevos_set:
            ci, se = ci_n.split('|', 1)
            clasif, sug = preclasificar_activo(ci, cat_ci_estado)
            try:
                query("""
                    INSERT INTO sctis.asset_request
                        (estado_codigo, asset_type, nombre_reportado, nombre_normalizado,
                         se_referencia, se_request_id, filas_afectadas, clasificacion,
                         sugerencia_alias, submission_id, requested_by)
                    VALUES (%s, 'CIRCUITO', %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (estado_codigo, asset_type, nombre_reportado, se_referencia)
                    DO UPDATE SET filas_afectadas = sctis.asset_request.filas_afectadas + EXCLUDED.filas_afectadas,
                                  updated_at = now()
                """, (estado_codigo, ci, normalizar_nombre(ci), se,
                      se_request_ids.get(se), filas_ci.get(ci_n, 0), clasif, sug,
                      submission_id, usuario_alias), fetch=False)
                requests_creados += 1
            except Exception:
                pass

    submission_id = result.get('submission_id')
    total_intentados = result.get('total_rows', insertados + len(errores))
    rechazados = len(errores)

    if submission_id:
        try:
            query("""
                UPDATE audit.submissions SET
                    accepted_count = %s,
                    rejected_count = %s,
                    validation_status = %s
                WHERE submission_id = %s
            """, (insertados, rechazados,
                  'VALIDATED' if rechazados == 0 else 'PARTIAL',
                  submission_id), fetch=False)
        except Exception:
            pass

        if rechazados > 0:
            user_id = g.user.get('user_id') if hasattr(g, 'user') and g.user else None
            if not user_id:
                try:
                    u = query_one("SELECT user_id FROM sctis.user_profiles WHERE username = %s",
                                  (session.get('user', {}).get('username', ''),))
                    user_id = u['user_id'] if u else None
                except Exception:
                    pass

            if user_id:
                error_records = []
                for i, err in enumerate(errores[:50]):
                    parts = err.split(': ', 1)
                    row_num = parts[0].replace('Fila ', '') if parts else ''
                    error_msg = parts[1] if len(parts) > 1 else err
                    error_records.append({
                        'row_numero': row_num,
                        'error': error_msg,
                    })

                try:
                    corr_file = generar_excel_correccion(error_records, token, estado_codigo)
                    corr_filename = os.path.basename(corr_file)
                except Exception:
                    corr_filename = None

                try:
                    query("""
                        INSERT INTO sctis.tarea_pendiente
                            (usuario_id, estado_codigo, tipo_tarea, submission_id,
                             descripcion, registros_total, registros_ok, registros_rechazados,
                             archivo_correccion, estado_tarea)
                        VALUES (%s, %s, 'CORREGIR_DATOS', %s, %s, %s, %s, %s, %s, 'PENDIENTE')
                    """, (
                        user_id, estado_codigo, submission_id,
                        f'Carga con {rechazados} registros rechazados de {total_intentados} totales. '
                        f'Archivo: {result.get("source_filename", "desconocido")}',
                        total_intentados, insertados, rechazados,
                        corr_filename,
                    ), fetch=False)
                except Exception:
                    pass

                if corr_filename:
                    try:
                        query("""
                            UPDATE audit.submissions SET correction_file = %s
                            WHERE submission_id = %s
                        """, (corr_filename, submission_id), fetch=False)
                    except Exception:
                        pass

        if requests_creados > 0:
            admin = query_one("""
                SELECT up.user_id FROM sctis.user_profiles up
                JOIN sctis.user_roles r ON r.role_id = up.role_id
                WHERE r.role_code = 'admin' AND up.is_active = true
                ORDER BY up.user_id LIMIT 1
            """)
            if admin:
                try:
                    query("""
                        INSERT INTO sctis.tarea_pendiente
                            (usuario_id, estado_codigo, tipo_tarea, submission_id,
                             descripcion, registros_total, registros_ok, registros_rechazados,
                             estado_tarea)
                        VALUES (%s, %s, 'APROBAR_ACTIVO', %s, %s, %s, %s, %s, 'PENDIENTE')
                    """, (
                        admin['user_id'], estado_codigo, submission_id,
                        f'{requests_creados} activo(s) (SE/CT) reportados fuera de catálogo '
                        f'pendientes de revisión. Archivo: {result.get("source_filename", "desconocido")}',
                        requests_creados, 0, requests_creados,
                    ), fetch=False)
                except Exception:
                    pass


    # -- Enviar notificación email si está configurado --
    try:
        email_config = query_one("SELECT valor FROM sctis.configuracion WHERE clave = 'notificacion_email'")
        if email_config and email_config['valor']:
            email_to = email_config['valor']
            asunto = f"Nueva Carga de Datos ({estado_codigo})"
            cuerpo = f"Se ha realizado una nueva carga de datos para el estado {estado_codigo}.\n"
            cuerpo += f"Registros procesados exitosamente: {insertados}\n"
            if rechazados > 0:
                cuerpo += f"Se generó una tarea de corrección por {rechazados} registros rechazados.\n"
            
            # En un entorno real se usaría la cuenta de servicio o un token válido de aplicación.
            # send_notification_email(access_token_app, email_to, asunto, cuerpo)
            pass
    except Exception as e:
        print("Error al intentar notificar", e)

    for f in [tmp_path, preview_path]:
        if os.path.exists(f):
            os.remove(f)
    session.pop('import_token', None)

    return jsonify({
        'ok': True,
        'insertados': insertados,
        'errores': errores[:20],
        'total_errores': len(errores),
        'submission_id': submission_id,
        'correccion_generada': rechazados > 0,
        'requests_creados': requests_creados,
    })


@bp.route('/api/importar/limpiar', methods=['POST'])
def limpiar_import():
    token = session.pop('import_token', None)
    if token:
        for f in [os.path.join(UPLOAD_DIR, f'{token}.xlsx'), os.path.join(UPLOAD_DIR, f'{token}_preview.json')]:
            if os.path.exists(f):
                os.remove(f)
    return jsonify({'ok': True})


@bp.route('/api/importar/sugerir-causas', methods=['POST'])
@login_required
def sugerir_causas_route():
    data = request.get_json() or {}
    causas_input = data.get('causas') or []
    if isinstance(causas_input, str):
        causas_input = [causas_input]

    _, causas_db = _load_caches()
    sugerencias = {}

    for c_text in causas_input:
        if not c_text:
            continue
        norm_orig = normalizar_texto(c_text)
        best_ratio = 0.0
        best_c = None

        for c in causas_db:
            norm_c = normalizar_texto(c['causa_nombre'])
            if norm_orig == norm_c:
                best_c = c
                best_ratio = 1.0
                break
            ratio = difflib.SequenceMatcher(None, norm_orig, norm_c).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_c = c

        if best_c:
            sugerencias[c_text] = {
                'causa_id': best_c['causa_id'],
                'causa_codigo': best_c['causa_codigo'],
                'causa_nombre': best_c['causa_nombre'],
                'confianza': round(best_ratio * 100)
            }

    return jsonify({'ok': True, 'sugerencias': sugerencias})

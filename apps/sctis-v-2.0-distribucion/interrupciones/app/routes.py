from flask import Blueprint, render_template, request, jsonify, g, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from app.db import query, query_one
from app.auth import login_required, user_estado_filter
from app.ai import suggest_cause, batch_suggest_causes, evaluate_quality_with_ai, detect_duplicates_with_ai

bp = Blueprint('main', __name__)

# ─── Páginas ────────────────────────────────────────────────

@bp.route('/')
@login_required
def index():
    return render_template('index.html')

@bp.route('/nuevo')
@login_required
def nuevo():
    user_estado = g.user.get('estado_codigo')
    if g.is_admin:
        estados = query("SELECT state_code, state_name FROM common.states ORDER BY state_name")
    else:
        estados = query("SELECT state_code, state_name FROM common.states WHERE state_code = %s ORDER BY state_name", (user_estado,))
    causas = query("SELECT causa_id, causa_codigo, causa_nombre FROM sctis.causa ORDER BY causa_nombre")
    return render_template('form.html', estados=estados, causas=causas, user_estado=user_estado)

@bp.route('/consulta')
@login_required
def consulta():
    return render_template('consulta.html')

# ─── API ─────────────────────────────────────────────────────

@bp.route('/api/interrupciones', methods=['GET'])
@login_required
def listar_interrupciones():
    estado = request.args.get('estado')
    mes = request.args.get('mes')
    estado_calculo = request.args.get('estado_calculo')
    where = []
    params = []
    if estado:
        where.append("ti.estado_codigo = %s")
        params.append(estado)
    if mes:
        where.append("ti.mes = %s")
        params.append(mes)
    if estado_calculo == 'REVISAR CALCULO':
        where.append("ti.estado_calculo = 'REVISAR CALCULO'")
    elif estado_calculo == 'CALCULO VALIDO':
        where.append("ti.estado_calculo = 'CALCULO VALIDO'")
    elif estado_calculo == 'INCOMPLETO_EXCEPCION_ADMIN':
        where.append("(ti.estado_calculo = 'INCOMPLETO_EXCEPCION_ADMIN' OR ti.es_excepcion_admin = 1)")
    elif estado_calculo == 'SIN DATOS':
        where.append("ti.estado_calculo IS NULL")

    estado_clause, estado_params = user_estado_filter()
    if estado_clause:
        where.append(estado_clause)
        params.extend(estado_params)

    sql = """
        SELECT ti.tira_id, ti.estado_codigo, ti.fecha_falla, ti.sistema,
               ti.subestacion, ti.subestacion_id, ti.circuito, ti.circuito_id,
               ti.jefatura, ti.fecha_inicio, ti.fecha_fin,
               ti.causa, ti.sub_causa, ti.causa_id, ti.observacion,
               ti.despachador, ti.despachador_id, ti.racion, ti.mw, ti.kva,
               ti.horas, ti.duracion_calculada::text AS duracion_calculada,
               ti.horas_calculadas, ti.diferencia_horas, ti.estado_calculo,
               ti.es_excepcion_admin, ti.audit_id,
               ti.mes, ti.sectores, ti.ciudad, ti.activo,
               s.state_name, c.causa_nombre AS causa_homologada,
               COALESCE(se.asset_name_normalizado, se.asset_name) AS subestacion_normalizado,
               COALESCE(ci.asset_name_normalizado, ci.asset_name) AS circuito_normalizado,
               CASE WHEN ti.subestacion_id IS NOT NULL THEN 'MATCH' ELSE 'SIN_MATCH' END AS se_match_estado,
               CASE WHEN ti.circuito_id IS NOT NULL THEN 'MATCH' ELSE 'SIN_MATCH' END AS ct_match_estado
        FROM sctis.tira_interrupcion ti
        JOIN common.states s ON s.state_code = ti.estado_codigo
        LEFT JOIN sctis.causa c ON c.causa_id = ti.causa_id
        LEFT JOIN common.assets se ON se.asset_id = ti.subestacion_id
        LEFT JOIN common.assets ci ON ci.asset_id = ti.circuito_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY ti.fecha_falla DESC, ti.tira_id"
    return jsonify(query(sql, params))

@bp.route('/api/interrupciones/<int:tira_id>', methods=['GET'])
def obtener_interrupcion(tira_id):
    row = query_one("""
        SELECT ti.tira_id, ti.estado_codigo, ti.fecha_falla, ti.sistema,
               ti.subestacion, ti.subestacion_id, ti.circuito, ti.circuito_id,
               ti.jefatura, ti.fecha_inicio, ti.fecha_fin,
               ti.causa, ti.sub_causa, ti.causa_id, ti.observacion,
               ti.despachador, ti.despachador_id, ti.racion, ti.mw, ti.kva,
               ti.horas, ti.duracion_calculada::text AS duracion_calculada,
               ti.horas_calculadas, ti.diferencia_horas, ti.estado_calculo,
               ti.mes, ti.sectores, ti.ciudad, ti.activo,
               s.state_name, c.causa_nombre AS causa_homologada,
               COALESCE(se.asset_name_normalizado, se.asset_name) AS subestacion_normalizado,
               COALESCE(ci.asset_name_normalizado, ci.asset_name) AS circuito_normalizado,
               CASE WHEN ti.subestacion_id IS NOT NULL THEN 'MATCH' ELSE 'SIN_MATCH' END AS se_match_estado,
               CASE WHEN ti.circuito_id IS NOT NULL THEN 'MATCH' ELSE 'SIN_MATCH' END AS ct_match_estado
        FROM sctis.tira_interrupcion ti
        JOIN common.states s ON s.state_code = ti.estado_codigo
        LEFT JOIN sctis.causa c ON c.causa_id = ti.causa_id
        LEFT JOIN common.assets se ON se.asset_id = ti.subestacion_id
        LEFT JOIN common.assets ci ON ci.asset_id = ti.circuito_id
        WHERE ti.tira_id = %s
    """, (tira_id,))
    if not row:
        return jsonify({'error': 'No encontrado'}), 404
    return jsonify(row)

def normalizar_nombre(nombre):
    """Limpia y normaliza un texto: elimina espacios extra, capitaliza palabras."""
    if not nombre:
        return nombre
    import re
    nombre = re.sub(r'\s+', ' ', nombre.strip())
    return ' '.join(w.capitalize() for w in nombre.split())


def asegurar_despachador(nombre_completo, estado_codigo):
    """Crea el despachador si no existe y retorna su ID."""
    if not nombre_completo or not nombre_completo.strip():
        return None
    nombre = normalizar_nombre(nombre_completo)
    existente = query_one(
        "SELECT despachador_id FROM sctis.despachador WHERE LOWER(nombre_completo) = LOWER(%s)",
        (nombre,)
    )
    if existente:
        return existente['despachador_id']
    result = query(
        "INSERT INTO sctis.despachador (nombre_completo, estado_codigo, activo) VALUES (%s, %s, true) RETURNING despachador_id",
        (nombre, estado_codigo)
    )
    return result[0]['despachador_id']


@bp.route('/api/interrupciones', methods=['POST'])
@login_required
def crear_interrupcion():
    data = {k: v for k, v in (request.get_json() or {}).items() if v is not None and v != ''}
    user_estado = g.user.get('estado_codigo')
    if user_estado:
        data['estado_codigo'] = user_estado

    subestacion_nombre = data.get('subestacion')
    circuito_nombre = data.get('circuito')

    if not subestacion_nombre and data.get('subestacion_id'):
        row = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (data['subestacion_id'],))
        if row:
            subestacion_nombre = row['asset_name']

    if not circuito_nombre and data.get('circuito_id'):
        row = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (data['circuito_id'],))
        if row:
            circuito_nombre = row['asset_name']

    nombre_desp = data.get('despachador')
    desp_id = data.get('despachador_id')
    if nombre_desp and not desp_id:
        desp_id = asegurar_despachador(nombre_desp, data.get('estado_codigo'))
        data['despachador_id'] = desp_id
        data['despachador'] = normalizar_nombre(nombre_desp)

    # Aplicar regla universal de extracción y reformato de fecha y hora
    from app.import_routes import normalizar_y_extraer_registro
    normalizar_y_extraer_registro(data)

    cols = ['estado_codigo','fecha_falla','sistema','subestacion','subestacion_id',
            'circuito','circuito_id','jefatura','fecha_inicio','fecha_fin',
            'causa','sub_causa','observacion','despachador','despachador_id',
            'racion','mw','kva','horas','mes','sectores','ciudad','causa_id']
    params = {}
    for c in cols:
        params[c] = data.get(c)
    if subestacion_nombre:
        params['subestacion'] = subestacion_nombre
    if circuito_nombre:
        params['circuito'] = circuito_nombre

    sql = f"""
        INSERT INTO sctis.tira_interrupcion
            ({', '.join(cols)})
        VALUES ({', '.join(f'%({c})s' for c in cols)})
        RETURNING tira_id
    """
    result = query(sql, params)
    return jsonify({'tira_id': result[0]['tira_id']}), 201

@bp.route('/api/interrupciones/<int:tira_id>', methods=['PUT'])
def actualizar_interrupcion(tira_id):
    data = request.get_json()
    data['tira_id'] = tira_id

    subestacion_nombre = data.get('subestacion')
    circuito_nombre = data.get('circuito')

    if not subestacion_nombre and data.get('subestacion_id'):
        row = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (data['subestacion_id'],))
        if row:
            subestacion_nombre = row['asset_name']

    if not circuito_nombre and data.get('circuito_id'):
        row = query_one("SELECT COALESCE(asset_name_normalizado, asset_name) AS asset_name FROM common.assets WHERE asset_id = %s", (data['circuito_id'],))
        if row:
            circuito_nombre = row['asset_name']

    data['subestacion'] = subestacion_nombre
    data['circuito'] = circuito_nombre

    # Aplicar regla universal de extracción y reformato de fecha y hora
    from app.import_routes import normalizar_y_extraer_registro
    normalizar_y_extraer_registro(data)

    sql = """
        UPDATE sctis.tira_interrupcion SET
            estado_codigo=%(estado_codigo)s, fecha_falla=%(fecha_falla)s,
            sistema=%(sistema)s, subestacion=%(subestacion)s, subestacion_id=%(subestacion_id)s,
            circuito=%(circuito)s, circuito_id=%(circuito_id)s,
            jefatura=%(jefatura)s, fecha_inicio=%(fecha_inicio)s, fecha_fin=%(fecha_fin)s,
            causa=%(causa)s, sub_causa=%(sub_causa)s, observacion=%(observacion)s,
            despachador=%(despachador)s, despachador_id=%(despachador_id)s,
            racion=%(racion)s, mw=%(mw)s, kva=%(kva)s,
            horas=%(horas)s, mes=%(mes)s, sectores=%(sectores)s, ciudad=%(ciudad)s,
            causa_id=%(causa_id)s
        WHERE tira_id = %(tira_id)s
    """
    query(sql, data, fetch=False)
    return jsonify({'ok': True})

@bp.route('/api/interrupciones/<int:tira_id>', methods=['DELETE'])
def eliminar_interrupcion(tira_id):
    query("DELETE FROM sctis.tira_interrupcion WHERE tira_id = %s", (tira_id,), fetch=False)
    return jsonify({'ok': True})

@bp.route('/api/causas')
def listar_causas():
    return jsonify(query("SELECT * FROM sctis.causa ORDER BY causa_nombre"))

@bp.route('/api/sub_causas')
def listar_sub_causas():
    causa_id = request.args.get('causa_id')
    if causa_id:
        return jsonify(query("SELECT * FROM sctis.sub_causa WHERE causa_id = %s ORDER BY sub_causa_nombre", (causa_id,)))
    return jsonify(query("SELECT sc.*, c.causa_nombre FROM sctis.sub_causa sc JOIN sctis.causa c ON c.causa_id = sc.causa_id ORDER BY sc.sub_causa_nombre"))

@bp.route('/api/despachadores')
def listar_despachadores():
    estado = request.args.get('estado')
    if estado:
        return jsonify(query("""
            SELECT * FROM sctis.despachador
            WHERE (estado_codigo = %s OR estado_codigo IS NULL)
              AND activo = true
            ORDER BY nombre_completo
        """, (estado,)))
    return jsonify(query("SELECT * FROM sctis.despachador WHERE activo = true ORDER BY nombre_completo"))

@bp.route('/api/estados')
def listar_estados():
    return jsonify(query("SELECT state_code, state_name FROM common.states ORDER BY state_name"))


COLUMNAS_FORMULARIO = [
    ('ESTADO', 15),
    ('SISTEMA', 12),
    ('DISTRITO', 12),
    ('SUBESTACIÓN', 20),
    ('CIRCUITO', 15),
    ('FECHA INICIO', 14),
    ('HORA INICIO', 10),
    ('FECHA FIN', 14),
    ('HORA FIN', 10),
    ('DURACIÓN', 10),
    ('CARGA', 8),
    ('FREC', 6),
    ('HORAS', 8),
    ('TTI CTO', 8),
    ('SEÑAL', 6),
    ('CAUSA', 20),
    ('SUB-CAUSA', 20),
    ('OBSERVACIÓN', 30),
    ('SECTORES', 25),
    ('CIUDAD', 15),
    ('KVA', 8),
]


def generar_formulario_carga(estado_codigo=None, estado_nombre=None):
    wb = openpyxl.Workbook()

    # ─── Hoja 1: Formulario de Carga ───────────────────────
    ws = wb.active
    ws.title = 'FORMULARIO DE CARGA'

    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, (header, width) in enumerate(COLUMNAS_FORMULARIO, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Fila 2: si estado está definido, pre-rellenar
    if estado_codigo:
        ws.cell(row=2, column=1, value=estado_nombre or estado_codigo)

    # Congelar panel
    ws.freeze_panes = 'A2'

    # ─── Hoja 2: Instructivo de Carga ──────────────────────
    ws2 = wb.create_sheet('INSTRUCTIVO DE CARGA')

    title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    subtitle_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    body_font = Font(name='Calibri', size=10)
    body_align = Alignment(wrap_text=True, vertical='top')

    row = 1
    ws2.cell(row=row, column=1, value='INSTRUCTIVO DE CARGA - FORMULARIO OFICIAL SCTIS').font = title_font
    row += 2

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 90

    instrucciones = [
        ('1. PROPÓSITO', [
            'Este formulario es el único instrumento homologado para la carga masiva de Tiras de Interrupción de Servicio Eléctrico.',
            'Cumple con los requisitos de ISO 27001 (seguridad de la información) e ISO 8000 (calidad de datos).',
        ]),
        ('2. FORMATO OBLIGATORIO', [
            'El archivo debe conservar exactamente las 21 columnas en el orden establecido.',
            'No se deben agregar, eliminar ni reordenar columnas.',
            'No se deben combinar celdas ni agregar formatos condicionales.',
            'Solo se acepta la primera pestaña del libro como fuente de datos.',
            'Cualquier modificación a la estructura será rechazada por el sistema.',
        ]),
        ('3. INSTRUCCIONES POR COLUMNA', [
            'A - ESTADO: Nombre del estado donde ocurrió la interrupción.',
            'B - SISTEMA: Sistema eléctrico (DISTRIBUCION, TRANSMISION, etc.).',
            'C - DISTRITO: Distrito o jefatura responsable.',
            'D - SUBESTACIÓN: Nombre de la subestación afectada.',
            'E - CIRCUITO: Nombre del circuito afectado.',
            'F - FECHA INICIO: Fecha de inicio de la interrupción (DD/MM/AAAA o AAAA-MM-DD).',
            'G - HORA INICIO: Hora de inicio de la interrupción (HH:MM).',
            'H - FECHA FIN: Fecha de reposición del servicio.',
            'I - HORA FIN: Hora de reposición del servicio.',
            'J - DURACIÓN: Duración del evento.',
            'M - HORAS: Horas de interrupción (valor numérico).',
            'P - CAUSA: Descripción de la causa original de la interrupción.',
            'Q - SUB-CAUSA: Detalle adicional de la causa.',
            'R - OBSERVACIÓN: Notas y observaciones.',
            'S - SECTORES: Sectores afectados.',
            'T - CIUDAD: Ciudad o localidad.',
            'U - KVA: Potencia afectada en KVA.',
        ]),
        ('4. VALIDACIONES DEL SISTEMA', [
            'El sistema validará automáticamente:',
            '  • Estructura del archivo (columnas y encabezados)',
            '  • Homologación de causas contra el catálogo oficial de 16 causas',
            '  • Correspondencia de subestaciones y circuitos con el catálogo del estado',
            '  • Normalización de nombres de despachadores',
            '  • Calidad de datos según reglas ISO 8000',
            '  • Detección de registros duplicados',
        ]),
        ('5. NO REPUDIO', [
            'El sistema registra en auditoría cada descarga de este formulario, cada intento de carga',
            'y cada operación realizada. El operador es responsable de la veracidad de los datos cargados.',
        ]),
        ('6. CONTACTO', [
            'Para soporte o solicitud de cambios al formulario, contacte a la Coordinación de SCTIS.',
        ]),
    ]

    for titulo, items in instrucciones:
        ws2.cell(row=row, column=1).value = ''
        ws2.cell(row=row, column=2, value=titulo).font = subtitle_font
        row += 1
        for item in items:
            ws2.cell(row=row, column=2, value=item).font = body_font
            ws2.cell(row=row, column=2).alignment = body_align
            row += 1
        row += 1

    return wb


@bp.route('/api/descargar-formulario')
@login_required
def descargar_formulario():
    estado_codigo = request.args.get('estado') or g.user.get('estado_codigo')
    estado_nombre = None
    if estado_codigo:
        row = query_one("SELECT state_name FROM common.states WHERE state_code = %s", (estado_codigo,))
        if row:
            estado_nombre = row['state_name']

    wb = generar_formulario_carga(estado_codigo, estado_nombre)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'Formulario Carga SCTIS{(" - " + estado_codigo) if estado_codigo else ""}.xlsx'

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )

@bp.route('/api/subestaciones')
def listar_subestaciones():
    estado = request.args.get('estado')
    if not estado:
        return jsonify([])
    rows = query("""
        SELECT asset_id, asset_code, 
               COALESCE(asset_name_normalizado, asset_name) AS asset_name,
               asset_name AS asset_name_original,
               state_code
        FROM common.assets
        WHERE asset_type = 'SUBSTATION'
          AND state_code = %s
          AND is_active = true
        ORDER BY asset_name
    """, (estado,))
    return jsonify(rows)

@bp.route('/api/circuitos')
def listar_circuitos():
    subestacion_id = request.args.get('subestacion_id')
    if not subestacion_id:
        return jsonify([])
    rows = query("""
        SELECT a.asset_id, a.asset_code, 
               COALESCE(a.asset_name_normalizado, a.asset_name) AS asset_name,
               a.asset_name AS asset_name_original,
               a.elemento_tipo,
               a.elemento_codigo,
               a.state_code
        FROM common.assets a
        WHERE a.asset_type = 'CIRCUITO'
          AND a.parent_asset_id = %s::bigint
          AND a.is_active = true
        ORDER BY a.asset_name
    """, (subestacion_id,))
    return jsonify(rows)


# ─── AI Endpoints ────────────────────────────────────

@bp.route('/api/ai/suggest-cause', methods=['POST'])
@login_required
def ai_suggest_cause():
    data = request.get_json() or {}
    original_cause = data.get('causa', '')
    sub_causa = data.get('sub_causa', '')
    if not original_cause.strip():
        return jsonify({'error': 'Se requiere una causa original'}), 400
    result = suggest_cause(original_cause, sub_causa)
    return jsonify(result)


@bp.route('/api/ai/suggest-causes-batch', methods=['POST'])
@login_required
def ai_suggest_causes_batch():
    data = request.get_json() or {}
    records = data.get('records', [])
    if not records:
        return jsonify({'error': 'Se requiere al menos un registro'}), 400
    results = batch_suggest_causes(records)
    return jsonify({'results': results})


@bp.route('/api/ai/quality-score', methods=['POST'])
@login_required
def ai_quality_score():
    data = request.get_json() or {}
    if not data:
        return jsonify({'error': 'Se requiere un registro para evaluar'}), 400
    result = evaluate_quality_with_ai(data)
    return jsonify(result)


@bp.route('/api/ai/detect-duplicates', methods=['POST'])
@login_required
def ai_detect_duplicates():
    data = request.get_json() or {}
    records = data.get('records', [])
    if not records:
        return jsonify({'error': 'Se requiere al menos un registro'}), 400
    results = detect_duplicates_with_ai(records)
    return jsonify({'results': results})


# ─── Dashboard Multi-Perspectiva ──────────────────────────────

@bp.route('/dashboard')
@login_required
def dashboard_page():
    return render_template('dashboard.html')


@bp.route('/api/dashboard')
@login_required
def dashboard_data():
    estado = request.args.get('estado')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    where = ['ti.activo = true']
    params = []

    if estado:
        where.append('ti.estado_codigo = %s')
        params.append(estado)
    if fecha_desde:
        where.append('ti.fecha_falla >= %s')
        params.append(fecha_desde)
    if fecha_hasta:
        where.append('ti.fecha_falla <= %s')
        params.append(fecha_hasta)

    where_sql = ' AND '.join(where)

    # ── KPIs Generales ──
    kpis = query_one(f"""
        SELECT
            COUNT(*) AS total_interrupciones,
            COALESCE(SUM(horas), 0) AS total_horas,
            COALESCE(AVG(horas), 0) AS horas_promedio,
            COALESCE(MAX(horas), 0) AS max_duracion,
            COALESCE(SUM(mw), 0) AS total_mw,
            COALESCE(SUM(racion), 0) AS total_clientes_afectados,
            COUNT(DISTINCT circuito) AS circuitos_afectados,
            COUNT(DISTINCT subestacion) AS subestaciones_afectadas,
            MIN(fecha_falla) AS primera_fecha,
            MAX(fecha_falla) AS ultima_fecha
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql}
    """, params)

    # ── Ingeniero Electricista: Causas por categoría ──
    causas_resumen = query(f"""
        SELECT c.causa_nombre AS causa, COUNT(*) AS cantidad,
               COALESCE(SUM(ti.horas), 0) AS horas_total,
               COALESCE(AVG(ti.horas), 0) AS horas_promedio
        FROM sctis.tira_interrupcion ti
        JOIN sctis.causa c ON c.causa_id = ti.causa_id
        WHERE {where_sql}
        GROUP BY c.causa_nombre
        ORDER BY cantidad DESC
    """, params)

    # ── Ingeniero Electricista: Interrupciones por mes ──
    por_mes = query(f"""
        SELECT ti.mes, COUNT(*) AS cantidad,
               COALESCE(SUM(ti.horas), 0) AS horas_total
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql} AND ti.mes IS NOT NULL
        GROUP BY ti.mes
        ORDER BY CASE ti.mes
            WHEN 'ENERO' THEN 1 WHEN 'FEBRERO' THEN 2 WHEN 'MARZO' THEN 3
            WHEN 'ABRIL' THEN 4 WHEN 'MAYO' THEN 5 WHEN 'JUNIO' THEN 6
            WHEN 'JULIO' THEN 7 WHEN 'AGOSTO' THEN 8 WHEN 'SEPTIEMBRE' THEN 9
            WHEN 'OCTUBRE' THEN 10 WHEN 'NOVIEMBRE' THEN 11 WHEN 'DICIEMBRE' THEN 12
        END
    """, params)

    # ── Ingeniero Mantenimiento: Circuitos más afectados ──
    circuitos_criticos = query(f"""
        SELECT ti.circuito, ti.subestacion, COUNT(*) AS fallas,
               COALESCE(SUM(ti.horas), 0) AS horas_total,
               COALESCE(AVG(ti.horas), 0) AS horas_promedio
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql} AND ti.circuito IS NOT NULL
        GROUP BY ti.circuito, ti.subestacion
        ORDER BY fallas DESC
        LIMIT 10
    """, params)

    # ── Ingeniero Mantenimiento: Subestaciones con más fallas ──
    subestaciones_fallas = query(f"""
        SELECT ti.subestacion, COUNT(*) AS fallas,
               COUNT(DISTINCT ti.circuito) AS circuitos_afectados,
               COALESCE(SUM(ti.horas), 0) AS horas_total
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql} AND ti.subestacion IS NOT NULL
        GROUP BY ti.subestacion
        ORDER BY fallas DESC
        LIMIT 10
    """, params)

    # ── Project Manager: Distribución por estado ──
    por_estado = query(f"""
        SELECT ti.estado_codigo, COUNT(*) AS cantidad,
               COALESCE(SUM(ti.horas), 0) AS horas_total
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql}
        GROUP BY ti.estado_codigo
        ORDER BY cantidad DESC
    """, params)

    # ── Project Manager: Top 5 interrupciones más largas ──
    top_largas = query(f"""
        SELECT ti.tira_id, ti.fecha_falla, ti.estado_codigo, ti.circuito,
               ti.subestacion, ti.causa, ti.horas, ti.mw, ti.kva
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql}
        ORDER BY ti.horas DESC NULLS LAST
        LIMIT 5
    """, params)

    # ── Científico de Datos: Detección de anomalías ──
    # Circuitos con fallas repetidas en poco tiempo
    repetidos = query(f"""
        SELECT ti.circuito, ti.estado_codigo,
               COUNT(*) AS repeticiones,
               MIN(ti.fecha_falla) AS primera_falla,
               MAX(ti.fecha_falla) AS ultima_falla
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql} AND ti.circuito IS NOT NULL
        GROUP BY ti.circuito, ti.estado_codigo
        HAVING COUNT(*) >= 3
        ORDER BY repeticiones DESC
        LIMIT 10
    """, params)

    # ── Científico de Datos: Patrón horario ──
    patron_horario = query(f"""
        SELECT EXTRACT(HOUR FROM ti.fecha_inicio) AS hora,
               COUNT(*) AS cantidad
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql} AND ti.fecha_inicio IS NOT NULL
        GROUP BY EXTRACT(HOUR FROM ti.fecha_inicio)
        ORDER BY hora
    """, params)

    # ── Científico de Datos: Distribución por sistema ──
    por_sistema = query(f"""
        SELECT COALESCE(ti.sistema, 'NO ESPECIFICADO') AS sistema,
               COUNT(*) AS cantidad,
               COALESCE(AVG(ti.horas), 0) AS horas_promedio
        FROM sctis.tira_interrupcion ti
        WHERE {where_sql}
        GROUP BY ti.sistema
        ORDER BY cantidad DESC
    """, params)

    # ── Alertas: Calidad de datos ──
    calidad_baja = query(f"""
        SELECT rqs.row_pk, rqs.score, rqs.rules_failed
        FROM sctis.record_quality_scores rqs
        WHERE rqs.table_name = 'tira_interrupcion'
        AND rqs.score < 80
        ORDER BY rqs.score ASC
        LIMIT 10
    """)

    # ── Alertas: Duplicados detectados ──
    duplicados = query("""
        SELECT dg.group_id, dg.group_ref, dg.description,
               COUNT(dm.member_id) AS miembros
        FROM sctis.duplicate_groups dg
        JOIN sctis.duplicate_members dm ON dm.group_id = dg.group_id
        GROUP BY dg.group_id, dg.group_ref, dg.description
        ORDER BY miembros DESC
        LIMIT 10
    """)

    # ── Estadísticas de calidad ISO 8000 ──
    stats_calidad = query_one("""
        SELECT
            COUNT(*) AS total_evaluados,
            COALESCE(AVG(score), 0) AS score_promedio,
            COUNT(*) FILTER (WHERE score >= 90) AS excelentes,
            COUNT(*) FILTER (WHERE score >= 70 AND score < 90) AS buenos,
            COUNT(*) FILTER (WHERE score < 70) AS deficientes
        FROM sctis.record_quality_scores
        WHERE table_name = 'tira_interrupcion'
    """)

    # ── Desglose de causas "OTRAS": causa homologada → causa original ──
    desglose_otras = query(f"""
        SELECT c.causa_nombre AS causa_homologada,
               COALESCE(NULLIF(TRIM(ti.causa), ''), 'SIN ESPECIFICAR') AS causa_original,
               COALESCE(NULLIF(TRIM(ti.sub_causa), ''), NULL) AS sub_causa,
               COUNT(*) AS cantidad,
               COALESCE(SUM(ti.horas), 0) AS horas_total
        FROM sctis.tira_interrupcion ti
        JOIN sctis.causa c ON c.causa_id = ti.causa_id
        WHERE ti.activo = true AND c.causa_codigo = 'OTRAS'
        GROUP BY c.causa_nombre, ti.causa, ti.sub_causa
        ORDER BY cantidad DESC
    """, params)

    return jsonify({
        'kpis': kpis,
        'causas_resumen': causas_resumen,
        'por_mes': por_mes,
        'circuitos_criticos': circuitos_criticos,
        'subestaciones_fallas': subestaciones_fallas,
        'por_estado': por_estado,
        'top_largas': top_largas,
        'repetidos': repetidos,
        'patron_horario': patron_horario,
        'por_sistema': por_sistema,
        'calidad_baja': calidad_baja,
        'duplicados': duplicados,
        'stats_calidad': stats_calidad,
        'desglose_otras': desglose_otras,
        'filtros': {'estado': estado, 'fecha_desde': fecha_desde, 'fecha_hasta': fecha_hasta}
    })


@bp.route('/api/audit/cargas-excepcionales', methods=['GET'])
@login_required
def listar_auditorias_cargas():
    rows = query("""
        SELECT audit_id, user_id, username, estado_codigo, filename, token,
               total_registros, registros_incompletos, declaracion_no_repudio,
               ip_address, created_at
        FROM "sctis.audit_admin_carga_excepcional"
        ORDER BY created_at DESC
    """)
    return jsonify(rows or [])


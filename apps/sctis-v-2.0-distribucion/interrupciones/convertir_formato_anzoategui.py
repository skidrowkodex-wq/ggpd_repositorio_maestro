#!/usr/bin/env python3
"""
Convierte el formato Excel de ANZOÁTEGUI al formato homologado SCTIS.

Formato ANZOÁTEGUI (original):
  A: Estado       B: Fecha        C: Sistema      D: Distrito
  E: Subestacion  F: Circuito     G: (vacío)      H: Hora Inicio
  I: (vacío)      J: Duración     K: Carga        L: Frec
  M: Horas        N: TTI          O: Señal        P: Causa
  Q: Sub-Causa    R: Observaciones S: Sectores    T: Ciudad
  U: kVa Instalados

Formato Homologado (esperado por la app):
  A: Estado       B: Sistema      C: Jefatura     D: Subestacion
  E: Circuito     F: Fecha Inicio G: Hora Inicio  H: Fecha Fin
  I: Hora Fin     J: Duración     K: Carga        L: Frec
  M: Horas        N: TTI          O: Señal        P: Causa
  Q: Sub-Causa    R: Observaciones S: Sectores    T: Ciudad
  U: kVa Instalados

Uso:
  python3 convertir_formato_anzoategui.py input.xlsx output.xlsx
"""

import sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime


def convertir(input_path, output_path):
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "TIRAS"

    # ── Encabezados homologados ──
    headers = [
        'Estado', 'Sistema', 'Jefatura', 'Subestacion', 'Circuito',
        'Fecha Inicio', 'Hora Inicio', 'Fecha Fin', 'Hora Fin',
        'Duración', 'Carga', 'Frec', 'Dur (Hora)', 'TTI',
        'Señal', 'Causa', '(Sub-Causa)', 'Observaciones',
        'Sectores', 'Ciudad', 'kVa Instalados'
    ]
    ws_out.append(headers)

    # Formato encabezados
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col_idx, cell in enumerate(ws_out[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # ── Mapeo columnas ANZO → Homologado ──
    # ANZO: A=Estado, B=Fecha, C=Sistema, D=Distrito, E=Subest, F=Circuito,
    #        G=(vacío), H=Hora Inicio, I=(vacío), J=Duración, K=Carga, L=Frec,
    #        M=Horas, N=TTI, O=Señal, P=Causa, Q=SubCausa, R=Observaciones,
    #        S=Sectores, T=Ciudad, U=kVa
    #
    # HOME: A=Estado, B=Sistema, C=Jefatura, D=Subest, E=Circuito,
    #        F=Fecha Inicio, G=Hora Inicio, H=Fecha Fin, I=Hora Fin,
    #        J=Duración, K=Carga, L=Frec, M=Horas, N=TTI, O=Señal,
    #        P=Causa, Q=SubCausa, R=Observaciones, S=Sectores, T=Ciudad, U=kVa

    col_map = {
        'A': 'estado',      # A → A
        'C': 'sistema',     # C → B
        'D': 'jefatura',    # D → C
        'E': 'subestacion', # E → D
        'F': 'circuito',    # F → E
        'B': 'fecha_inicio',# B → F
        'H': 'hora_inicio', # H → G
        'J': 'duracion',    # J → J
        'K': 'carga',       # K → K
        'L': 'frec',        # L → L
        'M': 'horas',       # M → M
        'N': 'tti',         # N → N
        'O': 'senal',       # O → O
        'P': 'causa',       # P → P
        'Q': 'sub_causa',   # Q → Q
        'R': 'observacion', # R → R
        'S': 'sectores',    # S → S
        'T': 'ciudad',      # T → T
        'U': 'kva',         # U → U
    }

    # Columna destino de cada campo
    dest_col = {
        'estado': 1, 'sistema': 2, 'jefatura': 3, 'subestacion': 4,
        'circuito': 5, 'fecha_inicio': 6, 'hora_inicio': 7,
        'duracion': 10, 'carga': 11, 'frec': 12, 'horas': 13,
        'tti': 14, 'senal': 15, 'causa': 16, 'sub_causa': 17,
        'observacion': 18, 'sectores': 19, 'ciudad': 20, 'kva': 21,
    }

    start_row = 2
    total = 0
    vacias = 0

    for row_idx in range(start_row, (ws_in.max_row or start_row) + 1):
        cells_in = list(ws_in.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        if not cells_in:
            continue
        row_data = cells_in[0]

        # Saltar filas vacías
        if not row_data or all(c is None or str(c).strip() == '' for c in row_data):
            vacias += 1
            continue

        # Verificar que la fila tenga datos útiles (al menos fecha o causa)
        fecha_val = row_data[1] if len(row_data) > 1 else None  # Col B = Fecha
        causa_val = row_data[15] if len(row_data) > 15 else None  # Col P = Causa
        if not fecha_val and not causa_val:
            vacias += 1
            continue

        # Construir fila de salida
        out_row = [''] * 21

        for src_col_letter, field_name in col_map.items():
            src_idx = ord(src_col_letter) - ord('A')
            if src_idx < len(row_data):
                val = row_data[src_idx]
                if val is not None:
                    # Normalizar fechas
                    if field_name == 'fecha_inicio':
                        if isinstance(val, datetime):
                            val = val.strftime('%Y-%m-%d')
                        elif isinstance(val, str):
                            val = val.strip()
                    dest_idx = dest_col.get(field_name)
                    if dest_idx:
                        out_row[dest_idx - 1] = val

        # Derivar fecha_fin y hora_fin desde fecha_inicio + horas
        fecha_inicio = out_row[5]  # Col F (índice 5)
        horas_val = out_row[12]    # Col M (índice 12) = Dur (Hora)
        if fecha_inicio and horas_val:
            try:
                if isinstance(fecha_inicio, str):
                    dt_inicio = datetime.fromisoformat(fecha_inicio)
                else:
                    dt_inicio = fecha_inicio
                h = float(horas_val)
                dt_fin = dt_inicio.replace(hour=0, minute=0, second=0)
                # Agregar horas como duración
                from datetime import timedelta
                dt_fin = dt_inicio + timedelta(hours=h)
                out_row[7] = dt_fin.strftime('%Y-%m-%d')  # H = Fecha Fin
                out_row[8] = dt_fin.strftime('%H:%M')      # I = Hora Fin
            except (ValueError, TypeError):
                pass

        ws_out.append(out_row)
        total += 1

    # Ajustar anchos de columna
    widths = [14, 8, 12, 25, 30, 12, 10, 12, 10, 10, 8, 6, 10, 8, 8, 35, 35, 40, 20, 15, 12]
    for i, w in enumerate(widths, 1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb_out.save(output_path)
    wb_in.close()

    print(f"Conversión completada:")
    print(f"  Filas con datos: {total}")
    print(f"  Filas vacías omitidas: {vacias}")
    print(f"  Archivo de salida: {output_path}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(f"Uso: python3 {sys.argv[0]} <input.xlsx> <output.xlsx>")
        sys.exit(1)
    convertir(sys.argv[1], sys.argv[2])

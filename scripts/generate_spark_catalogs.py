#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
⚡ CORPOELEC — GERENCIA GENERAL DE PLANIFICACIÓN DE DISTRIBUCIÓN (GGPD)
Generador de Catálogos Maestros y Tablas de Referencia en Excel (.xlsx)
para Consolidación y Normalización de Interrupciones en Google Spark / Data Lake
==============================================================================
"""

import os
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_spark_catalogs():
    json_path = 'apps/corpoelec-sigi-gestion-planificacion-distribucion/src/data/masterCatalogsLegacy.json'
    with open(json_path, 'r', encoding='utf-8') as f:
        legacy_data = json.load(f)

    output_dir = 'docs/catalogos_maestros_spark'
    os.makedirs(output_dir, exist_ok=True)

    # Styling definitions (CORPOELEC Institutional palette)
    navy_fill = PatternFill(start_color='0F2942', end_color='0F2942', fill_type='solid')
    blue_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    title_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    regular_font = Font(name='Calibri', size=10, color='1E293B')

    border_thin = Side(border_style='thin', color='CBD5E1')
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    def populate_sheet(ws, headers, rows_data, sheet_title=None):
        ws.views.sheetView[0].showGridLines = True
        start_row = 1
        if sheet_title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            t_cell = ws.cell(row=1, column=1, value=sheet_title)
            t_cell.font = title_font
            t_cell.fill = navy_fill
            t_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            start_row = 2

        # Headers
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=start_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = blue_fill if sheet_title else header_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = cell_border
        ws.row_dimensions[start_row].height = 26

        # Data rows
        for r_idx, row in enumerate(rows_data, start_row + 1):
            ws.row_dimensions[r_idx].height = 20
            is_even = (r_idx % 2 == 0)
            c_fill = zebra_fill if is_even else white_fill
            for c_idx, val in enumerate(row, 1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.font = regular_font
                c.fill = c_fill
                c.border = cell_border
                if isinstance(val, (int, float)):
                    c.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center')

        # Auto width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1 and sheet_title:
                    continue
                v_str = str(cell.value or '')
                if len(v_str) > max_len:
                    max_len = len(v_str)
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 50), 12)

    # 1. TABLA MAESTRA DE SUBESTACIONES SEN (871 SEs)
    se_headers = [
        'ID_ACTIVO_NORMALIZADO', 'REGION', 'ESTADO', 'TIPO_INSTALACION', 'NOMBRE_SUBESTACION',
        'NOMBRE_SE_NORM', 'ES_MOVIL', 'OBSERVACIONES_NORMALIZACION', 'ORIGEN',
        'ATENDIDA_POR', 'MUNICIPIO', 'AREA_SE_M2', 'TENSION_ENTRADA_KV',
        'TENSION_SECUNDARIA_KV', 'ALIMENTADOR_PRINCIPAL', 'TENSION_ALIMENTADOR_KV',
        'ESTADO_CONTROL', 'ESTADO_CARACTERIZACION'
    ]
    se_rows = []
    for se in legacy_data['caracterizacion']['subestaciones']:
        se_rows.append([
            se.get('codigo', ''),
            se.get('region', ''),
            se.get('estado', ''),
            se.get('tipo_instalacion', ''),
            se.get('nombre_se', ''),
            se.get('nombre_se_norm', ''),
            'SI' if se.get('es_movil') else 'NO',
            se.get('observaciones_normalizacion') or '',
            se.get('origen', ''),
            se.get('atendida_por', ''),
            se.get('municipio', ''),
            se.get('area_se_m2') or 'N/A',
            se.get('tension_entrada_kv') or 'N/A',
            se.get('tension_secundaria_kv') or 'N/A',
            se.get('alimentador_principal', '') or '',
            se.get('tension_alimentador_kv') or 'N/A',
            se.get('estado_control', 'CONTROLADO'),
            se.get('estado_caracterizacion', 'CARACTERIZADO')
        ])

    wb_se = openpyxl.Workbook()
    ws_se = wb_se.active
    ws_se.title = 'SUBESTACIONES_SEN'
    populate_sheet(ws_se, se_headers, se_rows, 'CORPOELEC GGPD — CATÁLOGO MAESTRO DE SUBESTACIONES DEL SEN (871 REGISTROS)')
    wb_se.save(os.path.join(output_dir, 'CATALOGO_MAESTRO_SUBESTACIONES_SEN.xlsx'))
    print(f'1. Subestaciones generadas: {len(se_rows)} registros')

    # 2. TABLA MAESTRA DE CIRCUITOS SEN (4,207 Circuitos)
    cto_headers = [
        'ID_CIRCUITO_NORMALIZADO', 'REGION', 'ESTADO', 'SUBESTACION_CABECERA', 'se_codigo',
        'NOMBRE_CIRCUITO', 'NOMBRE_CT_NORM', 'ELEMENTOS_TIPO', 'ELEMENTO_TECNICO_ESPECIFICO',
        'CODIGO_MANIOBRA_NORMA', 'DESCRIPCION_ELEMENTO_TECNICO', 'ES_COMPONENTE_PATIO',
        'OBSERVACIONES_NORMALIZACION', 'NIVEL_TENSION_KV', 'LONGITUD_KM', 'TIPO_RED',
        'ESTADO_CONTROL', 'ESTADO_CARACTERIZACION'
    ]
    cto_rows = []
    for cto in legacy_data['caracterizacion']['circuitos']:
        cto_rows.append([
            cto.get('codigo', ''),
            cto.get('region', ''),
            cto.get('estado', ''),
            cto.get('subestacion_cabecera', ''),
            cto.get('se_codigo', ''),
            cto.get('circuito', ''),
            cto.get('nombre_ct_norm', ''),
            cto.get('elemento_tipo', 'CIRCUITO'),
            cto.get('elemento_tecnico_especifico', 'ALIMENTADOR_CONVENCIONAL'),
            cto.get('codigo_maniobra_norma', 'CTO'),
            cto.get('descripcion_elemento_tecnico', 'Alimentador de Distribución Convencional'),
            'SI' if cto.get('es_componente_patio') else 'NO',
            cto.get('observaciones_normalizacion') or '',
            cto.get('nivel_tension_kv', 13.8),
            cto.get('km_total', 0.0),
            cto.get('tipo_red', 'AÉREO'),
            cto.get('estado_control', 'CONTROLADO'),
            cto.get('estado_caracterizacion', 'CARACTERIZADO')
        ])

    wb_cto = openpyxl.Workbook()
    ws_cto = wb_cto.active
    ws_cto.title = 'CIRCUITOS_SEN'
    populate_sheet(ws_cto, cto_headers, cto_rows, 'CORPOELEC GGPD — CATÁLOGO MAESTRO DE CIRCUITOS DE DISTRIBUCIÓN DEL SEN (4,207 REGISTROS)')
    wb_cto.save(os.path.join(output_dir, 'CATALOGO_MAESTRO_CIRCUITOS_SEN.xlsx'))
    print(f'2. Circuitos generados: {len(cto_rows)} registros')

    # 3. TABLA MAESTRA DE CAUSAS, SUB-CAUSAS Y REGLAS DE HOMOLOGACIÓN
    causas_oficiales = [
        ('ACCIDENTAL', 'Accidental', 'Eventos fortuitos externos, contacto de fauna, accidentes de tránsito contra postes.'),
        ('APERTURA_EMERGENCIA', 'Apertura por Emergencia', 'Maniobra forzada ejecutada por Despacho para evitar colapso de red o proteger vidas humanas.'),
        ('ATMOSFERICA', 'Atmosférica', 'Descargas eléctricas directas/indirectas, rayos, tormentas, fuertes precipitaciones o vendavales.'),
        ('ATRIBUIBLE_MANTENIMIENTO', 'Atribuible a Mantenimiento', 'Averías causadas por falta de mantenimiento preventivo, tornillería floja, sulfatación de bornes.'),
        ('ATRIBUIBLE_COORD_PROT', 'Atribuible a Coordinación de Protecciones', 'Disparos indebidos o descalibración en relés, reconectadores o fusibles de respaldo.'),
        ('BAJA_TENSION', 'Baja Tensión', 'Fallas originadas en la red secundaria de baja tensión que repercuten en el alimentador.'),
        ('COMPONENTE_DANADO', 'Componente Dañado', 'Rotura o deterioro físico de líneas, aisladores, seccionadores, crucetas, cortacorrientes o puentes.'),
        ('ERROR_OPERACIONES', 'Error de Operaciones', 'Falla originada por maniobra errónea de personal interno o incumplimiento de protocolo operativo.'),
        ('FALLA_EQUIPOS', 'Falla de Equipos', 'Avería interna en transformadores de potencia, interruptores de potencia, TT/TC o reguladores.'),
        ('FALLA_LINEA_115KV', 'Falla en Línea >= 115 KV', 'Pérdida de alimentación por eventos en el sistema de Transmisión (Líneas 115 kV, 230 kV, 400 kV).'),
        ('MANIOBRA_MT', 'Maniobra en Línea MT', 'Apertura o cierre manual para seccionamiento, transferencia de carga o libranza operativa.'),
        ('OTRAS', 'Otras', 'Causas no determinadas, en proceso de investigación técnica o no tipificadas.'),
        ('POR_TERCEROS', 'Por Terceros', 'Poda no autorizada de particulares, vandalismo, hurto de conductores o impacto de maquinaria.'),
        ('PROGRAMADA', 'Programada', 'Interrupción planificada con aviso previo para ejecución de obras del Plan de Mantenimiento / Pica y Poda.'),
        ('RACIONAMIENTO', 'Racionamiento', 'Corte preventivo por déficit de generación o límites de transmisión en el SEN.'),
        ('SOBRECARGA', 'Sobrecarga', 'Apertura por demanda superior a la capacidad nominal del circuito o equipo.'),
        ('SIN_TENSION_SE', 'Sin Tensión S/E', 'Pérdida total o parcial de tensión en las barras de la Subestación de potencia.'),
        ('SOBRECORRIENTE_FASE', 'Sobrecorriente en Fase', 'Operación de relé 50/51 por corriente elevada entre fases de media tensión.'),
        ('SOBRECORRIENTE_NEUTRO', 'Sobrecorriente en el Neutro', 'Operación de relé 50N/51N por corriente residual de desbalance o falla a tierra.'),
        ('SOBRECORRIENTE_FASE_NEUTRO', 'Sobrecorriente en Fase y Neutro', 'Operación combinada de relé 50/51 y 50N/51N.'),
        ('VEGETACION', 'Vegetación', 'Contacto de ramas, enredaderas o árboles sobre conductores de media tensión.'),
        ('PAC', 'PAC (Programación de Adecuación de Carga)', 'Interrupciones por administración de carga autorizadas por el Despacho Nacional.')
    ]

    causas_headers = ['CODIGO_CAUSA', 'NOMBRE_OFICIAL_CAUSA_SEN', 'DESCRIPCION_OPERATIVA']
    causas_rows = [[c[0], c[1], c[2]] for c in causas_oficiales]

    sub_causas_data = [
        ('COMPONENTE_DANADO', 'LINEA_ROTA_MT', 'Línea Rota en MT', 'Conductor de media tensión seccionado en vano.'),
        ('COMPONENTE_DANADO', 'PUENTE_ROTO_MT', 'Puente Roto en MT', 'Rotura de puente de interconexión o derivación.'),
        ('COMPONENTE_DANADO', 'PUNTO_ROTO_MT', 'Punto Roto en MT', 'Desprendimiento de terminal o conector.'),
        ('COMPONENTE_DANADO', 'PUNTO_CALIENTE', 'Punto Caliente', 'Sobrecalentamiento por resistencia de contacto alta detectado por termografía o averiado.'),
        ('COMPONENTE_DANADO', 'RECONDUCTOR', 'Reemplazo de Cortacorriente', 'Fusible o cortacorriente de expulsión destruido.'),
        ('PROGRAMADA', 'PCV', 'Poda (Control de Vegetación)', 'Mantenimiento programado para despeje de franja de servidumbre.'),
        ('PROGRAMADA', 'MANTENIMIENTO_PODA', 'Mantenimiento Programado Tipo Poda', 'Plan preventivo institucional de pica y poda.'),
        ('PROGRAMADA', 'MANTENIMIENTO_LINEA', 'Mantenimiento Preventivo de Línea', 'Sustitución de aisladores, postes y tensado.'),
        ('ATMOSFERICA', 'DESCARGA_ATMOSFERICA', 'Descarga Atmosférica', 'Impacto de rayo sobre la infraestructura.'),
        ('ATMOSFERICA', 'LLUVIAS', 'Fuertes Lluvias en la Zona', 'Condición climática adversa extrema.'),
        ('VEGETACION', 'RAMA_MT', 'Rama sobre Líneas de MT', 'Vegetación en contacto directo provocando cortocircuito.'),
        ('POR_TERCEROS', 'TERCEROS_PODA', 'Terceros podando árbol', 'Personal no autorizado talando o podando cerca de la red.'),
        ('POR_TERCEROS', 'HURTO_MATERIAL', 'Hurto de Conductor / Componente', 'Acto vandálico contra activos de distribución.'),
        ('ACCIDENTAL', 'IMPACTO_AVE', 'Impacto de Ave sobre Líneas MT', 'Contacto de fauna silvestre entre fases o fase-tierra.'),
        ('ACCIDENTAL', 'CHOQUE_POSTE', 'Colisión Vehicular', 'Impacto de vehículo automotor contra estructura o retenida.'),
        ('ATRIBUIBLE_MANTENIMIENTO', 'AJUSTE_TORNILLERIA', 'Ajuste de Tornillería', 'Aflojamiento en grapas o herrajes.'),
        ('FALLA_LINEA_115KV', 'DESCARGA_LINEA_115KV', 'Descarga Atmosférica en Línea 115 kV', 'Falla originada en el sistema troncal de transmisión.'),
        ('OTRAS', 'DESCONOCIDA', 'Causa Desconocida', 'No se localizó la causa tras recorrido completo de la brigada.'),
        ('APERTURA_EMERGENCIA', 'PAC', 'Programación de Adecuación de Carga', 'Apertura requerida por seguridad del SEN.')
    ]
    sub_headers = ['CODIGO_CAUSA_PADRE', 'CODIGO_SUB_CAUSA', 'NOMBRE_SUB_CAUSA_OFICIAL', 'DESCRIPCION']
    sub_rows = [[s[0], s[1], s[2], s[3]] for s in sub_causas_data]

    homologacion_rules = [
        ('MANIOBRA EN LINEA MT', 'PCV', '%MANTENIMIENTO PROGRAMADO%', 'PROGRAMADA', 'PCV', 'Asignar a Programada si sub_causa es PCV o mención a mantenimiento programado'),
        ('MANIOBRA EN LINEA MT', 'LINEA ROTA', '%LINEA ROTA%', 'COMPONENTE_DANADO', 'LINEA_ROTA_MT', 'Asignar a Componente Dañado si observación o subcausa indica línea rota'),
        ('MANIOBRA EN LINEA MT', 'PUENTE ROTO', '%PUENTE ROTO%', 'COMPONENTE_DANADO', 'PUENTE_ROTO_MT', 'Asignar a Componente Dañado si observación indica puente roto'),
        ('MANIOBRA EN LINEA MT', 'PUNTO CALIENTE', '%PUNTO CALIENTE%', 'COMPONENTE_DANADO', 'PUNTO_CALIENTE', 'Asignar a Componente Dañado si se menciona punto caliente'),
        ('MANIOBRA EN LINEA MT', 'TORNILLERIA', '%AJUSTES%TORNILLERIA%', 'ATRIBUIBLE_MANTENIMIENTO', 'AJUSTE_TORNILLERIA', 'Asignar a Mantenimiento por ajuste mecánico'),
        ('MANIOBRA EN LINEA MT', '(cualquiera)', '(sin coincidencia específica)', 'APERTURA_EMERGENCIA', 'PAC', 'Fallback de maniobras no tipificadas'),
        ('SOBRECORRIENTE EN FASE', '(cualquiera)', '%FUERTES LLUVIAS%', 'ATMOSFERICA', 'LLUVIAS', 'Lluvia extrema que provocó sobrecorriente'),
        ('SOBRECORRIENTE EN FASE', '(cualquiera)', '%TERCEROS% / %PODANDO% / %NO AUTOR%', 'POR_TERCEROS', 'TERCEROS_PODA', 'Terceros interviniendo en servidumbre'),
        ('SOBRECORRIENTE EN FASE', '(cualquiera)', '%IMPACTO DE AVE% / %FAUNA%', 'ACCIDENTAL', 'IMPACTO_AVE', 'Contacto con fauna'),
        ('SOBRECORRIENTE EN FASE', '(cualquiera)', '%RAMA% / %VEGETACI%', 'VEGETACION', 'RAMA_MT', 'Contacto de ramas'),
        ('SOBRECORRIENTE EN FASE', '(cualquiera)', '%LINEA ROTA%', 'COMPONENTE_DANADO', 'LINEA_ROTA_MT', 'Rotura de conductor en fase'),
        ('SOBRECORRIENTE EN FASE', '(cualquiera)', '(resto sin detalle)', 'OTRAS', 'DESCONOCIDA', 'Disparo de fase sin causa visible'),
        ('SOBRECORRIENTE EN EL NEUTRO', '(cualquiera)', '%RAMA% / %VEGETACI%', 'VEGETACION', 'RAMA_MT', 'Falla a tierra por vegetación'),
        ('SOBRECORRIENTE EN EL NEUTRO', '(cualquiera)', '(resto)', 'OTRAS', 'DESCONOCIDA', 'Disparo de neutro sin causa visible'),
        ('PAC', '(cualquiera)', '(cualquiera)', 'APERTURA_EMERGENCIA', 'PAC', 'Apertura de emergencia por adecuación de carga'),
        ('SIN TENSION S/E 115 KV', '(cualquiera)', '%DESCARGA ATMOSFERICA%', 'FALLA_LINEA_115KV', 'DESCARGA_LINEA_115KV', 'Falla troncal 115 kV por tormenta'),
        ('SIN TENSION S/E 115 KV', '(cualquiera)', '(resto)', 'FALLA_LINEA_115KV', 'DESCONOCIDA', 'Pérdida de barra 115 kV'),
        ('SOBRECORRIENTE EN FASE Y NEUTRO', '(cualquiera)', '(cualquiera)', 'OTRAS', 'DESCONOCIDA', 'Falla bifásica / trifásica a tierra no clasificada'),
        ('(VACÍO / NULL / ESPACIO)', '(cualquiera)', '(cualquiera)', 'OTRAS', 'DESCONOCIDA', 'Registro con causa omitida en despacho')
    ]
    homolog_headers = ['CAUSA_TEXTO_BRUTO', 'SUB_CAUSA_BRUTO', 'PATRON_OBSERVACION', 'CAUSA_OFICIAL_DESTINO', 'SUB_CAUSA_OFICIAL_DESTINO', 'REGLA_LOGICA']
    homolog_rows = [[h[0], h[1], h[2], h[3], h[4], h[5]] for h in homologacion_rules]

    sintax_fixes = [
        ('ABIERTO', 'Apertura', 'Normalización de verbo operativo'),
        ('DISPARO', 'Disparo', 'Capitalización canónica'),
        ('SEÑALIZANDO', 'señalizando', 'Minúscula en texto continuo'),
        ('DESCONIDA', 'desconocida', 'Corrección de error tipográfico (Monagas / Lara)'),
        ('SOBRECORREINTE', 'sobrecorriente', 'Corrección ortográfica'),
        ('VEGETACIÒN', 'vegetación', 'Corrección de acento grave a agudo'),
        ('08,:54:00,', '08:54:00', 'Eliminación de comas espurias de Crystal Reports (Monagas)'),
        ('18:00,:00,', '18:00:00', 'Eliminación de comas espurias'),
        ('  (doble espacio)', ' (espacio simple)', 'Compresión de espacios en blanco múltiples')
    ]
    sintax_headers = ['PATRON_ERRONEO_ORIGINAL', 'CORRECCION_NORMALIZADA', 'DESCRIPCION_ERROR']
    sintax_rows = [[s[0], s[1], s[2]] for s in sintax_fixes]

    wb_causas = openpyxl.Workbook()
    ws_c1 = wb_causas.active
    ws_c1.title = '01_CAUSAS_OFICIALES'
    populate_sheet(ws_c1, causas_headers, causas_rows, 'CATÁLOGO OFICIAL DE 22 CAUSAS NORMALIZADAS DEL SEN (GGPD)')

    ws_c2 = wb_causas.create_sheet(title='02_SUB_CAUSAS')
    populate_sheet(ws_c2, sub_headers, sub_rows, 'CATÁLOGO OFICIAL DE SUB-CAUSAS DE DISTRIBUCIÓN (GGPD)')

    ws_c3 = wb_causas.create_sheet(title='03_MATRIZ_HOMOLOGACION')
    populate_sheet(ws_c3, homolog_headers, homolog_rows, 'MATRIZ DETERMINISTA DE HOMOLOGACIÓN SEMÁNTICA (DE TEXTO BRUTO A CAUSA OFICIAL)')

    ws_c4 = wb_causas.create_sheet(title='04_CORRECCIONES_SINTACTICAS')
    populate_sheet(ws_c4, sintax_headers, sintax_rows, 'DICCIONARIO DE CORRECCIÓN SINTÁCTICA Y TIPOGRÁFICA ISO 8000')

    wb_causas.save(os.path.join(output_dir, 'CATALOGO_MAESTRO_CAUSAS_HOMOLOGACION.xlsx'))
    print('3. Causas y Homologación generadas.')

    # 4. TABLA MAESTRA DE ESTADOS, REGIONES Y ARCHIVOS DE ENTRADA
    estados_regiones_data = [
        ('AMA', 'AMAZONAS', 'GUAYANA', 'AMAZONAS26.xlsx', 'DISTRIBUCION', 1, 'Formato Amazonas (Distribución)', 'Cabecera estándar con FECHA FALLA'),
        ('ANZ', 'ANZOATEGUI', 'ORIENTAL', 'ANZOATEGUI26.xlsx', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Columna TTI y señal'),
        ('APU', 'APURE', 'LOS_LLANOS', 'APURE26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario Excel 97-2003'),
        ('ARA', 'ARAGUA', 'CENTRAL', 'ARAGUA26.xls', 'DETALLE', 1, 'Formato TIRAS Estándar', 'Hoja DETALLE en vez de DISTRIBUCION'),
        ('BAR', 'BARINAS', 'LOS_LLANOS', 'BARINAS26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario Excel 97-2003'),
        ('BOL', 'BOLIVAR', 'GUAYANA', 'BOLIVAR26.xlsx', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Formato estándar'),
        ('CAP', 'DISTRITO CAPITAL', 'CAPITAL', 'CAPITAL26.xlsx', 'Interrupciones_', 3, 'Formato Capital (Detección Fila 3)', 'Membrete en Filas 1-2, cabecera en fila 3'),
        ('CAR', 'CARABOBO', 'CENTRAL', 'CARABOBO26.xls', 'DISTR', 1, 'Formato TIRAS Estándar', 'Hoja nombrada DISTR'),
        ('COJ', 'COJEDES', 'LOS_LLANOS', 'COJEDES26.xlsx', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Formato estándar'),
        ('DEL', 'DELTA AMACURO', 'GUAYANA', 'DELTA26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('FAL', 'FALCON', 'OCCIDENTAL', 'FALCON26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('GUA', 'GUARICO', 'LOS_LLANOS', 'GUARICO26.xlsx', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Formato estándar'),
        ('LAR', 'LARA', 'OCCIDENTAL', 'LARA26.xlsx', 'DISTRIBUCION', 1, 'Formato TIRAS Multi-Hoja', 'Multi-pestaña: seleccionar DISTRIBUCION'),
        ('VAR', 'LA GUAIRA', 'CAPITAL', 'LGUA26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('MER', 'MERIDA', 'LOS_ANDES', 'MERIDA26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('MIR_GUA', 'MIRANDA (GUARENAS)', 'CAPITAL', 'MIRANDAGUARENAS26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Sub-región Guarenas-Guatire'),
        ('MIR_LTQ', 'MIRANDA (ALTOS MIRANDINOS)', 'CAPITAL', 'MIRANDALTQ26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Sub-región Los Teques'),
        ('MIR_TUY', 'MIRANDA (VALLES DEL TUY)', 'CAPITAL', 'MIRANDATUY26.XLS', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Sub-región Valles del Tuy'),
        ('MON', 'MONAGAS', 'ORIENTAL', 'MONAGAS26.xls', 'DISTRIBUCION', 15, 'Formato Monagas (Crystal Reports)', 'Membrete 14 filas, cabecera en fila 15, comas en horas'),
        ('NEV', 'NUEVA ESPARTA', 'INSULAR', 'NUEVA_ESP26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('POR', 'PORTUGUESA', 'LOS_LLANOS', 'PORTUGUESA26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('SUC', 'SUCRE', 'ORIENTAL', 'SUCRE26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('TAC', 'TACHIRA', 'LOS_ANDES', 'TACHIRA26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('TRU', 'TRUJILLO', 'LOS_ANDES', 'TRUJILLO26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 97-2003'),
        ('YAR', 'YARACUY', 'CENTRAL', 'YARACUY26.xlsx', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Formato estándar'),
        ('ZUL', 'ZULIA', 'OCCIDENTAL', 'ZUL_TOTZ_TTI_26.xls', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Binario 10.1 MB con 2,498 registros'),
        ('GEQ', 'GUAYANA ESEQUIBA', 'GUAYANA', 'GUAYANA_ESEQ26.xlsx', 'DISTRIBUCION', 1, 'Formato TIRAS Estándar', 'Entidad Territorial 25')
    ]
    edos_headers = [
        'CODIGO_ESTADO', 'NOMBRE_ESTADO', 'REGION_ADMINISTRATIVA', 'NOMBRE_ARCHIVO_DRIVE',
        'HOJA_OPERATIVA_DEFAULT', 'FILA_CABECERA_ORIGEN', 'MODELO_FORMATO_DETECTADO', 'PARTICULARIDAD_TECNICA'
    ]
    edos_rows = [[e[0], e[1], e[2], e[3], e[4], e[5], e[6], e[7]] for e in estados_regiones_data]

    wb_edos = openpyxl.Workbook()
    ws_ed = wb_edos.active
    ws_ed.title = 'ESTADOS_Y_REGIONES_SEN'
    populate_sheet(ws_ed, edos_headers, edos_rows, 'CORPOELEC GGPD — MATRIZ DE ENTIDADES FEDERALES, REGIONES Y PARÁMETROS DE INGESTA')
    wb_edos.save(os.path.join(output_dir, 'CATALOGO_MAESTRO_ESTADOS_REGIONES.xlsx'))
    print('4. Estados y Regiones generados.')

    # 5. MAPEO DE FORMATOS ESTADALES A ESTRUCTURA CANÓNICA SCTIS (A a U)
    mapeo_columnas_data = [
        ('A', 'ESTADO', 'estado_nombre', 'Texto (ej. AMAZONAS, ARAGUA)', 'DISTRITO / ESTADO / (Nombre del archivo)', 'Normalizar a mayúsculas sin tildes'),
        ('B', 'SISTEMA', 'sistema', 'Texto (DISTRIBUCION / TRANSMISION)', 'SISTEMA / SIST', 'Default a "DISTRIBUCION" si está vacío'),
        ('C', 'JEFATURA', 'jefatura', 'Texto (Distrito / Jefatura)', 'JEFATURA / DISTRITO / SECTOR', 'Asignar distrito operativo'),
        ('D', 'SUBESTACION', 'subestacion', 'Texto (Nombre canónico S/E)', 'S/E / SUBESTACION / SUB-ESTACION', 'Remover prefijos S/E, tensiones 115/13.8kV'),
        ('E', 'CIRCUITO', 'circuito', 'Texto (Nombre canónico CTO)', 'CIRCUITO / ALIMENTADOR', 'Remover números de posición entre paréntesis'),
        ('F', 'FECHA_INICIO', 'fecha_inicio', 'Fecha ISO (YYYY-MM-DD)', 'FECHA / FECHA INI / FECHA FALLA', 'Extraer componente fecha, formato YYYY-MM-DD'),
        ('G', 'HORA_INICIO', 'hora_inicio', 'Hora 24h (HH:MM:SS)', 'HORA INICIO / INICIO / HORA', 'Limpiar comas, convertir am/pm a 24 horas'),
        ('H', 'FECHA_FIN', 'fecha_fin', 'Fecha ISO (YYYY-MM-DD)', 'FECHA FIN / FECHA REST', 'Si vacía o nula, marcar como incompleto'),
        ('I', 'HORA_FIN', 'hora_fin', 'Hora 24h (HH:MM:SS)', 'HORA FIN / RESTITUCION / FIN', 'Limpiar comas, convertir am/pm a 24 horas'),
        ('J', 'DURACION', 'duracion_str', 'Texto (HH:MM:SS o horas)', 'DURACION / TIEMPO / T_TOTAL', 'Cálculo de diferencia temporal fin - inicio'),
        ('K', 'CARGA_MW', 'carga', 'Numérico decimal (MW / MVA)', 'CARGA / CARGA MW / KVA', 'Potencia interrumpida'),
        ('L', 'FRECUENCIA_HZ', 'frec', 'Numérico decimal (Hz)', 'FREC / FRECUENCIA', 'Default 60.0 Hz si no existe'),
        ('M', 'DURACION_HORAS', 'horas', 'Numérico decimal (Horas)', 'T.HORAS / HORAS / TOTAL HORAS', 'Calculado: (dt_fin - dt_ini).total_seconds() / 3600.0'),
        ('N', 'TTI_CTO', 'tti_cto', 'Numérico decimal (MWh o min)', 'TTI_CTO / TTI / INDICE', 'Tiempo Total de Interrupción'),
        ('O', 'SENAL_PROTECCION', 'senal', 'Texto (Relé / Bandera)', 'SEÑAL / SEÑALIZACION / RELÉ', '50/51, 50N, 87T, etc.'),
        ('P', 'CAUSA_HOMOLOGADA', 'causa', 'Texto (Catálogo 22 causas SEN)', 'CAUSA / MOTIVO', 'Homologar con matriz semántica de causas'),
        ('Q', 'SUB_CAUSA', 'sub_causa', 'Texto (Catálogo sub-causas)', 'SUB CAUSA / (SUB-CAUSA) / DETALLE', 'Tipificar según matriz de subcausas'),
        ('R', 'OBSERVACION', 'observacion', 'Texto descriptivo', 'OBSERVACION / OBSERVACIONES', 'Saneamiento sintáctico y corrección tipográfica'),
        ('S', 'SECTORES_AFECTADOS', 'sectores', 'Texto (Comunidades)', 'SECTORES / POBLACION / AFECTADOS', 'Listado de zonas sin servicio'),
        ('T', 'CIUDAD_MUNICIPIO', 'ciudad', 'Texto (Municipio / Ciudad)', 'CIUDAD / MUNICIPIO / PARROQUIA', 'Ubicación geográfica'),
        ('U', 'KVA_INSTALADOS', 'kva', 'Numérico decimal (kVA)', 'KVA / POTENCIA_KVA / CAPACIDAD', 'Capacidad nominal instalada')
    ]
    mapeo_headers = [
        'COL_CANONICA', 'NOMBRE_CAMPO_SCTIS', 'VARIABLE_INTERNA', 'TIPO_DATO_ISO8000',
        'ENCABEZADOS_EQUIVALENTES_ORIGEN', 'REGLA_DE_TRANSFORMACION_Y_NORMALIZACION'
    ]
    mapeo_rows = [[m[0], m[1], m[2], m[3], m[4], m[5]] for m in mapeo_columnas_data]

    wb_map = openpyxl.Workbook()
    ws_m = wb_map.active
    ws_m.title = 'MAPEO_COLUMNAS_SCTIS'
    populate_sheet(ws_m, mapeo_headers, mapeo_rows, 'CORPOELEC GGPD — MATRIZ DE NORMALIZACIÓN DE COLUMNAS (ISO 8000-110)')
    wb_map.save(os.path.join(output_dir, 'CATALOGO_MAESTRO_MAPEO_FORMATOS_ESTADALES.xlsx'))
    print('5. Mapeo de formatos generado.')

    # 6. PLANTILLA CANÓNICA NORMALIZADA CON EJEMPLOS
    template_headers = [
        'ESTADO', 'SISTEMA', 'JEFATURA', 'SUBESTACION', 'CIRCUITO',
        'FECHA_INICIO', 'HORA_INICIO', 'FECHA_FIN', 'HORA_FIN', 'DURACION',
        'CARGA_MW', 'FRECUENCIA_HZ', 'DURACION_HORAS', 'TTI_CTO', 'SENAL_PROTECCION',
        'CAUSA_HOMOLOGADA', 'SUB_CAUSA', 'OBSERVACION', 'SECTORES_AFECTADOS', 'CIUDAD_MUNICIPIO', 'KVA_INSTALADOS'
    ]
    template_sample_rows = [
        ['AMAZONAS', 'DISTRIBUCION', 'PUERTO AYACUCHO', 'PUERTO AYACUCHO 115 KV', 'CENTRO', '2026-08-01', '14:20:00', '2026-08-01', '16:45:00', '02:25:00', 4.5, 60.0, 2.4167, 10.875, '50/51', 'VEGETACION', 'RAMA_MT', 'Disparo de interruptor por rama sobre línea de media tensión. Seccionamiento y poda de emergencia.', 'Casco Central, Av. Río Negro, Sector La Granja', 'PUERTO AYACUCHO', 5000],
        ['ARAGUA', 'DISTRIBUCION', 'MARACAY SUR', 'DELICIAS', 'LAS DELICIAS', '2026-08-02', '08:15:00', '2026-08-02', '09:30:00', '01:15:00', 6.2, 59.9, 1.2500, 7.750, '50N', 'ATMOSFERICA', 'DESCARGA_ATMOSFERICA', 'Descarga atmosférica provocó actuación de fusible cortacorriente.', 'Urb. Las Delicias, Calicanto, Av. Principal', 'MARACAY', 7500],
        ['MONAGAS', 'DISTRIBUCION', 'MATURIN', 'INDIO MATURIN', 'LOS CORTIJOS', '2026-08-03', '10:05:00', '2026-08-03', '12:50:00', '02:45:00', 3.8, 60.0, 2.7500, 10.450, '50/51', 'COMPONENTE_DANADO', 'PUENTE_ROTO_MT', 'Puente roto en media tensión en estructura N° 45. Reparación y reconexión.', 'Sector Los Cortijos, La Muralla', 'MATURIN', 4000],
        ['ZULIA', 'DISTRIBUCION', 'MARACAIBO NORTE', 'BELLA VISTA', 'SANTA RITA', '2026-08-04', '18:00:00', '2026-08-04', '20:30:00', '02:30:00', 8.5, 60.0, 2.5000, 21.250, '87T', 'PROGRAMADA', 'PCV', 'Mantenimiento programado tipo poda en circuito Santa Rita.', 'Av. Santa Rita, Calle 72, Sector Paraíso', 'MARACAIBO', 10000],
        ['DISTRITO CAPITAL', 'DISTRIBUCION', 'CENTRO', 'SANTA ROSA', 'BELLAS ARTES', '2026-08-05', '11:10:00', '2026-08-05', '11:45:00', '00:35:00', 5.0, 60.1, 0.5833, 2.917, '50/51', 'ACCIDENTAL', 'IMPACTO_AVE', 'Impacto de ave sobre conductores de salida en pórtico.', 'Bellas Artes, Parque Central, Av. México', 'CARACAS', 8000]
    ]

    wb_temp = openpyxl.Workbook()
    ws_t = wb_temp.active
    ws_t.title = 'FORMULARIO_NORMALIZADO_SCTIS'
    populate_sheet(ws_t, template_headers, template_sample_rows, 'CORPOELEC GGPD — FORMULARIO DE CONSOLIDACIÓN NORMALIZADO SCTIS v2.0 (21 COLUMNAS)')
    wb_temp.save(os.path.join(output_dir, 'PLANTILLA_SCTIS_CONSOLIDADO_NORMALIZADO.xlsx'))
    print('6. Plantilla normalizada generada.')

    # 7. LIBRO MAESTRO UNIFICADO MULTI-HOJA TODO-EN-UNO
    wb_uni = openpyxl.Workbook()
    
    # Sheet 1: Subestaciones
    ws_u1 = wb_uni.active
    ws_u1.title = '01_SUBESTACIONES_SEN'
    populate_sheet(ws_u1, se_headers, se_rows, 'CATÁLOGO MAESTRO DE SUBESTACIONES DEL SEN (871 REGISTROS)')

    # Sheet 2: Circuitos
    ws_u2 = wb_uni.create_sheet(title='02_CIRCUITOS_SEN')
    populate_sheet(ws_u2, cto_headers, cto_rows, 'CATÁLOGO MAESTRO DE CIRCUITOS DE DISTRIBUCIÓN DEL SEN (4,207 REGISTROS)')

    # Sheet 3: Causas Oficiales
    ws_u3 = wb_uni.create_sheet(title='03_CAUSAS_OFICIALES')
    populate_sheet(ws_u3, causas_headers, causas_rows, 'CATÁLOGO OFICIAL DE 22 CAUSAS NORMALIZADAS (GGPD)')

    # Sheet 4: Sub Causas
    ws_u4 = wb_uni.create_sheet(title='04_SUB_CAUSAS')
    populate_sheet(ws_u4, sub_headers, sub_rows, 'CATÁLOGO DE SUB-CAUSAS')

    # Sheet 5: Matriz Homologación
    ws_u5 = wb_uni.create_sheet(title='05_MATRIZ_HOMOLOGACION')
    populate_sheet(ws_u5, homolog_headers, homolog_rows, 'MATRIZ DE HOMOLOGACIÓN SEMÁNTICA DETERMINISTA')

    # Sheet 6: Estados y Regiones
    ws_u6 = wb_uni.create_sheet(title='06_ESTADOS_Y_REGIONES')
    populate_sheet(ws_u6, edos_headers, edos_rows, 'MATRIZ DE ENTIDADES FEDERALES, REGIONES Y ARCHIVOS DE ORIGEN')

    # Sheet 7: Mapeo Columnas
    ws_u7 = wb_uni.create_sheet(title='07_MAPEO_COLUMNAS')
    populate_sheet(ws_u7, mapeo_headers, mapeo_rows, 'MAPEO DE COLUMNAS ESTADALES A FORMATO CANÓNICO SCTIS')

    # Sheet 8: Plantilla Estructura
    ws_u8 = wb_uni.create_sheet(title='08_PLANTILLA_CANONICA')
    populate_sheet(ws_u8, template_headers, template_sample_rows, 'ESTRUCTURA DE CONSOLIDACIÓN SEMANAL Y MENSUAL (21 COLUMNAS)')

    unificado_path = os.path.join(output_dir, 'CATALOGO_MAESTRO_UNIFICADO_SCTIS_SEN.xlsx')
    wb_uni.save(unificado_path)
    print(f'7. Libro maestro unificado generado exitosamente en: {unificado_path}')

if __name__ == '__main__':
    build_spark_catalogs()

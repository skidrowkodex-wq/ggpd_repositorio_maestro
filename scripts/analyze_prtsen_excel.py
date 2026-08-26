import openpyxl
import pandas as pd
import numpy as np

file_path = '/home/skidrowkodex/Documentos/Repositorio_Maestro/data/PRTSEN_PRT_DISTRIBUCION_FICHA_CONSOLIDADO.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
print("Hojas en el libro:", wb.sheetnames)

sheet_name = 'PRTSEN'
sheet = wb[sheet_name]

data = list(sheet.iter_rows(values_only=True))
headers = [str(h) if h is not None else f"COL_{i}" for i, h in enumerate(data[0])]
rows = data[1:]

df = pd.DataFrame(rows, columns=headers)
print(f"\n=== ANÁLISIS DE LA HOJA: {sheet_name} ===")
print(f"Total Filas de Proyectos: {len(df)}")
print(f"Total Columnas: {len(df.columns)}")

print("\n--- DETALLE DE TODAS LAS COLUMNAS ---")
for i, col in enumerate(df.columns):
    valid_count = int(df[col].notna().sum())
    series_clean = df[col].dropna()
    sample = series_clean.iloc[0] if len(series_clean) > 0 else 'VACÍA'
    print(f"{i+1:02d}. [{col[:35]:<35}] -> No-Nulos: {valid_count:4d}/{len(df)} | Muestra: {str(sample)[:40]}")

print("\n--- TIPO DE INSTALACIÓN ---")
tipo_col = [c for c in df.columns if 'TIPO DE INSTALACION' in str(c).upper()]
if tipo_col:
    print(df[tipo_col[0]].value_counts(dropna=False))

print("\n--- ESTATUS DE LOS PROYECTOS ---")
estatus_col = [c for c in df.columns if 'ESTATUS' in str(c).upper()]
if estatus_col:
    print(df[estatus_col[0]].value_counts(dropna=False))

print("\n--- RESUMEN FINANCIERO (MONTO INVERSIÓN) ---")
monto_cols = [c for c in df.columns if 'MONTO' in str(c).upper() or 'INVERSION' in str(c).upper()]
for c in monto_cols:
    s = pd.to_numeric(df[c], errors='coerce')
    print(f"Columna [{c}]:")
    print(f"  Proyectos con monto válido: {int(s.notna().sum())}")
    print(f"  Inversión Total: ${s.sum():,.2f} USD")
    print(f"  Promedio por Proyecto: ${s.mean():,.2f} USD")
    print(f"  Rango: Min ${s.min():,.2f} | Max ${s.max():,.2f} USD")

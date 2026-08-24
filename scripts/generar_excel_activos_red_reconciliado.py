#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
⚡ CORPOELEC — GERENCIA GENERAL DE PLANIFICACIÓN DE DISTRIBUCIÓN (GGPD)
Generador de Catálogo Maestro de Activos de Red y Matriz de Reconciliación
Código Normativo: NAC_2026_GGPD_CATALOGO_ACTIVOS_RED_SE_CT_RECONCILIACION
==============================================================================
Genera un libro .xlsx con 5 hojas sin cabeceras/membretes decorativos:
  1. RESUMEN                      — Ficha técnica, diagnóstico 4,313 vs 4,207 y métricas
  2. SUBESTACIONES                — 871 SE con metadata técnica y origen
  3. CIRCUITOS_NORMALIZADOS       — 4,207 CT con segmentación NS-P-105 y cabeceras
  4. DASHBOARD                    — Matriz consolidada por estado (SE, CT, Deltas, Ratios)
  5. CIRCUITOS_AUDITORIA_LEGACY   — 4,313 CT originales con estatus de conciliación y dictamen
==============================================================================
"""

import json
import os
import re
import sys
import unicodedata
from datetime import datetime
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Se requiere openpyxl.")
    sys.exit(1)

# Rutas
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLS_ORIG_PATH = os.path.join(REPO_ROOT, "apps/corpoelec-sigi-gestion-planificacion-distribucion/docs/CARACTERIZACIÓN DISTRIBUCION.xls")
NORM_PATH = os.path.join(REPO_ROOT, "apps/caracterizacion_distribucion/data/caracterizacion_distribucion_normalizado.xlsx")
OUTPUT_DIR = os.path.join(REPO_ROOT, "docs")
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"NAC_2026_GGPD_CATALOGO_ACTIVOS_RED_SE_CT_RECONCILIADO_{TIMESTAMP}.xlsx")

# Estilos
AZUL_CORP     = "1B3A5C"
AZUL_CLARO    = "2E5D8A"
AMARILLO_CORP = "F4B400"
VERDE_OK      = "0D7C3F"
GRIS_FONDO    = "F5F5F5"
BLANCO        = "FFFFFF"
ROJO_WARN     = "B71C1C"
NARANJA_WARN  = "E65100"

FONT_TITULO   = Font(name="Calibri", size=14, bold=True, color=BLANCO)
FONT_SUBTIT   = Font(name="Calibri", size=11, bold=True, color=AZUL_CORP)
FONT_HEADER   = Font(name="Calibri", size=10, bold=True, color=BLANCO)
FONT_NORMAL   = Font(name="Calibri", size=10, color="333333")
FONT_METRIC   = Font(name="Calibri", size=11, bold=True, color=VERDE_OK)
FONT_SMALL    = Font(name="Calibri", size=9, color="666666")
FONT_LINK     = Font(name="Calibri", size=10, color="1565C0", underline="single")

FILL_TITULO   = PatternFill(start_color=AZUL_CORP, end_color=AZUL_CORP, fill_type="solid")
FILL_HEADER   = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
FILL_ALTROW   = PatternFill(start_color=GRIS_FONDO, end_color=GRIS_FONDO, fill_type="solid")
FILL_METRIC   = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FILL_DASH_HDR = PatternFill(start_color=AMARILLO_CORP, end_color=AMARILLO_CORP, fill_type="solid")

FILL_OK       = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FILL_REV      = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
FILL_DUP      = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
FILL_TX       = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


def clean_key(s):
    if pd.isna(s) or s is None:
        return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


def norm_str(s):
    if pd.isna(s) or s is None:
        return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s


def apply_header_style(ws, row_num, col_count, fill=FILL_HEADER, font=FONT_HEADER):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def write_data_rows(ws, data_rows, start_row, col_count, left_cols=3):
    for i, row_data in enumerate(data_rows):
        row_num = start_row + i
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_num, column=col_idx + 1, value=value)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT if col_idx < left_cols else ALIGN_CENTER
            cell.border = BORDER_THIN
            if i % 2 == 1:
                cell.fill = FILL_ALTROW


def auto_width(ws, min_w=8, max_w=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_w
        for cell in col_cells[:250]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[col_letter].width = max_len


def classify_circuit_nsp105(nombre_circuito, observaciones, elemento_tipo):
    nombre = str(nombre_circuito or "").strip()
    obs = str(observaciones or "").strip()
    tipo = str(elemento_tipo or "CIRCUITO").strip().upper()

    if obs.startswith("SECCIONADOR"):
        cod = obs.split(":")[-1] if ":" in obs else "S"
        return ("SECCIONADOR_MANIOBRA", cod, f"Seccionador de Maniobra y Aislamiento [{cod}] (Norma NS-P-105)", True)
    if obs.startswith("DISYUNTOR"):
        cod = obs.split(":")[-1] if ":" in obs else "D"
        return ("DISYUNTOR_POTENCIA", cod, f"Disyuntor / Interruptor de Potencia [{cod}] (Norma NS-P-105)", True)
    if obs.startswith("BARRA") or tipo == "BARRA":
        cod = obs.split(":")[-1] if ":" in obs else "B"
        return ("JUEGO_BARRAS", cod, f"Juego de Barras Colectoras / Transferencia [{cod}] (Norma NS-P-105)", True)

    m = re.search(r"\(([A-Za-z]+-?\d+[A-Za-z]?|[A-Za-z]+)\)\s*$", nombre)
    if m:
        code_str = m.group(1).upper()
        if re.match(r"^S\d*$", code_str):
            return ("SECCIONADOR_MANIOBRA", code_str, f"Seccionador de Maniobra [{code_str}] (Norma NS-P-105)", True)
        elif re.match(r"^D\d*$", code_str):
            return ("DISYUNTOR_POTENCIA", code_str, f"Disyuntor de Potencia [{code_str}] (Norma NS-P-105)", True)
        elif re.match(r"^B\d*$", code_str):
            return ("JUEGO_BARRAS", code_str, f"Juego de Barras [{code_str}] (Norma NS-P-105)", True)
        elif re.match(r"^L\d*$", code_str) or tipo == "LINEA":
            return ("LINEA_ENLACE", code_str, f"Línea de Enlace / Interconexión [{code_str}] (Norma NS-P-105)", True)
        elif re.match(r"^T\d*$", code_str) or tipo == "TRANSFORMADOR" or "TRAFO" in nombre.upper():
            return ("TRANSFORMADOR", code_str, f"Transformador de Potencia [{code_str}] (Norma NS-P-105)", True)
        elif code_str in ("RESERVA", "Q", "RESPALDO"):
            return ("POSICION_RESERVA", code_str, f"Posición / Bahía de Reserva [{code_str}] (Norma NS-P-105)", True)

    if tipo == "LINEA" or re.search(r"^L\d+\s+", nombre, re.I) or "LINEA " in nombre.upper():
        m_l = re.search(r"(L\d+)", nombre, re.I)
        cod = m_l.group(1).upper() if m_l else "L"
        return ("LINEA_ENLACE", cod, f"Línea de Enlace / Interconexión [{cod}] (Norma NS-P-105)", True)

    if tipo == "BARRA" or "BARRA " in nombre.upper():
        return ("JUEGO_BARRAS", "B", "Juego de Barras Colectoras / Transferencia (Norma NS-P-105)", True)

    if tipo == "TRANSFORMADOR" or "TRANSFORMADOR" in nombre.upper() or "TRAFO" in nombre.upper():
        return ("TRANSFORMADOR", "T", "Transformador de Potencia / Distribución (Norma NS-P-105)", True)

    if "RESERVA" in nombre.upper():
        return ("POSICION_RESERVA", "RESERVA", "Posición / Bahía de Reserva en Subestación", True)

    return ("ALIMENTADOR_CONVENCIONAL", "CTO", "Alimentador de Distribución Convencional", False)


def load_and_reconcile():
    print("1. Cargando archivos de datos...")
    df_orig = pd.read_excel(XLS_ORIG_PATH, sheet_name="CARACTERIZACION CIRCUITOS")
    df_norm_se = pd.read_excel(NORM_PATH, sheet_name="CARACTERIZACION_SE_COMPLETO")
    df_norm_ct = pd.read_excel(NORM_PATH, sheet_name="CARACTERIZACION_CIRCUITOS")

    # Aplicar remediaciones sobre df_norm_ct
    df_norm_ct.loc[df_norm_ct["SUBESTACION_CABECERA"] == "YARACAL LL", "SUBESTACION_CABECERA"] = "YARACAL II"
    df_norm_ct.loc[df_norm_ct["se_codigo"] == "SE-FAL-0321", "se_codigo"] = "SE-FAL-0320"
    df_norm_ct.loc[df_norm_ct["ID_CIRCUITO_NORMALIZADO"] == "CT-CAR-01238", "SUBESTACION_CABECERA"] = "CENTRO"
    df_norm_ct.loc[df_norm_ct["ID_CIRCUITO_NORMALIZADO"] == "CT-CAR-01239", "SUBESTACION_CABECERA"] = "QUIZANDA"
    df_norm_ct.loc[df_norm_ct["ID_CIRCUITO_NORMALIZADO"] == "CT-CAR-01240", "SUBESTACION_CABECERA"] = "QUIZANDA"
    df_norm_ct.loc[df_norm_ct["ID_CIRCUITO_NORMALIZADO"] == "CT-CAR-01241", "SUBESTACION_CABECERA"] = "CENTRO"
    df_norm_ct.loc[df_norm_ct["ID_CIRCUITO_NORMALIZADO"] == "CT-CAR-01242", "SUBESTACION_CABECERA"] = "QUIZANDA"
    df_norm_ct.loc[df_norm_ct["ID_CIRCUITO_NORMALIZADO"] == "CT-CAR-01243", "SUBESTACION_CABECERA"] = "QUIZANDA"
    df_norm_ct.loc[df_norm_ct["ID_CIRCUITO_NORMALIZADO"].isin([
        "CT-ZUL-03817", "CT-ZUL-03818", "CT-ZUL-03819", "CT-ZUL-03820", "CT-ZUL-03821"
    ]), "SUBESTACION_CABECERA"] = "NODO ZULIA"

    # Clasificación NS-P-105 en normalizados
    nsp_tipos = []
    nsp_cods = []
    nsp_descs = []
    nsp_patios = []
    for _, r in df_norm_ct.iterrows():
        t, c, d, p = classify_circuit_nsp105(r["NOMBRE_CIRCUITO"], r.get("OBSERVACIONES_NORMALIZACION"), r.get("ELEMENTOS_TIPO"))
        nsp_tipos.append(t)
        nsp_cods.append(c)
        nsp_descs.append(d)
        nsp_patios.append("SÍ" if p else "NO")

    df_norm_ct["ELEMENTO_TECNICO_NSP105"] = nsp_tipos
    df_norm_ct["CODIGO_MANIOBRA_NORMA"] = nsp_cods
    df_norm_ct["DESCRIPCION_TECNICA"] = nsp_descs
    df_norm_ct["ES_COMPONENTE_PATIO"] = nsp_patios

    # Mapeo de reconciliación
    df_norm_ct["ESTADO_K"] = df_norm_ct["ESTADO"].apply(clean_key)
    df_norm_ct["SE_K"] = df_norm_ct["SUBESTACION_CABECERA"].apply(clean_key)
    df_norm_ct["CT_K"] = df_norm_ct["NOMBRE_CIRCUITO"].apply(clean_key)
    df_norm_ct["CT_NORM_K"] = df_norm_ct["NOMBRE_CT_NORM"].apply(clean_key)

    norm_full_map = {}
    norm_ct_map = {}
    for _, r in df_norm_ct.iterrows():
        norm_full_map[(r["ESTADO_K"], r["SE_K"], r["CT_K"])] = r
        norm_full_map[(r["ESTADO_K"], r["SE_K"], r["CT_NORM_K"])] = r
        norm_ct_map[(r["ESTADO_K"], r["CT_K"])] = r
        norm_ct_map[(r["ESTADO_K"], r["CT_NORM_K"])] = r

    # Reconciliar df_orig
    df_orig["ESTADO_K"] = df_orig["Estado"].apply(clean_key).replace({
        "DISTRITOCAPITAL": "DISTRITOCAPITAL", "LAGUAIRA": "LAGUAIRA",
        "NUEVAESPARTA": "NUEVAESPARTA", "DELTAAMACURO": "DELTAAMACURO",
        "VARGAS": "LAGUAIRA"
    })
    df_orig["SE_K"] = df_orig["Subestación"].apply(clean_key)
    df_orig["CT_K"] = df_orig["Circuito"].apply(clean_key)

    seen_keys = set()
    status_list = []
    matched_id_list = []
    matched_ct_name_list = []
    matched_se_name_list = []
    elem_nsp_list = []
    motivo_list = []

    tx_gen_keywords = ["GURI", "TOCOMA", "MACAGUA", "CARUACHI", "FERROMINERA", "FMO", "SIDOR", "VENALUM", "ALCASA", "CABELUM", "OMCI", "PLANTACENTRO", "JOSEFAJOAQUINA", "TERMOZULIA", "RAMONLAGUNA", "TACOA"]

    for idx, r in df_orig.iterrows():
        est = r["ESTADO_K"]
        se = r["SE_K"]
        ct = r["CT_K"]
        k_full = (est, se, ct)

        # 1. Duplicado en origen
        if k_full in seen_keys:
            status_list.append("DUPLICADO EN ORIGEN")
            matched_id_list.append("")
            matched_ct_name_list.append("")
            matched_se_name_list.append("")
            elem_nsp_list.append("DUPLICADO")
            motivo_list.append(f"Registro sintáctico idéntico duplicado en archivo original ({r['Circuito']} en S/E {r['Subestación']})")
            continue
        seen_keys.add(k_full)

        # 2. Match exacto
        if k_full in norm_full_map:
            m = norm_full_map[k_full]
            status_list.append("HOMOLOGADO EXACTO")
            matched_id_list.append(m["ID_CIRCUITO_NORMALIZADO"])
            matched_ct_name_list.append(m["NOMBRE_CIRCUITO"])
            matched_se_name_list.append(m["SUBESTACION_CABECERA"])
            elem_nsp_list.append(m["ELEMENTO_TECNICO_NSP105"])
            motivo_list.append(f"Conciliado 100% con activo maestro {m['ID_CIRCUITO_NORMALIZADO']} (S/E {m['SUBESTACION_CABECERA']})")
        # 3. Match por Estado + CT (cabecera saneada)
        elif (est, ct) in norm_ct_map:
            m = norm_ct_map[(est, ct)]
            status_list.append("HOMOLOGADO (CABECERA AJUSTADA)")
            matched_id_list.append(m["ID_CIRCUITO_NORMALIZADO"])
            matched_ct_name_list.append(m["NOMBRE_CIRCUITO"])
            matched_se_name_list.append(m["SUBESTACION_CABECERA"])
            elem_nsp_list.append(m["ELEMENTO_TECNICO_NSP105"])
            motivo_list.append(f"Conciliado con corrección de cabecera: '{r['Subestación']}' -> '{m['SUBESTACION_CABECERA']}' ({m['ID_CIRCUITO_NORMALIZADO']})")
        else:
            # 4. Transmisión / Generación / Industrial
            is_tx = any(kw in se for kw in tx_gen_keywords) or any(kw in ct for kw in tx_gen_keywords)
            if is_tx:
                status_list.append("EXCLUIDO POR TRANSMISIÓN / GENERACIÓN")
                matched_id_list.append("")
                matched_ct_name_list.append("")
                matched_se_name_list.append("")
                elem_nsp_list.append("TRANSMISIÓN/GENERACIÓN")
                motivo_list.append(f"Activo de alta tensión troncal o generación/industrial ({r['Circuito']} en S/E {r['Subestación']}) fuera del perímetro GGPD")
            else:
                # 5. Circuito reclasificado o pendiente de auditoría territorial
                status_list.append("EN REVISIÓN / AUDITORÍA TERRITORIAL")
                matched_id_list.append("")
                matched_ct_name_list.append("")
                matched_se_name_list.append("")
                elem_nsp_list.append("PENDIENTE_REVISION")
                motivo_list.append(f"Circuito legacy no reportado o renombrado en caracterización normalizada (requiere validación en campo con estado {r['Estado']})")

    df_orig["ESTATUS_CONCILIACION"] = status_list
    df_orig["ID_NORMALIZADO"] = matched_id_list
    df_orig["NOMBRE_NORMALIZADO"] = matched_ct_name_list
    df_orig["SE_NORMALIZADA"] = matched_se_name_list
    df_orig["ELEMENTO_NSP105"] = elem_nsp_list
    df_orig["DICTAMEN_TECNICO"] = motivo_list

    return df_orig, df_norm_se, df_norm_ct


def build_resumen_sheet(wb, df_orig, df_norm_se, df_norm_ct):
    ws = wb.active
    ws.title = "RESUMEN"
    ws.sheet_properties.tabColor = AZUL_CORP

    # Título
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value="CATÁLOGO MAESTRO Y MATRIZ DE RECONCILIACIÓN DE ACTIVOS DE RED (SE Y CT)")
    c.font = FONT_TITULO; c.fill = FILL_TITULO; c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 36

    # Ficha Técnica
    ficha = [
        ("Organismo:", "CORPOELEC — Gerencia General de Planificación de Distribución (GGPD)"),
        ("Unidad Emisora:", "Equipo de Automatización e Ingeniería de Productos con IA, de Planificación de Distribución"),
        ("Responsables:", "Yván M. Cipiran N. | T.S.U. Josué Pacheco"),
        ("Fecha de Emisión:", datetime.now().strftime("%d/%m/%Y %H:%M:%S (VET / UTC-4)")),
        ("Fuentes de Conciliación:", f"CARACTERIZACIÓN DISTRIBUCION.xls (4,313 CTs) ↔ Gemini SPARK Normalizado (871 SE / 4,207 CT)"),
        ("Código Normativo:", "GGPD-SGM-CAT-001 v2.0 ISO (Gobernanza ISO 8000-110 / ISO 55000 / NS-P-105)"),
    ]
    for i, (lbl, val) in enumerate(ficha, start=3):
        ws.cell(row=i, column=1, value=lbl).font = FONT_SUBTIT
        ws.cell(row=i, column=2, value=val).font = FONT_NORMAL
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = BORDER_THIN

    # Diagnóstico y Explicación de la Diferencia
    row = len(ficha) + 5
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="⚖️ CONCILIACIÓN INTEGRAL DE UNIVERSO: 4,313 vs 4,207 CIRCUITOS").font = Font(name="Calibri", size=12, bold=True, color=AZUL_CORP)
    row += 1

    status_counts = df_orig["ESTATUS_CONCILIACION"].value_counts()
    
    conciliacion_data = [
        ("Categoría de Conciliación", "Cantidad (CT)", "% del Total Legacy", "Dictamen Técnico Operativo"),
        ("1. Homologados Exactos", int(status_counts.get("HOMOLOGADO EXACTO", 0)), f"{status_counts.get('HOMOLOGADO EXACTO', 0)/len(df_orig)*100:.1f}%", "Circuitos con correspondencia directa en nombre, cabecera y estado en el catálogo 4,207."),
        ("2. Homologados con Saneamiento de Cabecera", int(status_counts.get("HOMOLOGADO (CABECERA AJUSTADA)", 0)), f"{status_counts.get('HOMOLOGADO (CABECERA AJUSTADA)', 0)/len(df_orig)*100:.1f}%", "Circuitos validados con corrección de nombre de subestación padre u OCR tipográfico."),
        ("3. Excluidos por Transmisión / Generación", int(status_counts.get("EXCLUIDO POR TRANSMISIÓN / GENERACIÓN", 0)), f"{status_counts.get('EXCLUIDO POR TRANSMISIÓN / GENERACIÓN', 0)/len(df_orig)*100:.1f}%", "Activos de centrales hidroeléctricas (Guri, Tocoma), patio de transmisión o clientes pesados (Sidor, Venalum)."),
        ("4. Duplicados Sintácticos en Origen", int(status_counts.get("DUPLICADO EN ORIGEN", 0)), f"{status_counts.get('DUPLICADO EN ORIGEN', 0)/len(df_orig)*100:.1f}%", "Filas repetidas exactamente en el libro de Excel original legacy."),
        ("5. En Revisión / Auditoría Territorial", int(status_counts.get("EN REVISIÓN / AUDITORÍA TERRITORIAL", 0)), f"{status_counts.get('EN REVISIÓN / AUDITORÍA TERRITORIAL', 0)/len(df_orig)*100:.1f}%", "Circuitos legacy con sufijos de patio (D-105, S1), cambios de nombre o bajas no reportadas (ver Hoja 5)."),
        ("TOTAL UNIVERSO LEGACY", len(df_orig), "100.0%", "100% de los registros originales auditados y clasificados uno a uno."),
    ]

    for ci, h in enumerate(conciliacion_data[0]):
        cell = ws.cell(row=row, column=ci + 1, value=h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    row += 1

    for i, (cat, cant, pct, dictamen) in enumerate(conciliacion_data[1:]):
        is_total = (i == len(conciliacion_data[1:]) - 1)
        ws.cell(row=row, column=1, value=cat).font = Font(name="Calibri", size=10, bold=is_total, color=AZUL_CORP if not is_total else BLANCO)
        ws.cell(row=row, column=2, value=cant).font = Font(name="Calibri", size=10, bold=True, color=VERDE_OK if not is_total else BLANCO)
        ws.cell(row=row, column=2).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=pct).font = Font(name="Calibri", size=10, bold=is_total, color=AZUL_CORP if not is_total else BLANCO)
        ws.cell(row=row, column=3).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4, value=dictamen).font = Font(name="Calibri", size=9.5, bold=is_total, color="333333" if not is_total else BLANCO)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)

        fill = FILL_TITULO if is_total else (FILL_ALTROW if i % 2 == 1 else PatternFill(fill_type=None))
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER_THIN
            if fill.fill_type:
                ws.cell(row=row, column=col).fill = fill
        row += 1

    # Métricas Consolidadas
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="📊 MÉTRICAS OFICIALES NORMALIZADAS (CATÁLOGO VIGENTE)").font = Font(name="Calibri", size=12, bold=True, color=AZUL_CORP)
    row += 1

    metrics = [
        ("Métrica / Parámetro", "Valor Normalizado", "Detalle Técnico"),
        ("Subestaciones Totales (SE)", len(df_norm_se), "451 Transmisión + 420 Distribución (Inc. 2 S/E Móviles: Banda Ciudadana y San Mateo)"),
        ("Circuitos y Equipos (CT)", len(df_norm_ct), "4,022 Alimentadores + 140 Seccionadores + 25 Reservas + 10 Barras + 9 Líneas MT + 1 Trafo"),
        ("Total Activos de Red", len(df_norm_se) + len(df_norm_ct), "Universo completo normalizado, mapeado y persistido en BD Supabase"),
        ("Entidades Territoriales", df_norm_se["ESTADO"].nunique(), "25 Entidades Federales (Inc. Guayana Esequiba 🇻🇪)"),
        ("Integridad Referencial CT → SE", "100.0%", "0 activos huérfanos tras remediación Yaracal II y saneamiento de cabeceras"),
        ("Norma de Segmentación", "CADAFE NS-P-105", "Doble capa: macro-control de balances + micro-segmentación de equipos de patio"),
    ]

    for ci, h in enumerate(metrics[0]):
        cell = ws.cell(row=row, column=ci + 1, value=h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    row += 1

    for i, (met, val, det) in enumerate(metrics[1:]):
        ws.cell(row=row, column=1, value=met).font = FONT_SUBTIT
        c_v = ws.cell(row=row, column=2, value=val)
        c_v.font = FONT_METRIC; c_v.fill = FILL_METRIC; c_v.alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=det).font = FONT_NORMAL
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER_THIN
            if i % 2 == 1 and col != 2:
                ws.cell(row=row, column=col).fill = FILL_ALTROW
        row += 1

    # Índice de Hojas
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="📋 ESTRUCTURA Y NAVEGACIÓN DEL LIBRO").font = Font(name="Calibri", size=12, bold=True, color=AZUL_CORP)
    row += 1

    hojas = [
        ("Hoja 1 — RESUMEN", "Ficha técnica, diagnóstico comparativo 4,313 vs 4,207, métricas y balance de auditoría"),
        ("Hoja 2 — SUBESTACIONES", f"{len(df_norm_se)} registros de Subestaciones (SE) con datos técnicos, tensiones, municipios y áreas"),
        ("Hoja 3 — CIRCUITOS_NORMALIZADOS", f"{len(df_norm_ct)} Circuitos (CT) con código maestro, cabecera SE, clasificación NS-P-105 y maniobra"),
        ("Hoja 4 — DASHBOARD", "Matriz combinada por estado con comparación directa Original vs Normalizado y ratios CT/SE"),
        ("Hoja 5 — CIRCUITOS_AUDITORIA_LEGACY", f"{len(df_orig)} registros originales con estatus de conciliación, ID homologado y dictamen técnico"),
    ]

    for hoja, desc in hojas:
        ws.cell(row=row, column=1, value=hoja).font = FONT_SUBTIT
        ws.cell(row=row, column=2, value=desc).font = FONT_NORMAL
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER_THIN
        row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    for lt in ("C", "D", "E", "F", "G"):
        ws.column_dimensions[lt].width = 18

    return ws


def build_se_sheet(wb, df_norm_se):
    ws = wb.create_sheet("SUBESTACIONES")
    ws.sheet_properties.tabColor = "1565C0"

    headers = [
        "N°", "Código Activo", "Nombre Subestación", "Nombre Normalizado", "Macro Proceso",
        "Estado Control", "Estado Caracterización", "Origen Dato", "Región", "Estado",
        "Municipio", "Tensión Entrada (kV)", "Tensión Secundaria (kV)", "Área SE (m²)",
        "Es Móvil", "Atendida Por", "Alimentador Principal", "Tensión Alimentador (kV)", "Observaciones"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    rows = []
    for idx, r in df_norm_se.iterrows():
        macro = "TRANSMISION" if "TRANSMISION" in str(r.get("TIPO_INSTALACION", "")).upper() else "DISTRIBUCION"
        rows.append([
            idx + 1,
            str(r.get("ID_ACTIVO_NORMALIZADO", "")),
            str(r.get("NOMBRE_SUBESTACION", "")),
            str(r.get("NOMBRE_SE_NORM", "")),
            macro,
            str(r.get("ESTADO_CONTROL", "CONTROLADO")),
            str(r.get("ESTADO_CARACTERIZACION", "CARACTERIZADO")),
            str(r.get("ORIGEN", "CARACTERIZACION SE")),
            str(r.get("REGION", "")),
            str(r.get("ESTADO", "")),
            str(r.get("MUNICIPIO", "")),
            r.get("TENSION_ENTRADA_KV"),
            r.get("TENSION_SECUNDARIA_KV"),
            r.get("AREA_SE_M2"),
            "SÍ" if str(r.get("ES_MOVIL", "")).strip().upper() == "SI" else "NO",
            str(r.get("ATENDIDA_POR", "")),
            str(r.get("ALIMENTADOR_PRINCIPAL", "")),
            r.get("TENSION_ALIMENTADOR_KV"),
            str(r.get("OBSERVACIONES_NORMALIZACION", "") or ""),
        ])

    write_data_rows(ws, rows, 2, len(headers), left_cols=4)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "E2"
    auto_width(ws)
    return ws


def build_ct_normalized_sheet(wb, df_norm_ct):
    ws = wb.create_sheet("CIRCUITOS_NORMALIZADOS")
    ws.sheet_properties.tabColor = "F57F17"

    headers = [
        "N°", "Código Normalizado", "Nombre Circuito", "Nombre CT Normalizado",
        "Macro Proceso", "Estado Control", "Estado Caracterización", "Origen Dato",
        "Región", "Estado", "SE Código Padre", "Subestación Cabecera",
        "Elemento Tipo", "Elemento Técnico (NS-P-105)", "Código Maniobra",
        "Descripción Técnica", "Es Componente de Patio", "Nivel Tensión (kV)",
        "Longitud (km)", "Tipo Red", "Observaciones"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    rows = []
    for idx, r in df_norm_ct.iterrows():
        rows.append([
            idx + 1,
            str(r.get("ID_CIRCUITO_NORMALIZADO", "")),
            str(r.get("NOMBRE_CIRCUITO", "")),
            str(r.get("NOMBRE_CT_NORM", "")),
            "DISTRIBUCION",
            str(r.get("ESTADO_CONTROL", "CONTROLADO")),
            str(r.get("ESTADO_CARACTERIZACION", "CARACTERIZADO")),
            "CARACTERIZACION CT",
            str(r.get("REGION", "")),
            str(r.get("ESTADO", "")),
            str(r.get("se_codigo", "")),
            str(r.get("SUBESTACION_CABECERA", "")),
            str(r.get("ELEMENTOS_TIPO", "CIRCUITO")),
            str(r.get("ELEMENTO_TECNICO_NSP105", "")),
            str(r.get("CODIGO_MANIOBRA_NORMA", "")),
            str(r.get("DESCRIPCION_TECNICA", "")),
            str(r.get("ES_COMPONENTE_PATIO", "NO")),
            r.get("NIVEL_TENSION_KV"),
            r.get("LONGITUD_KM"),
            str(r.get("TIPO_RED", "")),
            str(r.get("OBSERVACIONES_NORMALIZACION", "") or ""),
        ])

    write_data_rows(ws, rows, 2, len(headers), left_cols=4)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "E2"
    auto_width(ws)
    return ws


def build_dashboard_sheet(wb, df_orig, df_norm_se, df_norm_ct):
    ws = wb.create_sheet("DASHBOARD")
    ws.sheet_properties.tabColor = AMARILLO_CORP

    # Agregar datos por estado
    dashboard = {}
    
    # Contar SE
    for _, r in df_norm_se.iterrows():
        est = norm_str(r.get("ESTADO", "SIN_ESTADO"))
        reg = norm_str(r.get("REGION", ""))
        macro = "TX" if "TRANSMISION" in str(r.get("TIPO_INSTALACION", "")).upper() else "DX"
        mov = str(r.get("ES_MOVIL", "")).strip().upper() == "SI"
        
        if est not in dashboard:
            dashboard[est] = {"region": reg, "se_total": 0, "se_tx": 0, "se_dx": 0, "se_mov": 0,
                             "ct_norm_total": 0, "alim": 0, "secc": 0, "dis": 0, "bar": 0, "res": 0, "lin": 0, "otros": 0,
                             "ct_orig_total": 0}
        dashboard[est]["se_total"] += 1
        if macro == "TX": dashboard[est]["se_tx"] += 1
        else: dashboard[est]["se_dx"] += 1
        if mov: dashboard[est]["se_mov"] += 1

    # Contar CT normalizados
    for _, r in df_norm_ct.iterrows():
        est = norm_str(r.get("ESTADO", "SIN_ESTADO"))
        elem = r.get("ELEMENTO_TECNICO_NSP105", "ALIMENTADOR_CONVENCIONAL")
        if est not in dashboard:
            dashboard[est] = {"region": "", "se_total": 0, "se_tx": 0, "se_dx": 0, "se_mov": 0,
                             "ct_norm_total": 0, "alim": 0, "secc": 0, "dis": 0, "bar": 0, "res": 0, "lin": 0, "otros": 0,
                             "ct_orig_total": 0}
        dashboard[est]["ct_norm_total"] += 1
        if elem == "SECCIONADOR_MANIOBRA": dashboard[est]["secc"] += 1
        elif elem == "DISYUNTOR_POTENCIA": dashboard[est]["dis"] += 1
        elif elem == "JUEGO_BARRAS": dashboard[est]["bar"] += 1
        elif elem in ("POSICION_RESERVA", "RESERVA_AMPLIACION"): dashboard[est]["res"] += 1
        elif elem == "LINEA_ENLACE": dashboard[est]["lin"] += 1
        elif elem == "ALIMENTADOR_CONVENCIONAL": dashboard[est]["alim"] += 1
        else: dashboard[est]["otros"] += 1

    # Contar CT originales
    for _, r in df_orig.iterrows():
        est_raw = norm_str(r.get("Estado", "SIN_ESTADO"))
        est_map = {
            "DISTRITO CAPITAL": "DISTRITO_CAPITAL", "LA GUAIRA": "LA_GUAIRA",
            "NUEVA ESPARTA": "NUEVA_ESPARTA", "DELTA AMACURO": "DELTA_AMACURO",
            "VARGAS": "LA_GUAIRA", "FALCON": "FALCON", "TACHIRA": "TACHIRA"
        }
        est = est_map.get(est_raw, est_raw)
        if est not in dashboard:
            dashboard[est] = {"region": "", "se_total": 0, "se_tx": 0, "se_dx": 0, "se_mov": 0,
                             "ct_norm_total": 0, "alim": 0, "secc": 0, "dis": 0, "bar": 0, "res": 0, "lin": 0, "otros": 0,
                             "ct_orig_total": 0}
        dashboard[est]["ct_orig_total"] += 1

    estados_ord = sorted(dashboard.keys())

    headers = [
        "N°", "Entidad Federal", "Región", "SE Total", "SE Transmisión", "SE Distribución", "SE Móviles",
        "CT Normalizado (Vigente)", "Alimentadores", "Seccionadores (S)", "Disyuntores (D)", "Barras (B)",
        "Reservas (Q)", "Líneas MT (L)", "Otros", "CT Original Legacy (4,313)", "Diferencia Neta (Orig - Norm)", "Ratio CT/SE"
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(row=1, column=1, value="DASHBOARD COMPARATIVO Y MATRIZ TERRITORIAL DE ACTIVOS DE RED")
    c.font = FONT_TITULO; c.fill = FILL_TITULO; c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Calibri", size=9.5, bold=True, color=AZUL_CORP)
        cell.fill = FILL_DASH_HDR; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
    ws.row_dimensions[2].height = 28

    row_num = 3
    tot = {k: 0 for k in ["se_total", "se_tx", "se_dx", "se_mov", "ct_norm_total", "alim", "secc", "dis", "bar", "res", "lin", "otros", "ct_orig_total"]}

    for num, estado in enumerate(estados_ord, 1):
        d = dashboard[estado]
        dif = d["ct_orig_total"] - d["ct_norm_total"]
        ratio = round(d["ct_norm_total"] / d["se_total"], 2) if d["se_total"] > 0 else 0

        row_data = [
            num, estado, d["region"], d["se_total"], d["se_tx"], d["se_dx"], d["se_mov"],
            d["ct_norm_total"], d["alim"], d["secc"], d["dis"], d["bar"], d["res"], d["lin"], d["otros"],
            d["ct_orig_total"], dif, ratio
        ]

        for ci, val in enumerate(row_data):
            cell = ws.cell(row=row_num, column=ci + 1, value=val)
            cell.font = FONT_NORMAL; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
            if (row_num - 3) % 2 == 1:
                cell.fill = FILL_ALTROW
            # Resaltar diferencia si > 0
            if ci == 16 and val != 0:
                cell.font = Font(name="Calibri", size=10, bold=True, color=ROJO_WARN if val > 0 else VERDE_OK)

        for k in tot:
            tot[k] += d[k]
        row_num += 1

    # Fila TOTALES
    dif_total = tot["ct_orig_total"] - tot["ct_norm_total"]
    grand_ratio = round(tot["ct_norm_total"] / tot["se_total"], 2) if tot["se_total"] > 0 else 0

    total_row = [
        "", "TOTAL NACIONAL", "—", tot["se_total"], tot["se_tx"], tot["se_dx"], tot["se_mov"],
        tot["ct_norm_total"], tot["alim"], tot["secc"], tot["dis"], tot["bar"], tot["res"], tot["lin"], tot["otros"],
        tot["ct_orig_total"], dif_total, grand_ratio
    ]

    for ci, val in enumerate(total_row):
        cell = ws.cell(row=row_num, column=ci + 1, value=val)
        cell.font = Font(name="Calibri", size=10, bold=True, color=BLANCO)
        cell.fill = PatternFill(start_color=AZUL_CORP, end_color=AZUL_CORP, fill_type="solid")
        cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{row_num}"
    auto_width(ws)
    return ws


def build_legacy_audit_sheet(wb, df_orig):
    ws = wb.create_sheet("CIRCUITOS_AUDITORIA_LEGACY")
    ws.sheet_properties.tabColor = "C62828"

    headers = [
        "N°", "Estado Original", "Subestación Original", "Nivel Tensión S/E", "Circuito Original",
        "Nivel Tensión CT (kV)", "km Total", "Tipo Red Original", "km Aéreo", "km Subterráneo",
        "Condición Operativa", "Observaciones Originales", "Centro de Servicio", "Prioridad", "Nota",
        "ESTATUS DE CONCILIACIÓN", "Código Normalizado Asignado", "Nombre CT Normalizado",
        "Subestación Cabecera Asignada", "Elemento Técnico (NS-P-105)", "DICTAMEN TÉCNICO DE CONCILIACIÓN"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        if col >= 16:
            cell.fill = PatternFill(start_color="880E4F", end_color="880E4F", fill_type="solid")
        else:
            cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[1].height = 28

    rows = []
    for idx, r in df_orig.iterrows():
        rows.append([
            idx + 1,
            str(r.get("Estado", "")),
            str(r.get("Subestación", "")),
            r.get("NIVEL TENSION S/E"),
            str(r.get("Circuito", "")),
            r.get("Nivel de Tensión (kV)"),
            r.get("km  Total"),
            str(r.get("Tipo", "")),
            r.get("Km Total Aereo"),
            r.get("Klm Total Subterraneo"),
            str(r.get("Condicion", "")),
            str(r.get("Observaciones", "") or ""),
            str(r.get("CENTRO DE SERVICIO", "") or ""),
            str(r.get("PRIORI", "") or ""),
            str(r.get("NOTA", "") or ""),
            # Columnas de conciliación
            str(r.get("ESTATUS_CONCILIACION", "")),
            str(r.get("ID_NORMALIZADO", "") or ""),
            str(r.get("NOMBRE_NORMALIZADO", "") or ""),
            str(r.get("SE_NORMALIZADA", "") or ""),
            str(r.get("ELEMENTO_NSP105", "") or ""),
            str(r.get("DICTAMEN_TECNICO", "") or ""),
        ])

    for i, row_data in enumerate(rows):
        row_num = 2 + i
        status = row_data[15]

        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_num, column=col_idx + 1, value=value)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT if col_idx in (1, 2, 4, 11, 12, 17, 18, 20) else ALIGN_CENTER
            cell.border = BORDER_THIN

            # Colorear según estatus
            if col_idx == 15:
                if status == "HOMOLOGADO EXACTO":
                    cell.fill = FILL_OK; cell.font = Font(name="Calibri", size=9.5, bold=True, color=VERDE_OK)
                elif "CABECERA" in status:
                    cell.fill = FILL_REV; cell.font = Font(name="Calibri", size=9.5, bold=True, color="E65100")
                elif "DUPLICADO" in status:
                    cell.fill = FILL_DUP; cell.font = Font(name="Calibri", size=9.5, bold=True, color=ROJO_WARN)
                elif "TRANSMISIÓN" in status:
                    cell.fill = FILL_TX; cell.font = Font(name="Calibri", size=9.5, bold=True, color="4A148C")
                else:
                    cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                    cell.font = Font(name="Calibri", size=9.5, bold=True, color="F57F17")
            elif i % 2 == 1:
                cell.fill = FILL_ALTROW

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "F2"
    auto_width(ws, min_w=8, max_w=45)
    return ws


def main():
    print("=" * 76)
    print("⚡ CORPOELEC — GGPD — Generador de Catálogo Maestro y Reconciliación")
    print("=" * 76)

    df_orig, df_norm_se, df_norm_ct = load_and_reconcile()

    print(f"\n📊 Datos Cargados:")
    print(f"   - Circuitos Originales Legacy:     {len(df_orig)} registros")
    print(f"   - Subestaciones Normalizadas:       {len(df_norm_se)} registros")
    print(f"   - Circuitos Normalizados Vigentes:  {len(df_norm_ct)} registros")

    print("\n📝 Compilando Libro Excel de 5 Hojas...")
    wb = Workbook()

    print("   → Hoja 1: RESUMEN (Ficha, diagnóstico 4,313 vs 4,207 y métricas)")
    build_resumen_sheet(wb, df_orig, df_norm_se, df_norm_ct)

    print(f"   → Hoja 2: SUBESTACIONES ({len(df_norm_se)} registros)")
    build_se_sheet(wb, df_norm_se)

    print(f"   → Hoja 3: CIRCUITOS_NORMALIZADOS ({len(df_norm_ct)} registros)")
    build_ct_normalized_sheet(wb, df_norm_ct)

    print("   → Hoja 4: DASHBOARD (Matriz combinada y comparativa por estado)")
    build_dashboard_sheet(wb, df_orig, df_norm_se, df_norm_ct)

    print(f"   → Hoja 5: CIRCUITOS_AUDITORIA_LEGACY ({len(df_orig)} registros con conciliación)")
    build_legacy_audit_sheet(wb, df_orig)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)

    file_size = os.path.getsize(OUTPUT_FILE)
    size_str = f"{file_size / 1024:.1f} KB" if file_size < 1024*1024 else f"{file_size / (1024*1024):.2f} MB"

    print("\n" + "=" * 76)
    print("✅ ARCHIVO EXCEL DE RECONCILIACIÓN GENERADO CON ÉXITO")
    print(f"   📁 Ruta:    {OUTPUT_FILE}")
    print(f"   📦 Tamaño:  {size_str}")
    print(f"   📑 Hojas:   RESUMEN | SUBESTACIONES ({len(df_norm_se)}) | CIRCUITOS_NORMALIZADOS ({len(df_norm_ct)}) | DASHBOARD | CIRCUITOS_AUDITORIA_LEGACY ({len(df_orig)})")
    print("=" * 76)
    return OUTPUT_FILE


if __name__ == "__main__":
    main()

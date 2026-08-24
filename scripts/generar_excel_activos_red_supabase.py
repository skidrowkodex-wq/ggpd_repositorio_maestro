#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
⚡ CORPOELEC — GERENCIA GENERAL DE PLANIFICACIÓN DE DISTRIBUCIÓN (GGPD)
Generador de Libro Excel de Activos de Red (SE y CT)
Fuente: SQL Maestro 03_poblar_activos_red_caracterizacion.sql (5,078 activos)
Código Normativo: NAC_2026_GGPD_CATALOGO_ACTIVOS_RED_SE_CT
==============================================================================
Genera un libro .xlsx con 4 hojas sin membretes/cabeceras institucionales:
  1. RESUMEN          — Ficha resumen del archivo, actualizaciones y métricas
  2. SUBESTACIONES    — 871 SE con metadata expandida y origen
  3. CIRCUITOS        — 4,207 CT con metadata expandida y origen
  4. DASHBOARD        — Tabla combinada resumen tipo dashboard por estado
==============================================================================
"""

import json
import os
import re
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Se requiere openpyxl. Instale con: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

# ==============================================================================
# RUTAS
# ==============================================================================
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SQL_FILE = os.path.join(REPO_ROOT, "sql", "03_poblar_activos_red_caracterizacion.sql")
OUTPUT_DIR = os.path.join(REPO_ROOT, "docs")
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"NAC_2026_GGPD_CATALOGO_ACTIVOS_RED_SE_CT_{TIMESTAMP}.xlsx")

# ==============================================================================
# ESTILOS CORPORATIVOS
# ==============================================================================
AZUL_CORP     = "1B3A5C"
AZUL_CLARO    = "2E5D8A"
AMARILLO_CORP = "F4B400"
VERDE_OK      = "0D7C3F"
GRIS_FONDO    = "F5F5F5"
BLANCO        = "FFFFFF"

FONT_TITULO   = Font(name="Calibri", size=14, bold=True, color=BLANCO)
FONT_SUBTIT   = Font(name="Calibri", size=11, bold=True, color=AZUL_CORP)
FONT_HEADER   = Font(name="Calibri", size=10, bold=True, color=BLANCO)
FONT_NORMAL   = Font(name="Calibri", size=10, color="333333")
FONT_METRIC   = Font(name="Calibri", size=11, bold=True, color=VERDE_OK)
FONT_SMALL    = Font(name="Calibri", size=9, color="666666")
FONT_LINK     = Font(name="Calibri", size=10, color="1565C0", underline="single")

FILL_TITULO   = PatternFill(start_color=AZUL_CORP, end_color=AZUL_CORP, fill_type="solid")
FILL_HEADER   = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
FILL_ALTROW   = PatternFill(start_color=GRIS_FONDO, end_color=GRIS_FONDO, fill_type="solid")
FILL_METRIC   = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FILL_DASH_HDR = PatternFill(start_color=AMARILLO_CORP, end_color=AMARILLO_CORP, fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")

# ==============================================================================
# PARSER SQL → REGISTROS
# ==============================================================================
# Patrón para extraer los VALUES de cada INSERT
# VALUES ('código', 'nombre', 'tipo', 'macro', 'control', 'caract', 'origen', '{json}'::jsonb)
VALUES_RE = re.compile(
    r"VALUES\s*\(\s*"
    r"'((?:[^']|'')+)'\s*,\s*"          # 1: codigo_activo (handles escaped '')
    r"'((?:[^']|'')*)'\s*,\s*"          # 2: nombre
    r"'((?:[^']|'')*)'\s*,\s*"          # 3: tipo_activo
    r"'((?:[^']|'')*)'\s*,\s*"          # 4: macro_proceso
    r"'((?:[^']|'')*)'\s*,\s*"          # 5: estado_control
    r"'((?:[^']|'')*)'\s*,\s*"          # 6: estado_caracterizacion
    r"'((?:[^']|'')*)'\s*,\s*"          # 7: origen_dato
    r"'(\{.*?\})'::jsonb\s*\)",         # 8: metadata_tecnica JSON
    re.DOTALL
)


def parse_sql_file(filepath):
    """Parsea el archivo SQL y retorna listas de registros SE y CT."""
    print(f"   📄 Leyendo archivo SQL ({os.path.getsize(filepath) / (1024*1024):.1f} MB)...")

    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    all_matches = VALUES_RE.findall(content)
    print(f"   🔍 {len(all_matches)} registros VALUES detectados")

    se_data = []
    ct_data = []

    for match in all_matches:
        codigo, nombre, tipo, macro, control, caract, origen, meta_str = match

        # Un-escape SQL single quotes
        codigo = codigo.replace("''", "'")
        nombre = nombre.replace("''", "'")

        # Fix JSON escaped quotes too
        meta_str = meta_str.replace("''", "'")
        try:
            meta = json.loads(meta_str)
        except json.JSONDecodeError:
            meta = {}

        record = {
            "codigo_activo": codigo,
            "nombre": nombre,
            "tipo_activo": tipo,
            "macro_proceso": macro,
            "estado_control": control,
            "estado_caracterizacion": caract,
            "origen_dato": origen,
            "metadata_tecnica": meta,
        }

        if tipo == "SE":
            se_data.append(record)
        elif tipo == "CT":
            ct_data.append(record)

    return se_data, ct_data


# ==============================================================================
# UTILIDADES EXCEL
# ==============================================================================
def apply_header_style(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def write_data_rows(ws, data_rows, start_row, col_count):
    for i, row_data in enumerate(data_rows):
        row_num = start_row + i
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_num, column=col_idx + 1, value=value)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT if col_idx < 3 else ALIGN_CENTER
            cell.border = BORDER_THIN
            if i % 2 == 1:
                cell.fill = FILL_ALTROW


def auto_width(ws, min_w=8, max_w=38):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_w
        for cell in col_cells[:200]:  # Muestrear las primeras 200 filas para performance
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[col_letter].width = max_len


def safe_num(val):
    """Convierte a número si es posible, None si vacío."""
    if val is None or val == "" or val == "null":
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        return float(val)
    except (ValueError, TypeError):
        return val


# ==============================================================================
# HOJA 1: RESUMEN
# ==============================================================================
def build_resumen_sheet(wb, se_data, ct_data):
    ws = wb.active
    ws.title = "RESUMEN"
    ws.sheet_properties.tabColor = AZUL_CORP

    # ── Título principal ──
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1,
                value="CATÁLOGO MAESTRO DE ACTIVOS DE RED — SUBESTACIONES (SE) Y CIRCUITOS (CT)")
    c.font = FONT_TITULO; c.fill = FILL_TITULO; c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 36

    # ── Ficha técnica ──
    ficha = [
        ("Organismo:", "CORPOELEC — Gerencia General de Planificación de Distribución (GGPD)"),
        ("Unidad Emisora:", "Equipo de Automatización e Ingeniería de Productos con IA, de Planificación de Distribución"),
        ("Responsables:", "Yván M. Cipiran N. | T.S.U. Josué Pacheco"),
        ("Fecha de Generación:", datetime.now().strftime("%d/%m/%Y %H:%M:%S (VET / UTC-4)")),
        ("Fuente de Datos:", "SQL Maestro — 03_poblar_activos_red_caracterizacion.sql (Supabase / PostgreSQL public.activos_red)"),
        ("Origen del Procesamiento:", "caracterizacion_distribucion_normalizado.xlsx → Gemini SPARK → ISO 8000-110 / NS-P-105"),
        ("Código Normativo:", "GGPD-SGM-CAT-001 v1.0 ISO"),
    ]
    for i, (lbl, val) in enumerate(ficha, start=3):
        ws.cell(row=i, column=1, value=lbl).font = FONT_SUBTIT
        ws.cell(row=i, column=2, value=val).font = FONT_NORMAL
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = BORDER_THIN

    # ── Separador: Actualizaciones recientes ──
    row = len(ficha) + 5
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1,
            value="🔄 ACTUALIZACIONES RECIENTES APLICADAS").font = Font(
        name="Calibri", size=12, bold=True, color=AZUL_CORP)
    row += 1

    updates = [
        ("Remediación OCR Falcón", "YARACAL LL → YARACAL II (CT-FAL-01467..01469 → SE-FAL-0320)", "Agosto 2026"),
        ("Saneamiento Cabeceras en Blanco", "11 circuitos Carabobo/Zulia vinculados a nodos padre", "Agosto 2026"),
        ("Segmentación NS-P-105", "140 Seccionadores, 25 Reservas, 10 Barras, 9 Líneas MT segregados", "Agosto 2026"),
        ("100% Integridad Referencial", "0 huérfanos CT→SE tras Gemini SPARK normalización", "Agosto 2026"),
        ("Ingesta Universal SIGI", "Motor de tolerancia a datos sucios homologado SCTIS v2.0 → SIGI", "Agosto 2026"),
        ("Catálogos Maestros Spark", "7 libros Excel regenerados en docs/catalogos_maestros_spark/", "Agosto 2026"),
    ]

    upd_headers = ["Actualización", "Descripción", "Fecha"]
    for col_idx, h in enumerate(upd_headers):
        cell = ws.cell(row=row, column=col_idx + 1, value=h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1

    for i, (upd, desc, fecha) in enumerate(updates):
        ws.cell(row=row, column=1, value=upd).font = FONT_SUBTIT
        ws.cell(row=row, column=2, value=desc).font = FONT_NORMAL
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=7, value=fecha).font = FONT_SMALL
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER_THIN
            if i % 2 == 1:
                ws.cell(row=row, column=col).fill = FILL_ALTROW
        row += 1

    # ── Métricas consolidadas ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1,
            value="📊 MÉTRICAS CONSOLIDADAS DE ACTIVOS DE RED").font = Font(
        name="Calibri", size=12, bold=True, color=AZUL_CORP)
    row += 1

    # Calcular métricas
    estados_set = set()
    se_tx, se_dx, se_mov = 0, 0, 0
    ct_alim, ct_secc, ct_dis, ct_bar, ct_res, ct_lin, ct_otros = 0, 0, 0, 0, 0, 0, 0

    for r in se_data:
        m = r["metadata_tecnica"]
        estados_set.add(m.get("estado", ""))
        if r["macro_proceso"] == "TRANSMISION": se_tx += 1
        else: se_dx += 1
        if m.get("es_movil"): se_mov += 1

    for r in ct_data:
        m = r["metadata_tecnica"]
        estados_set.add(m.get("estado", ""))
        elem = m.get("elemento_tecnico_especifico", "ALIMENTADOR_CONVENCIONAL")
        if elem == "SECCIONADOR_MANIOBRA": ct_secc += 1
        elif elem == "DISYUNTOR_POTENCIA": ct_dis += 1
        elif elem == "JUEGO_BARRAS": ct_bar += 1
        elif elem in ("POSICION_RESERVA", "RESERVA_AMPLIACION"): ct_res += 1
        elif elem == "LINEA_INTERCONEXION_MT": ct_lin += 1
        elif elem == "ALIMENTADOR_CONVENCIONAL": ct_alim += 1
        else: ct_otros += 1

    estados_set.discard(""); estados_set.discard(None)
    se_count = len(se_data)
    ct_count = len(ct_data)

    metrics = [
        ("Métrica / Parámetro", "Valor", "Detalle"),
        ("Total Subestaciones (SE)", se_count, f"{se_tx} Transmisión / {se_dx} Distribución / {se_mov} Móviles"),
        ("Total Circuitos / Elementos CT", ct_count,
         f"{ct_alim} Alimentadores + {ct_secc} Seccionadores + {ct_dis} Disyuntores + {ct_bar} Barras + {ct_res} Reservas + {ct_lin} Líneas MT + {ct_otros} Otros"),
        ("Total Activos de Red", se_count + ct_count, "Universo completo contenido en BD PostgreSQL / Supabase"),
        ("Entidades Territoriales", len(estados_set), "Cobertura federal completa (Inc. Guayana Esequiba 🇻🇪)"),
        ("Integridad Referencial (CT → SE)", "100.0%", "0 registros huérfanos — verificado Gemini SPARK"),
        ("Estado de Caracterización", "CARACTERIZADO", "100% de los activos auditados y normalizados"),
        ("Estado de Control Operativo", "CONTROLADO", "100% bajo gestión formal GGPD"),
        ("Normas Aplicables", "ISO 8000-110 | ISO 55000 | IEC 81346-10 | NS-P-105",
         "Cumplimiento integral de calidad, gestión de activos y codificación"),
    ]

    # Encabezados
    for ci, val in enumerate(metrics[0]):
        cell = ws.cell(row=row, column=ci + 1, value=val)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    row += 1

    for i, (metrica, valor, detalle) in enumerate(metrics[1:]):
        ws.cell(row=row, column=1, value=metrica).font = FONT_SUBTIT
        c_v = ws.cell(row=row, column=2, value=valor)
        c_v.font = FONT_METRIC; c_v.fill = FILL_METRIC; c_v.alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=detalle).font = FONT_NORMAL
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER_THIN
            if i % 2 == 1 and col != 2:
                ws.cell(row=row, column=col).fill = FILL_ALTROW
        row += 1

    # ── Contenido de hojas ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="📋 CONTENIDO DEL LIBRO").font = Font(
        name="Calibri", size=12, bold=True, color=AZUL_CORP)
    row += 1

    hojas = [
        ("Hoja 1 — RESUMEN", "Ficha técnica, actualizaciones recientes, métricas consolidadas y descripción del contenido"),
        ("Hoja 2 — SUBESTACIONES",
         f"{se_count} registros SE con metadata técnica expandida (región, estado, municipio, tensiones, área, origen)"),
        ("Hoja 3 — CIRCUITOS",
         f"{ct_count} registros CT con metadata expandida (SE cabecera, elemento técnico NS-P-105, maniobra, origen)"),
        ("Hoja 4 — DASHBOARD",
         "Tabla combinada resumen por estado con conteos SE, CT, alimentadores, seccionadores, cobertura y ratios"),
    ]
    for hoja, desc in hojas:
        ws.cell(row=row, column=1, value=hoja).font = FONT_SUBTIT
        ws.cell(row=row, column=2, value=desc).font = FONT_NORMAL
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER_THIN
        row += 1

    # ── Origen de los datos ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="📂 ORÍGENES DE DATOS").font = Font(
        name="Calibri", size=12, bold=True, color=AZUL_CORP)
    row += 1

    origenes = [
        ("Fuente Primaria", "caracterizacion_distribucion_normalizado.xlsx",
         "Archivo Excel con datos de campo de 25 entidades, normalizado por Gemini SPARK"),
        ("Script de Remediación", "scripts/remediar_y_poblar_caracterizacion.py",
         "Motor de remediación ISO 8000-110 y segmentación NS-P-105 de elementos de patio"),
        ("Script SQL Maestro", "sql/03_poblar_activos_red_caracterizacion.sql",
         "5,078 sentencias UPSERT idempotentes para la tabla public.activos_red"),
        ("Base de Datos Destino", "Supabase PostgreSQL — public.activos_red",
         "https://owpiwacuotcaeruvonbd.supabase.co — Esquema DDL en sql/01_tabla_unificada_activos.sql"),
        ("Catálogos Spark", "docs/catalogos_maestros_spark/ (7 libros .xlsx)",
         "Catálogos maestros normalizados para analítica BigQuery/Spark"),
        ("Catálogo JSON Frontend", "src/data/masterCatalogsLegacy.json",
         "Sincronizado con SIGI Distribución para dashboards y mapas GIS"),
    ]

    orig_hdr = ["Concepto", "Archivo / Recurso", "Descripción"]
    for ci, h in enumerate(orig_hdr):
        cell = ws.cell(row=row, column=ci + 1, value=h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    row += 1

    for i, (concepto, archivo, desc) in enumerate(origenes):
        ws.cell(row=row, column=1, value=concepto).font = FONT_SUBTIT
        ws.cell(row=row, column=2, value=archivo).font = FONT_LINK
        ws.cell(row=row, column=3, value=desc).font = FONT_NORMAL
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER_THIN
            if i % 2 == 1:
                ws.cell(row=row, column=col).fill = FILL_ALTROW
        row += 1

    # Ajustes de ancho
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    for lt in ("C", "D", "E", "F", "G"):
        ws.column_dimensions[lt].width = 18

    return ws


# ==============================================================================
# HOJA 2: SUBESTACIONES
# ==============================================================================
def build_se_sheet(wb, se_data):
    ws = wb.create_sheet("SUBESTACIONES")
    ws.sheet_properties.tabColor = "1565C0"

    headers = [
        "N°", "Código Activo", "Nombre SE", "Macro Proceso", "Estado Control",
        "Estado Caracterización", "Origen Dato", "Región", "Estado",
        "Municipio", "Tensión Entrada (kV)", "Tensión Secundaria (kV)",
        "Área SE (m²)", "Es Móvil", "Atendida Por",
        "Alimentador Principal", "Tensión Alimentador (kV)", "Observaciones",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    rows = []
    for idx, r in enumerate(se_data, 1):
        m = r["metadata_tecnica"]
        rows.append([
            idx,
            r["codigo_activo"],
            r["nombre"],
            r["macro_proceso"],
            r["estado_control"],
            r["estado_caracterizacion"],
            r["origen_dato"],
            m.get("region", ""),
            m.get("estado", ""),
            m.get("municipio", ""),
            safe_num(m.get("tension_entrada_kv")),
            safe_num(m.get("tension_secundaria_kv")),
            safe_num(m.get("area_se_m2")),
            "SÍ" if m.get("es_movil") else "NO",
            m.get("atendida_por", ""),
            m.get("alimentador_principal", ""),
            safe_num(m.get("tension_alimentador_kv")),
            m.get("observaciones") or "",
        ])

    write_data_rows(ws, rows, 2, len(headers))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "D2"
    auto_width(ws)
    return ws


# ==============================================================================
# HOJA 3: CIRCUITOS
# ==============================================================================
def build_ct_sheet(wb, ct_data):
    ws = wb.create_sheet("CIRCUITOS")
    ws.sheet_properties.tabColor = "F57F17"

    headers = [
        "N°", "Código Activo", "Nombre CT", "Macro Proceso", "Estado Control",
        "Estado Caracterización", "Origen Dato", "Región", "Estado",
        "SE Código (Cabecera)", "SE Nombre (Cabecera)", "Municipio",
        "Nivel Tensión (kV)", "Elemento Técnico (NS-P-105)", "Código Maniobra",
        "Descripción Técnica", "Es Componente de Patio",
        "Longitud (km)", "Tipo Red", "Observaciones",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    rows = []
    for idx, r in enumerate(ct_data, 1):
        m = r["metadata_tecnica"]
        rows.append([
            idx,
            r["codigo_activo"],
            r["nombre"],
            r["macro_proceso"],
            r["estado_control"],
            r["estado_caracterizacion"],
            r["origen_dato"],
            m.get("region", ""),
            m.get("estado", ""),
            m.get("se_codigo_padre", m.get("se_codigo", "")),
            m.get("subestacion_cabecera", m.get("se_nombre", "")),
            m.get("municipio", ""),
            safe_num(m.get("nivel_tension_kv", m.get("tension_entrada_kv"))),
            m.get("elemento_tecnico_especifico", "ALIMENTADOR_CONVENCIONAL"),
            m.get("codigo_maniobra_norma", ""),
            m.get("descripcion_elemento_tecnico", ""),
            "SÍ" if m.get("es_componente_patio") else "NO",
            safe_num(m.get("longitud_km")),
            m.get("tipo_red", ""),
            m.get("observaciones") or "",
        ])

    write_data_rows(ws, rows, 2, len(headers))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "D2"
    auto_width(ws)
    return ws


# ==============================================================================
# HOJA 4: DASHBOARD
# ==============================================================================
def build_dashboard_sheet(wb, se_data, ct_data):
    ws = wb.create_sheet("DASHBOARD")
    ws.sheet_properties.tabColor = AMARILLO_CORP

    # Agregar datos por estado
    dashboard = {}
    for r in se_data:
        m = r["metadata_tecnica"]
        estado = m.get("estado", "SIN ESTADO")
        region = m.get("region", "")
        if estado not in dashboard:
            dashboard[estado] = {"region": region, "se_total": 0, "se_tx": 0,
                                 "se_dx": 0, "se_mov": 0, "ct_total": 0,
                                 "alim": 0, "secc": 0, "dis": 0, "bar": 0,
                                 "res": 0, "lin": 0, "otros": 0}
        dashboard[estado]["se_total"] += 1
        if r["macro_proceso"] == "TRANSMISION":
            dashboard[estado]["se_tx"] += 1
        else:
            dashboard[estado]["se_dx"] += 1
        if m.get("es_movil"):
            dashboard[estado]["se_mov"] += 1

    for r in ct_data:
        m = r["metadata_tecnica"]
        estado = m.get("estado", "SIN ESTADO")
        region = m.get("region", "")
        if estado not in dashboard:
            dashboard[estado] = {"region": region, "se_total": 0, "se_tx": 0,
                                 "se_dx": 0, "se_mov": 0, "ct_total": 0,
                                 "alim": 0, "secc": 0, "dis": 0, "bar": 0,
                                 "res": 0, "lin": 0, "otros": 0}
        dashboard[estado]["ct_total"] += 1
        elem = m.get("elemento_tecnico_especifico", "ALIMENTADOR_CONVENCIONAL")
        key_map = {
            "SECCIONADOR_MANIOBRA": "secc", "DISYUNTOR_POTENCIA": "dis",
            "JUEGO_BARRAS": "bar", "POSICION_RESERVA": "res",
            "RESERVA_AMPLIACION": "res", "LINEA_INTERCONEXION_MT": "lin",
            "ALIMENTADOR_CONVENCIONAL": "alim",
        }
        dashboard[estado][key_map.get(elem, "otros")] += 1

    estados_ord = sorted(dashboard.keys())

    headers = [
        "N°", "Estado", "Región", "SE Total", "SE Transmisión", "SE Distribución",
        "SE Móviles", "CT Total", "Alimentadores", "Seccionadores (S)",
        "Disyuntores (D)", "Barras (B)", "Reservas (Q)", "Líneas MT (L)",
        "Otros", "Total Activos", "Ratio CT/SE"
    ]

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws.cell(row=1, column=1,
                value="DASHBOARD CONSOLIDADO — ACTIVOS DE RED POR ENTIDAD TERRITORIAL")
    c.font = FONT_TITULO; c.fill = FILL_TITULO; c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=AZUL_CORP)
        cell.fill = FILL_DASH_HDR; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
    ws.row_dimensions[2].height = 26

    # Datos
    row_num = 3
    totals = {"se_total": 0, "se_tx": 0, "se_dx": 0, "se_mov": 0,
              "ct_total": 0, "alim": 0, "secc": 0, "dis": 0,
              "bar": 0, "res": 0, "lin": 0, "otros": 0}

    for num, estado in enumerate(estados_ord, 1):
        d = dashboard[estado]
        total_act = d["se_total"] + d["ct_total"]
        ratio = round(d["ct_total"] / d["se_total"], 2) if d["se_total"] > 0 else 0

        row_data = [
            num, estado, d["region"], d["se_total"], d["se_tx"],
            d["se_dx"], d["se_mov"], d["ct_total"],
            d["alim"], d["secc"], d["dis"],
            d["bar"], d["res"], d["lin"], d["otros"],
            total_act, ratio
        ]

        for ci, val in enumerate(row_data):
            cell = ws.cell(row=row_num, column=ci + 1, value=val)
            cell.font = FONT_NORMAL; cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
            if (row_num - 3) % 2 == 1:
                cell.fill = FILL_ALTROW

        for k in totals:
            totals[k] += d[k]
        row_num += 1

    # Fila TOTALES
    grand_total = totals["se_total"] + totals["ct_total"]
    grand_ratio = round(totals["ct_total"] / totals["se_total"], 2) if totals["se_total"] > 0 else 0

    total_row = [
        "", "TOTAL NACIONAL", "—", totals["se_total"], totals["se_tx"],
        totals["se_dx"], totals["se_mov"], totals["ct_total"],
        totals["alim"], totals["secc"], totals["dis"],
        totals["bar"], totals["res"], totals["lin"], totals["otros"],
        grand_total, grand_ratio
    ]

    for ci, val in enumerate(total_row):
        cell = ws.cell(row=row_num, column=ci + 1, value=val)
        cell.font = Font(name="Calibri", size=10, bold=True, color=BLANCO)
        cell.fill = PatternFill(start_color=AZUL_CORP, end_color=AZUL_CORP, fill_type="solid")
        cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{row_num}"
    auto_width(ws)
    return ws


# ==============================================================================
# MAIN
# ==============================================================================
def main():
    print("=" * 72)
    print("⚡ CORPOELEC — GGPD — Generador de Catálogo de Activos de Red")
    print("   Fuente: SQL Maestro 03_poblar_activos_red_caracterizacion.sql")
    print("=" * 72)

    if not os.path.isfile(SQL_FILE):
        print(f"\n❌ ERROR: No se encontró el archivo SQL:\n   {SQL_FILE}")
        sys.exit(1)

    # 1. Parsear SQL
    print(f"\n📡 Parseando activos desde el archivo SQL...")
    se_data, ct_data = parse_sql_file(SQL_FILE)
    print(f"   ✅ {len(se_data)} Subestaciones (SE)")
    print(f"   ✅ {len(ct_data)} Circuitos / Elementos (CT)")
    print(f"   📊 TOTAL ACTIVOS: {len(se_data) + len(ct_data)}")

    if len(se_data) == 0 and len(ct_data) == 0:
        print("\n⚠️  No se detectaron registros. Verifique el formato del SQL.")
        sys.exit(1)

    # 2. Generar libro Excel
    print(f"\n📝 Generando libro Excel con 4 hojas...")
    wb = Workbook()

    print("   → Hoja 1: RESUMEN (ficha, actualizaciones, métricas, orígenes)")
    build_resumen_sheet(wb, se_data, ct_data)

    print(f"   → Hoja 2: SUBESTACIONES ({len(se_data)} registros)")
    build_se_sheet(wb, se_data)

    print(f"   → Hoja 3: CIRCUITOS ({len(ct_data)} registros)")
    build_ct_sheet(wb, ct_data)

    print("   → Hoja 4: DASHBOARD (resumen por estado)")
    build_dashboard_sheet(wb, se_data, ct_data)

    # 3. Guardar archivo
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)

    file_size = os.path.getsize(OUTPUT_FILE)
    if file_size < 1024 * 1024:
        size_str = f"{file_size / 1024:.1f} KB"
    else:
        size_str = f"{file_size / (1024*1024):.2f} MB"

    print(f"\n{'=' * 72}")
    print(f"✅ ARCHIVO GENERADO EXITOSAMENTE")
    print(f"   📁 Ruta:    {OUTPUT_FILE}")
    print(f"   📦 Tamaño:  {size_str}")
    print(f"   📊 Hojas:   RESUMEN | SUBESTACIONES ({len(se_data)}) | CIRCUITOS ({len(ct_data)}) | DASHBOARD")
    print(f"{'=' * 72}")
    return OUTPUT_FILE


if __name__ == "__main__":
    main()

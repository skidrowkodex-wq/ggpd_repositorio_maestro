import os
import glob
import pandas as pd
import openpyxl

target_dir = '/home/skidrowkodex/Documentos/Repositorio_Maestro/apps-refactorizadas/SCPPE-REF/prtsen_drive'

def scan_drive_folder():
    print(f"============================================================")
    print(f"🔍 INSPECCIÓN DE ARCHIVOS EN: {target_dir}")
    print(f"============================================================")
    
    if not os.path.exists(target_dir):
        print(f"❌ El directorio no existe todavía.")
        return

    all_files = []
    for root, dirs, files in os.walk(target_dir):
        for f in files:
            full_path = os.path.join(root, f)
            rel_path = os.path.relpath(full_path, target_dir)
            size_kb = round(os.path.getsize(full_path) / 1024, 2)
            all_files.append((rel_path, full_path, size_kb))

    print(f"Total de archivos detectados: {len(all_files)}")
    if len(all_files) == 0:
        print("⏳ Carpeta vacía o en proceso de descarga...")
        return

    for rel, full, size in sorted(all_files):
        ext = os.path.splitext(full)[1].lower()
        print(f"\n📄 [{ext.upper()}] {rel} ({size} KB)")
        
        if ext in ['.xlsx', '.xls']:
            try:
                wb = openpyxl.load_workbook(full, read_only=True, data_only=True)
                print(f"   Hojas: {wb.sheetnames}")
                for sname in wb.sheetnames:
                    sheet = wb[sname]
                    rows = list(sheet.iter_rows(values_only=True))
                    print(f"   - Hoja '{sname}': {len(rows)} filas")
                    if len(rows) > 0:
                        first_row = [str(c) for c in rows[0] if c is not None]
                        print(f"     Cabeceras ({len(first_row)}): {first_row[:8]}")
            except Exception as e:
                print(f"   ⚠️ Error leyendo Excel: {e}")

if __name__ == '__main__':
    scan_drive_folder()

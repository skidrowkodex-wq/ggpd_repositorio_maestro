#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
⚡ CORPOELEC - GGPD / INSFORGE-BK MASTER CATALOGS DEPLOYER
Crea y puebla todas las tablas de catálogos maestros compartidos:
- cat_macro_procesos (8 Macro-procesos / 10 Procesos)
- cat_tipos_equipo_potencia (SCEIN 12 familias de equipos y componentes)
- cat_causas_interrupcion (SCTIS 22 Causas Oficiales SEN)
- cat_subcausas_interrupcion (SCTIS Sub-causas homologadas)
- cat_familias_materiales (21 Familias de Materiales e Insumos)
- cat_items_materiales_precios (804 Insumos con Precios en EUR y Unidades)
- cat_tipos_restriccion_operativa (Restricciones de Circuitos MT)
- cat_estados_operativos_activo (Estados de activos ISO 55000)
==============================================================================
"""

import json
import openpyxl
import subprocess
import os

def escape_sql(val):
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        return str(val)
    clean = str(val).replace("'", "''").strip()
    return f"'{clean}'"

def run_query(sql):
    res = subprocess.run(['npx', '@insforge/cli', 'db', 'query', sql], capture_output=True, text=True)
    if res.returncode != 0:
        print("SQL Error:", res.stderr, res.stdout)
        raise RuntimeError(f"Query failed with code {res.returncode}")
    return res.stdout

def main():
    print("1. Creating DDL for all master catalog tables in core...")
    ddl_sql = """
    -- 1. Macro-procesos de la GGPD
    CREATE TABLE IF NOT EXISTS core.cat_macro_procesos (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        codigo_proceso VARCHAR(30) NOT NULL UNIQUE,
        nombre_proceso VARCHAR(150) NOT NULL,
        macro_proceso_padre VARCHAR(100),
        sistema_responsable VARCHAR(100),
        frecuencia_corte VARCHAR(50) DEFAULT 'SEMANAL',
        descripcion TEXT,
        activo BOOLEAN NOT NULL DEFAULT true,
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );

    -- 2. Familias y Tipos de Equipos de Potencia (SCEIN / ISO 55000)
    CREATE TABLE IF NOT EXISTS core.cat_tipos_equipo_potencia (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        id_tipo_equipo VARCHAR(30) NOT NULL UNIQUE,
        categoria_equipo VARCHAR(50) NOT NULL,
        nombre_elemento_estandar VARCHAR(150) NOT NULL,
        siglas_componente VARCHAR(50),
        niveles_kv_tipicos VARCHAR(100),
        acciones_operativas_habituales TEXT,
        criticidad_defecto VARCHAR(20) DEFAULT 'ALTA',
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );

    -- 3. Causas Oficiales de Interrupción SEN (SCTIS)
    CREATE TABLE IF NOT EXISTS core.cat_causas_interrupcion (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        codigo_causa VARCHAR(50) NOT NULL UNIQUE,
        nombre_oficial_causa_sen VARCHAR(150) NOT NULL,
        descripcion_operativa TEXT,
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );

    -- 4. Sub-Causas de Interrupción (SCTIS)
    CREATE TABLE IF NOT EXISTS core.cat_subcausas_interrupcion (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        codigo_sub_causa VARCHAR(50) NOT NULL UNIQUE,
        codigo_causa_padre VARCHAR(50) NOT NULL REFERENCES core.cat_causas_interrupcion(codigo_causa) ON UPDATE CASCADE,
        nombre_sub_causa_oficial VARCHAR(150) NOT NULL,
        descripcion TEXT,
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );

    -- 5. Familias de Materiales e Insumos
    CREATE TABLE IF NOT EXISTS core.cat_familias_materiales (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        codigo_familia VARCHAR(50) NOT NULL UNIQUE,
        nombre_familia VARCHAR(150) NOT NULL,
        descripcion TEXT,
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );

    -- 6. Items de Materiales, Unidades y Precios de Referencia (Precios EUR)
    CREATE TABLE IF NOT EXISTS core.cat_items_materiales_precios (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        descripcion_item VARCHAR(255) NOT NULL,
        codigo_familia VARCHAR(50),
        unidad_medida VARCHAR(20) NOT NULL DEFAULT 'UN',
        precio_referencia_eur NUMERIC(12,4) NOT NULL DEFAULT 0.0,
        moneda VARCHAR(10) NOT NULL DEFAULT 'EUR',
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );

    -- 7. Tipos de Restricción Operativa (Circuitos MT)
    CREATE TABLE IF NOT EXISTS core.cat_tipos_restriccion_operativa (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        codigo_restriccion VARCHAR(50) NOT NULL UNIQUE,
        nombre_restriccion VARCHAR(150) NOT NULL,
        descripcion TEXT,
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );

    -- 8. Estados Operativos de Activos (ISO 55000)
    CREATE TABLE IF NOT EXISTS core.cat_estados_operativos_activo (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        codigo_estado_operativo VARCHAR(30) NOT NULL UNIQUE,
        nombre_estado VARCHAR(100) NOT NULL,
        descripcion TEXT,
        creado_en TIMESTAMPTZ NOT NULL DEFAULT clock_timestamp()
    );
    """
    run_query(ddl_sql)
    print("DDL created successfully!")

    # 1. Populate Macro-Procesos
    print("2. Populating cat_macro_procesos...")
    procesos_data = [
        ('01_SCTIS', 'Seguimiento y Control de Tiras de Interrupciones', 'MP-06 Calidad del Servicio e Interrupciones', 'SCTIS V2.0', 'SEMANAL', 'Ingesta y balance de interrupciones y energía no suministrada (ENS)'),
        ('02_SCEIN', 'Seguimiento y Control de Equipos Indisponibles', 'MP-01 Mantenimiento de Subestaciones', 'SCEIN V3.0', 'SEMANAL', 'Inventario, diagnóstico y plan de atención de equipos averiados en Subestaciones'),
        ('03_SCPPE', 'Seguimiento y Control de Planes y Proyectos Especiales / Viáticos', 'MP-08 Planificación Estratégica y Finanzas', 'SCPPE V3.0', 'SEMANAL', 'Control presupuestario de viáticos SAMC y seguimiento de proyectos'),
        ('04_SCMTP', 'Seguimiento y Control de Minutas y Tareas de Planificación', 'MP-08 Planificación Estratégica', 'SCMTP V2.0', 'SEMANAL', 'Digitalización de minutas, compromisos ministeriales y acuerdos institucionales'),
        ('05_SCPYP', 'Control de Mantenimiento: Pica y Poda de Vegetación', 'MP-02 Mantenimiento de Redes de Media Tensión', 'SIGI Ingesta Hub', 'SEMANAL', 'Seguimiento de kilómetros despejados y mantenimiento de servidumbre en circuitos'),
        ('06_SCDES', 'Control de Mantenimiento: Desmalezamiento de Patios en SE', 'MP-01 Mantenimiento de Subestaciones', 'SIGI Ingesta Hub', 'SEMANAL', 'Limpieza, desmalezamiento y control químico en patios de subestaciones'),
        ('07_SCALU', 'Mantenimiento e Instalación de Alumbrado Público', 'MP-04 Instalaciones de Alumbrado Público', 'SIGI Ingesta Hub', 'MENSUAL', 'Instalación y sustitución de luminarias, atención casos 1x10 VenApp'),
        ('08_SCRES', 'Levantamiento y Plan de Atención de Restricciones Operativas', 'MP-02 Mantenimiento de Redes de Media Tensión', 'SIGI Ingesta Hub', 'SEMANAL', 'Gestión de cuellos de botella y limitaciones operativas en alimentadores'),
        ('09_SCBTE', 'Mantenimiento y Adecuación de Redes de Baja Tensión', 'MP-03 Mantenimiento de Redes de Baja Tensión', 'SIGI Ingesta Hub', 'MENSUAL', 'Balanceo de cargas BT, sustitución de acometidas y conductores secundarios'),
        ('10_SCTRA', 'Gestión y Tasa de Falla de Transformadores de Distribución', 'MP-05 Gestión de Transformadores', 'SIGI Ingesta Hub / MDM', 'SEMANAL', 'Averías, desincorporación y reemplazo de transformadores de distribución')
    ]
    p_vals = []
    for cod, nom, mp, sis, frec, desc in procesos_data:
        p_vals.append(f"({escape_sql(cod)}, {escape_sql(nom)}, {escape_sql(mp)}, {escape_sql(sis)}, {escape_sql(frec)}, {escape_sql(desc)})")
    
    run_query(f"""
    INSERT INTO core.cat_macro_procesos (codigo_proceso, nombre_proceso, macro_proceso_padre, sistema_responsable, frecuencia_corte, descripcion)
    VALUES {', '.join(p_vals)}
    ON CONFLICT (codigo_proceso) DO UPDATE SET
        nombre_proceso = EXCLUDED.nombre_proceso,
        macro_proceso_padre = EXCLUDED.macro_proceso_padre,
        sistema_responsable = EXCLUDED.sistema_responsable,
        frecuencia_corte = EXCLUDED.frecuencia_corte,
        descripcion = EXCLUDED.descripcion;
    """)

    # 2. Populate Equipos SCEIN
    print("3. Populating cat_tipos_equipo_potencia...")
    wb_eq = openpyxl.load_workbook('docs/catalogos_maestros_spark/CATALOGO_MAESTRO_EQUIPOS_SCEIN.xlsx')
    ws_eq = wb_eq['TIPOS_EQUIPOS_SCEIN']
    eq_vals = []
    for r in range(3, ws_eq.max_row+1):
        id_eq = ws_eq.cell(row=r, column=1).value
        if not id_eq:
            continue
        cat = ws_eq.cell(row=r, column=2).value
        nom = ws_eq.cell(row=r, column=3).value
        sig = ws_eq.cell(row=r, column=4).value
        kvs = ws_eq.cell(row=r, column=5).value
        acc = ws_eq.cell(row=r, column=6).value
        cri = ws_eq.cell(row=r, column=7).value or 'ALTA'
        eq_vals.append(f"({escape_sql(id_eq)}, {escape_sql(cat)}, {escape_sql(nom)}, {escape_sql(sig)}, {escape_sql(kvs)}, {escape_sql(acc)}, {escape_sql(cri)})")

    if eq_vals:
        run_query(f"""
        INSERT INTO core.cat_tipos_equipo_potencia (id_tipo_equipo, categoria_equipo, nombre_elemento_estandar, siglas_componente, niveles_kv_tipicos, acciones_operativas_habituales, criticidad_defecto)
        VALUES {', '.join(eq_vals)}
        ON CONFLICT (id_tipo_equipo) DO UPDATE SET
            categoria_equipo = EXCLUDED.categoria_equipo,
            nombre_elemento_estandar = EXCLUDED.nombre_elemento_estandar,
            siglas_componente = EXCLUDED.siglas_componente,
            niveles_kv_tipicos = EXCLUDED.niveles_kv_tipicos,
            acciones_operativas_habituales = EXCLUDED.acciones_operativas_habituales,
            criticidad_defecto = EXCLUDED.criticidad_defecto;
        """)

    # 3. Populate Causas SCTIS
    print("4. Populating cat_causas_interrupcion & cat_subcausas_interrupcion...")
    wb_ca = openpyxl.load_workbook('docs/catalogos_maestros_spark/CATALOGO_MAESTRO_CAUSAS_HOMOLOGACION.xlsx')
    ws_c1 = wb_ca['01_CAUSAS_OFICIALES']
    c_vals = []
    for r in range(3, ws_c1.max_row+1):
        cod = ws_c1.cell(row=r, column=1).value
        if not cod:
            continue
        nom = ws_c1.cell(row=r, column=2).value
        desc = ws_c1.cell(row=r, column=3).value
        c_vals.append(f"({escape_sql(cod)}, {escape_sql(nom)}, {escape_sql(desc)})")

    if c_vals:
        run_query(f"""
        INSERT INTO core.cat_causas_interrupcion (codigo_causa, nombre_oficial_causa_sen, descripcion_operativa)
        VALUES {', '.join(c_vals)}
        ON CONFLICT (codigo_causa) DO UPDATE SET
            nombre_oficial_causa_sen = EXCLUDED.nombre_oficial_causa_sen,
            descripcion_operativa = EXCLUDED.descripcion_operativa;
        """)

    # Subcausas
    ws_c2 = wb_ca['02_SUB_CAUSAS']
    sc_vals = []
    for r in range(3, ws_c2.max_row+1):
        padre = ws_c2.cell(row=r, column=1).value
        cod_sub = ws_c2.cell(row=r, column=2).value
        if not cod_sub or not padre:
            continue
        nom_sub = ws_c2.cell(row=r, column=3).value
        desc_sub = ws_c2.cell(row=r, column=4).value
        sc_vals.append(f"({escape_sql(cod_sub)}, {escape_sql(padre)}, {escape_sql(nom_sub)}, {escape_sql(desc_sub)})")

    if sc_vals:
        run_query(f"""
        INSERT INTO core.cat_subcausas_interrupcion (codigo_sub_causa, codigo_causa_padre, nombre_sub_causa_oficial, descripcion)
        VALUES {', '.join(sc_vals)}
        ON CONFLICT (codigo_sub_causa) DO UPDATE SET
            codigo_causa_padre = EXCLUDED.codigo_causa_padre,
            nombre_sub_causa_oficial = EXCLUDED.nombre_sub_causa_oficial,
            descripcion = EXCLUDED.descripcion;
        """)

    # 4. Populate Familias de Materiales
    print("5. Populating cat_familias_materiales...")
    with open('apps/corpoelec-sigi-gestion-planificacion-distribucion/src/data/masterCatalogsLegacy.json') as f:
        legacy_data = json.load(f)

    fam_list = legacy_data['materiales']['familias']
    fam_vals = []
    for fam in fam_list:
        code_fam = fam.replace(' ', '_').replace('(', '').replace(')', '').replace('/', '_').upper()
        fam_vals.append(f"({escape_sql(code_fam)}, {escape_sql(fam)}, {escape_sql('Familia de insumos ' + fam)})")

    run_query(f"""
    INSERT INTO core.cat_familias_materiales (codigo_familia, nombre_familia, descripcion)
    VALUES {', '.join(fam_vals)}
    ON CONFLICT (codigo_familia) DO UPDATE SET
        nombre_familia = EXCLUDED.nombre_familia,
        descripcion = EXCLUDED.descripcion;
    """)

    # 5. Populate Items Materiales y Precios (804 items)
    print("6. Populating cat_items_materiales_precios (804 items)...")
    precios_list = legacy_data['precios']
    batch_size = 100
    for i in range(0, len(precios_list), batch_size):
        chunk = precios_list[i:i+batch_size]
        item_vals = []
        for it in chunk:
            desc = it.get('descripcion', '').strip()
            und = it.get('unidad', 'UN').strip().upper()
            pr = float(it.get('precio_eur', 0.0))
            item_vals.append(f"({escape_sql(desc)}, {escape_sql(und)}, {pr}, 'EUR')")
        
        run_query(f"""
        INSERT INTO core.cat_items_materiales_precios (descripcion_item, unidad_medida, precio_referencia_eur, moneda)
        VALUES {', '.join(item_vals)};
        """)
        print(f"  Inserted material prices up to {min(i+batch_size, len(precios_list))}/{len(precios_list)}")

    # 6. Populate Tipos de Restricción Operativa & Estados Operativos
    print("7. Populating cat_tipos_restriccion_operativa & cat_estados_operativos_activo...")
    restricciones = [
        ('AISLAMIENTO', 'Restricción por Aislamiento Dañado', 'Pérdida de rigidez dieléctrica o rotura de aisladores'),
        ('CONDENSADOR', 'Restricción por Banco de Condensadores', 'Indisponibilidad de compensación reactiva'),
        ('CONECTORES', 'Restricción por Conectores / Puntos Calientes', 'Conexiones sulfatadas o sobrecalentadas'),
        ('SECCIONAMIENTO', 'Restricción por Equipos de Seccionamiento', 'Cortacorrientes o seccionadores inoperables'),
        ('ESTRUCTURA', 'Restricción por Estructuras / Postes', 'Postes chocados, socavados o corroídos'),
        ('CONDUCTOR', 'Restricción por Conductor Subdimensionado / Dañado', 'Calibre insuficiente para la demanda o hebras cortadas'),
        ('VEGETACION', 'Restricción por Vegetación / Franja de Servidumbre', 'Árboles y ramas invadiendo servidumbre de paso'),
        ('TRANSFORMADOR', 'Restricción por Transformador de Distribución', 'Sobrecarga térmica o fuga de aceite dieléctrico')
    ]
    r_vals = [f"({escape_sql(c)}, {escape_sql(n)}, {escape_sql(d)})" for c, n, d in restricciones]
    run_query(f"""
    INSERT INTO core.cat_tipos_restriccion_operativa (codigo_restriccion, nombre_restriccion, descripcion)
    VALUES {', '.join(r_vals)}
    ON CONFLICT (codigo_restriccion) DO UPDATE SET
        nombre_restriccion = EXCLUDED.nombre_restriccion,
        descripcion = EXCLUDED.descripcion;
    """)

    estados_op = [
        ('OPERATIVO', 'Operativo / En Servicio', 'Activo en condiciones nominales de funcionamiento'),
        ('INDISPONIBLE_AVERIA', 'Indisponible por Avería', 'Activo fuera de servicio por falla física o eléctrica'),
        ('EN_MANTENIMIENTO', 'En Mantenimiento', 'Activo sometido a intervención preventiva o correctiva'),
        ('FUERA_DE_SERVICIO', 'Fuera de Servicio Programado', 'Activo desenergizado por consignación o maniobra operativa'),
        ('RESERVA_FRIA', 'Reserva Fría', 'Equipo desconectado disponible para entrada diferida'),
        ('RESERVA_CALIENTE', 'Reserva Caliente', 'Equipo energizado sin carga listo para transferencia inmediata'),
        ('DESINCORPORADO', 'Desincorporado / Baja Técnica', 'Activo retirado definitivamente del sistema')
    ]
    e_vals = [f"({escape_sql(c)}, {escape_sql(n)}, {escape_sql(d)})" for c, n, d in estados_op]
    run_query(f"""
    INSERT INTO core.cat_estados_operativos_activo (codigo_estado_operativo, nombre_estado, descripcion)
    VALUES {', '.join(e_vals)}
    ON CONFLICT (codigo_estado_operativo) DO UPDATE SET
        nombre_estado = EXCLUDED.nombre_estado,
        descripcion = EXCLUDED.descripcion;
    """)

    # 7. Create Views in Public & Enable RLS
    print("8. Creating Views in public schema and enabling RLS...")
    v_sql = """
    CREATE OR REPLACE VIEW public.cat_macro_procesos AS SELECT * FROM core.cat_macro_procesos;
    CREATE OR REPLACE VIEW public.cat_tipos_equipo_potencia AS SELECT * FROM core.cat_tipos_equipo_potencia;
    CREATE OR REPLACE VIEW public.cat_causas_interrupcion AS SELECT * FROM core.cat_causas_interrupcion;
    CREATE OR REPLACE VIEW public.cat_subcausas_interrupcion AS SELECT * FROM core.cat_subcausas_interrupcion;
    CREATE OR REPLACE VIEW public.cat_familias_materiales AS SELECT * FROM core.cat_familias_materiales;
    CREATE OR REPLACE VIEW public.cat_items_materiales_precios AS SELECT * FROM core.cat_items_materiales_precios;
    CREATE OR REPLACE VIEW public.cat_tipos_restriccion_operativa AS SELECT * FROM core.cat_tipos_restriccion_operativa;
    CREATE OR REPLACE VIEW public.cat_estados_operativos_activo AS SELECT * FROM core.cat_estados_operativos_activo;

    ALTER TABLE core.cat_macro_procesos ENABLE ROW LEVEL SECURITY;
    ALTER TABLE core.cat_tipos_equipo_potencia ENABLE ROW LEVEL SECURITY;
    ALTER TABLE core.cat_causas_interrupcion ENABLE ROW LEVEL SECURITY;
    ALTER TABLE core.cat_subcausas_interrupcion ENABLE ROW LEVEL SECURITY;
    ALTER TABLE core.cat_familias_materiales ENABLE ROW LEVEL SECURITY;
    ALTER TABLE core.cat_items_materiales_precios ENABLE ROW LEVEL SECURITY;
    ALTER TABLE core.cat_tipos_restriccion_operativa ENABLE ROW LEVEL SECURITY;
    ALTER TABLE core.cat_estados_operativos_activo ENABLE ROW LEVEL SECURITY;

    DROP POLICY IF EXISTS p_sel_mp ON core.cat_macro_procesos;
    CREATE POLICY p_sel_mp ON core.cat_macro_procesos FOR SELECT TO authenticated, anon USING (true);

    DROP POLICY IF EXISTS p_sel_eq ON core.cat_tipos_equipo_potencia;
    CREATE POLICY p_sel_eq ON core.cat_tipos_equipo_potencia FOR SELECT TO authenticated, anon USING (true);

    DROP POLICY IF EXISTS p_sel_ca ON core.cat_causas_interrupcion;
    CREATE POLICY p_sel_ca ON core.cat_causas_interrupcion FOR SELECT TO authenticated, anon USING (true);

    DROP POLICY IF EXISTS p_sel_sca ON core.cat_subcausas_interrupcion;
    CREATE POLICY p_sel_sca ON core.cat_subcausas_interrupcion FOR SELECT TO authenticated, anon USING (true);

    DROP POLICY IF EXISTS p_sel_fm ON core.cat_familias_materiales;
    CREATE POLICY p_sel_fm ON core.cat_familias_materiales FOR SELECT TO authenticated, anon USING (true);

    DROP POLICY IF EXISTS p_sel_it ON core.cat_items_materiales_precios;
    CREATE POLICY p_sel_it ON core.cat_items_materiales_precios FOR SELECT TO authenticated, anon USING (true);

    DROP POLICY IF EXISTS p_sel_ro ON core.cat_tipos_restriccion_operativa;
    CREATE POLICY p_sel_ro ON core.cat_tipos_restriccion_operativa FOR SELECT TO authenticated, anon USING (true);

    DROP POLICY IF EXISTS p_sel_eo ON core.cat_estados_operativos_activo;
    CREATE POLICY p_sel_eo ON core.cat_estados_operativos_activo FOR SELECT TO authenticated, anon USING (true);
    """
    run_query(v_sql)
    print("ALL MASTER CATALOGS DEPLOYED SUCCESSFULLY!")

if __name__ == '__main__':
    main()
